"""
regeneration_lecture.py — Brique 1 de la "régénération partielle" (4e bloc à
venir dans l'app).

Rôle de cette brique : NE RIEN CALCULER. Juste lire un planning déjà rempli
(et modifié à la main) et préparer, pour UNE semaine et UN OU PLUSIEURS
jours de cette semaine, tout ce dont la brique 2 (recalcul CP-SAT) aura
besoin :

1. Les règles (mêmes onglets de préparation que le Bloc 3, même lecture —
   Planning_type/Paramètres/Affectations visibles, le reste en _prep_*).
2. Les "contraintes figées" du/des jour(s) à régénérer : tout ce qui est
   déjà noté en Accueil/Réunion/Absence (H/I/J) + notes agents (W-Z) —
   converties au même format que les "événements" que le moteur CP-SAT
   sait déjà utiliser (parse_evenements), pour brancher directement dessus
   sans rien réinventer. Les colonnes B à G (qui travaille où) de ces
   jours-là ne sont PAS lues comme contraintes : elles seront effacées et
   recalculées par la brique 2.
3. Les heures déjà travaillées cette semaine par chaque agent, sur les
   jours qui NE sont PAS régénérés (pour que la brique 2 puisse démarrer
   son compteur d'équité hebdomadaire au bon endroit, plutôt qu'à zéro).
4. Les conflits déjà présents dans le/les jour(s) à régénérer (ex. un
   agent noté absent ET en réunion au même moment) — réutilise telle
   quelle la détection du Bloc 3 (planning_checker.py), restreinte à cette
   zone. Ces conflits doivent être réglés par la personne AVANT de relancer
   le calcul, sinon la brique 2 recevrait une contrainte impossible à
   respecter (elle ne peut pas deviner laquelle des deux notes est la
   bonne).

Rien n'est écrit dans un fichier Excel ici — uniquement de la lecture, pour
qu'on puisse la valider avant de toucher au moteur de calcul.
"""

from collections import defaultdict
import re

from planning_engine_cpsat import (
    parse_parametres, parse_affectations, parse_horaires_agents,
    parse_roulement_samedi, parse_planning_type, parse_besoins_jeunesse,
    parse_jours_speciaux,
)
from planning_checker import (
    lire_jours_semaine, construire_occurrences_jour, fusionner_occurrences,
    charger_donnees_preparation, JOUR_CAPITALISE,
    est_ignore, est_eloise, ALL_AGENTS_CONNUS,
)

import openpyxl
from io import BytesIO


# Types d'occurrence qui viennent des colonnes B à G (affectation de service
# public) — celles-là sont EFFACÉES pour le(s) jour(s) à régénérer, jamais
# gardées comme contrainte.
TYPES_AFFECTATION = {'RDC', 'Adulte', 'M & F', 'Jeunesse'}
# Types qui viennent de H/I/J + notes W-Z — celles-là restent, elles
# deviennent des contraintes figées pour la brique 2.
TYPES_EVENEMENT = {'Accueil/Animation', 'Réunion', 'Absence'}


from dataclasses import dataclass


def _nom_propre(segment):
    """Dérive un nom d'événement 'propre' (sans prénoms ni horaire) à partir
    du texte brut d'un segment, ex. '- rdv médical (Macha)' -> 'rdv médical',
    'Reunion pôle (Anne-Françoise, ..., 10h-11h30)' -> 'Reunion pôle'.
    Utilisé pour alimenter la vue par agent (generer_vue_agent), qui attend
    un nom déjà nettoyé — contrairement aux messages d'alerte, qui eux
    utilisent le texte complet ('nom_affichage') pour rester précis."""
    s = (segment or '').strip()
    if s.startswith('- '):
        s = s[2:].strip()
    s = re.sub(r'\s*\([^)]*\)\s*$', '', s).strip()
    return s or (segment or '').strip()


@dataclass
class ConflitFixe:
    """Un chevauchement entre deux éléments FIXES (Accueil/Réunion/Absence
    ou notes W-Z) pour un même agent, qui restera vrai même après
    régénération — donc à signaler visuellement sur le planning final,
    pas à corriger automatiquement (le programme ne peut pas savoir lequel
    des deux événements est le bon)."""
    agent: str
    jour: str
    date: str
    evenement_1: dict
    evenement_2: dict


class ErreurRegeneration(Exception):
    """Erreur bloquante détectée pendant la lecture (semaine/jour introuvable,
    onglets de préparation absents, etc.) — à afficher telle quelle à
    l'utilisatrice dans l'app, pas une erreur technique à décoder."""
    pass


def lire_planning_pour_regeneration(file_bytes, semaine_num, jours_a_regenerer):
    """
    Point d'entrée de la brique 1.

    - file_bytes : contenu du fichier planning du mois déjà rempli (bytes).
    - semaine_num : numéro de la semaine à traiter (1, 2, 3...).
    - jours_a_regenerer : liste de jours à régénérer au sein de cette
      semaine, ex. ['Mercredi'] ou ['Mercredi', 'Jeudi']. Les noms de jour
      sont insensibles à la casse ('mercredi' fonctionne aussi).

    Retourne un dict prêt à être consommé par la brique 2 :
    {
        'semaine_num': int,
        'jours_regeneres': [...],      # jours confirmés, orthographe canonique
        'jours_fixes': [...],          # les autres jours de cette même semaine
        'prep': {...},                 # données de préparation (règles)
        'evenements_regeneres': [...], # au format parse_evenements() du moteur
        'heures_deja_semaine': {agent: minutes},  # sur les jours fixes
        'conflits': [Anomalie],        # zone régénérée uniquement
        'jours_data_bruts': {...},     # bloc complet de chaque jour (pour la brique 2)
    }

    Ne modifie rien, ne recalcule rien — lecture seule.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    nom_onglet = f'Semaine_{semaine_num}'
    if nom_onglet not in wb.sheetnames:
        raise ErreurRegeneration(
            f"Je ne trouve pas d'onglet '{nom_onglet}' dans ce fichier. "
            f"Onglets présents : {', '.join(wb.sheetnames)}."
        )
    ws = wb[nom_onglet]

    # ── 1. Règles (mêmes onglets cachés que le Bloc 3) ─────────────────
    prep = charger_donnees_preparation(wb)
    if prep is None:
        raise ErreurRegeneration(
            "Ce fichier ne contient pas les onglets de préparation "
            "(Paramètres, Affectations, _prep_..., etc.). Il a probablement "
            "été généré avec une ancienne version de l'outil, ou ces "
            "onglets ont été supprimés. La régénération partielle a besoin "
            "de ces règles pour fonctionner — impossible de continuer avec "
            "ce fichier."
        )
    if 'erreur_lecture' in prep:
        raise ErreurRegeneration(
            f"Les onglets de préparation sont présents mais illisibles "
            f"({prep['erreur_lecture']}). Impossible de continuer en toute "
            f"sécurité."
        )
    manquants = prep.get('manquants', [])
    if manquants:
        raise ErreurRegeneration(
            f"Onglet(s) de préparation manquant(s) dans ce fichier : "
            f"{', '.join(manquants)}. La régénération a besoin de toutes "
            f"les règles pour ne pas proposer un planning qui les viole."
        )

    # ── 2. Découpage de la semaine en jours ─────────────────────────────
    jours_data = lire_jours_semaine(ws)
    if not jours_data:
        raise ErreurRegeneration(f"Aucun jour reconnu dans l'onglet '{nom_onglet}'.")

    jours_dispo = {j['jour']: j for j in jours_data}  # clé = 'MARDI', 'MERCREDI'...

    cibles_norm = [j.strip().upper() for j in jours_a_regenerer]
    introuvables = [j for j in cibles_norm if j not in jours_dispo]
    if introuvables:
        raise ErreurRegeneration(
            f"Jour(s) demandé(s) introuvable(s) dans '{nom_onglet}' : "
            f"{', '.join(introuvables)}. Jours présents cette semaine-là : "
            f"{', '.join(jours_dispo.keys())}."
        )

    jours_regeneres = cibles_norm
    jours_fixes = [j for j in jours_dispo if j not in cibles_norm]

    # ── 3. Contraintes figées (Accueil/Réunion/Absence) ──────────────────
    # On lit H/I/J + notes W-Z (déjà combinées par les formules du fichier)
    # et on ne garde QUE les occurrences de type Accueil/Animation, Réunion,
    # Absence — jamais les affectations B-G (RDC/Adulte/M&F/Jeunesse).
    #
    # Fait pour TOUS les jours de la semaine (régénérés ET fixes) :
    # - pour les jours à régénérer, ces événements servent de CONTRAINTES
    #   FIGÉES au moteur de calcul (brique 2) → 'evenements_regeneres' ;
    # - pour les jours fixes, ils servent seulement à reconstruire
    #   correctement l'onglet "vue par agent" (brique 3), qui couvre
    #   toujours la semaine entière → inclus en plus dans
    #   'evenements_tous_jours'. Corrige la limite du 22/08 : avant, cet
    #   onglet n'était reconstruit (donc le lien direct avec les notes
    #   W-Z restait actif) que si TOUTE la semaine était régénérée.
    def _extraire_evenements_jour(jour_key):
        jour_data = jours_dispo[jour_key]
        date_str = jour_data.get('date_str')
        evs = []
        occ_brutes = construire_occurrences_jour(jour_data, ALL_AGENTS_CONNUS)
        for agent, liste in occ_brutes.items():
            if est_ignore(agent) or est_eloise(agent):
                # Eloïse : jamais une contrainte agent (cf. consigne du 20/08 —
                # sa présence dans un texte d'événement est ignorée partout).
                continue
            for occ in fusionner_occurrences(liste):
                if occ['type'] not in TYPES_EVENEMENT:
                    continue  # occurrence B-G : pas une contrainte, sera effacée
                # Convention du moteur normal (parse_evenements) : un congé
                # porte le nom EXACT 'congé' (mot nu), pas le texte détaillé
                # — sinon la vue par agent ne le reconnaît pas comme tel et
                # l'affiche comme un événement quelconque au lieu du gris
                # "Congé" attendu (bug découvert le 20/08 sur Stéphane).
                # Pour les autres événements, 'nom' doit être PROPRE (sans
                # prénoms ni horaire) — c'est ce qu'attend la vue par agent
                # (label_evenement_sans_noms) pour son affichage "sans nom" ;
                # le texte complet reste disponible via 'nom_affichage' pour
                # les messages d'alerte, qui eux doivent rester précis.
                detail_norm = occ['detail'].strip().lower()
                if detail_norm.startswith('congé') or detail_norm.startswith('conge'):
                    nom = 'congé'
                else:
                    nom = _nom_propre(occ['detail'])
                evs.append({
                    'date': date_str,
                    'cs': occ['debut'],
                    'ce': occ['fin'],
                    'nom': nom,
                    'nom_affichage': occ['detail'],  # texte complet, pour les messages/alertes
                    'type': occ['type'],
                    'agents': [agent],
                })
        return evs

    evenements_regeneres = []
    for jour_key in jours_regeneres:
        evenements_regeneres.extend(_extraire_evenements_jour(jour_key))

    evenements_fixes = []
    for jour_key in jours_fixes:
        evenements_fixes.extend(_extraire_evenements_jour(jour_key))

    evenements_tous_jours = evenements_regeneres + evenements_fixes

    # ── 4. Heures déjà travaillées cette semaine, sur les jours FIXES ───
    # Sert à la brique 2 pour démarrer son compteur d'équité hebdomadaire
    # au bon endroit plutôt qu'à zéro (le moteur calcule l'équité jour après
    # jour, remise à zéro chaque nouvelle semaine — cf. planning_engine_cpsat).
    # ⚠️ Limite connue (déjà signalée) : si le jour régénéré n'est pas le
    # dernier de la semaine, les jours fixes APRÈS lui ne seront pas
    # rééquilibrés en retour — seul le compteur AVANT est correctement pris
    # en compte ici.
    heures_deja_semaine = defaultdict(int)
    for jour_key in jours_fixes:
        jour_data = jours_dispo[jour_key]
        occ_brutes = construire_occurrences_jour(jour_data, ALL_AGENTS_CONNUS)
        for agent, liste in occ_brutes.items():
            if est_ignore(agent) or est_eloise(agent):
                continue
            for occ in fusionner_occurrences(liste):
                if occ['type'] in TYPES_AFFECTATION:
                    heures_deja_semaine[agent] += (occ['fin'] - occ['debut'])

    # ── 5. Conflits qui RESTERONT vrais après régénération ──────────────
    # Important : on ne réutilise PAS ici la vérification complète du
    # Bloc 3 (verifier_jour), qui mélange des règles portant sur les
    # affectations B-G (RDC/Adulte/M&F/Jeunesse) — celles-là seront
    # effacées et recalculées, donc un conflit qui les implique n'aura
    # plus de sens une fois le nouveau planning généré (ex. "Stéphane en
    # congé mais affecté en M&F" disparaît de lui-même puisque le solveur
    # ne pourra plus le proposer en M&F). Seuls les chevauchements ENTRE
    # DEUX ÉLÉMENTS FIXES (Accueil/Réunion/Absence, ou notes W-Z) restent
    # vrais quoi qu'il arrive — ce sont ceux-là, et ceux-là seulement,
    # qu'on doit signaler visuellement sur le planning final (demande du
    # 20/08 : le cas Anne-Françoise doit apparaître avec une alerte, pas
    # bloquer ni disparaître).
    conflits = []
    evenements_par_agent_jour = defaultdict(list)
    for ev in evenements_regeneres:
        for agent in ev['agents']:
            evenements_par_agent_jour[(agent, ev['date'])].append(ev)
    for (agent, date_str), evs in evenements_par_agent_jour.items():
        for i in range(len(evs)):
            for j in range(i + 1, len(evs)):
                a, b = evs[i], evs[j]
                if a is b:
                    continue
                if a['nom_affichage'] == b['nom_affichage'] and a['cs'] == b['cs'] and a['ce'] == b['ce']:
                    continue  # même événement compté deux fois, pas un conflit
                if a['cs'] < b['ce'] and b['cs'] < a['ce']:
                    jour_aff = next((jd['jour'] for jd in jours_dispo.values()
                                      if jd.get('date_str') == date_str), '?')
                    conflits.append(ConflitFixe(
                        agent=agent, jour=jour_aff, date=date_str,
                        evenement_1=a, evenement_2=b,
                    ))

    return {
        'semaine_num': semaine_num,
        'jours_regeneres': jours_regeneres,
        'jours_fixes': jours_fixes,
        'prep': prep,
        'evenements_regeneres': evenements_regeneres,
        'evenements_tous_jours': evenements_tous_jours,
        'heures_deja_semaine': dict(heures_deja_semaine),
        'conflits': conflits,
        'jours_data_bruts': jours_dispo,
    }


def resumer_lecture(resultat):
    """Petit résumé lisible (texte simple) de ce que la brique 1 a trouvé —
    pratique pour l'affichage Streamlit et pour le débogage."""
    lignes = []
    lignes.append(f"Semaine {resultat['semaine_num']} : "
                   f"{', '.join(resultat['jours_regeneres'])} à régénérer, "
                   f"{', '.join(resultat['jours_fixes'])} conservé(s) tel(s) quel(s).")
    lignes.append(f"{len(resultat['evenements_regeneres'])} contrainte(s) figée(s) "
                   f"(Accueil/Réunion/Absence) trouvée(s) sur le(s) jour(s) à régénérer.")
    if resultat['heures_deja_semaine']:
        total_h = ', '.join(
            f"{a} : {m/60:.1f}h" for a, m in sorted(resultat['heures_deja_semaine'].items())
        )
        lignes.append(f"Heures déjà travaillées cette semaine (jours fixes) : {total_h}")
    if resultat['conflits']:
        lignes.append(f"⚠️ {len(resultat['conflits'])} conflit(s) entre éléments fixes "
                       f"détecté(s) — resteront visibles (alerte) sur le planning régénéré :")
        for c in resultat['conflits']:
            lignes.append(f"  {c.jour} — {c.agent} : "
                           f"« {c.evenement_1['nom_affichage']} » ({c.evenement_1['cs']//60}h{c.evenement_1['cs']%60:02d}-"
                           f"{c.evenement_1['ce']//60}h{c.evenement_1['ce']%60:02d}) "
                           f"⟷ « {c.evenement_2['nom_affichage']} » ({c.evenement_2['cs']//60}h{c.evenement_2['cs']%60:02d}-"
                           f"{c.evenement_2['ce']//60}h{c.evenement_2['ce']%60:02d})")
    else:
        lignes.append("Aucun conflit entre éléments fixes détecté.")
    return '\n'.join(lignes)

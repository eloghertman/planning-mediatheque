# -*- coding: utf-8 -*-
"""
sources_to_evenements.py
=========================

Construit l'onglet "Événements" (format attendu par planning_engine.py /
parse_evenements) à partir des fichiers sources bruts que la médiathèque tient
déjà, plutôt que de les recopier à la main.

Fichiers sources gérés (chacun optionnel — on ne fournit que ceux qu'on a) :

  - Congé équipe            (parse_conges)
  - Accueil libre crèche    (parse_accueil_creche)
  - Accueil de classe       (parse_accueil_classe)
  - Lecture du jeudi matin  (parse_lecture_jeudi_matin)
  - Calendrier des événements déjà saisi, si on veut le réinjecter tel quel
    (parse_calendrier_evenements) — passthrough, aucune interprétation.

Chaque parseur retourne une liste d'"événements" (dict) au format commun :

    {
        'date':  datetime.date,
        'debut': str ou None   # "10h", "14h30"...
        'fin':   str ou None
        'nom':   str            # nom de l'événement
        'agents': [str, ...]    # noms d'agents déjà normalisés
        'alert': bool           # True → à surligner en jaune + commentaire
        'alert_reason': str ou None
    }

`build_onglet_evenements(events, out_path)` écrit ensuite un classeur Excel
avec un onglet "Événements" dans le format texte attendu par
`planning_engine.parse_evenements` (Date FR texte | Début | Fin | Nom | Agents).
Les cases incomplètes (alert=True) sont surlignées en jaune avec un
commentaire expliquant pourquoi, pour qu'Elo puisse compléter d'un coup d'œil.

RÈGLES VALIDÉES AVEC L'UTILISATRICE (session du 14/08) :
  - Fichier congés : n'importe quelle lettre (C, CS, M, récup...) = un congé
    journée complète (9h-19h). Une valeur numérique < 1 (ex: 0.5) = demi-
    journée dont on ignore si c'est le matin ou l'après-midi → case Début/Fin
    laissée vide + surlignée en jaune, à compléter par Elo.
  - Eloïse (la directrice) n'est jamais incluse dans les agents d'un
    événement — elle n'est pas planifiée automatiquement (règle générale du
    projet, cf. contexte).
  - Lydie a quitté l'équipe — toujours ignorée si elle apparaît dans un
    fichier source.
  - Accueil de classe : si aucun intervenant n'est indiqué dans le fichier
    source alors que l'événement existe (école renseignée), la case Agents
    est laissée vide + surlignée en jaune.
  - Référence en cas de désaccord : les fichiers sources bruts font foi, pas
    un ancien onglet Événements rempli à la main (peut contenir des ajouts
    faits "de tête" par Elo, non traçables plusieurs mois après).
"""

import re
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment

# ══════════════════════════════════════════════════════════════
#  RÉFÉRENTIEL COMMUN
# ══════════════════════════════════════════════════════════════

JOURS_FR = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi', 'dimanche']
MOIS_FR = {1: 'janvier', 2: 'février', 3: 'mars', 4: 'avril', 5: 'mai', 6: 'juin',
           7: 'juillet', 8: 'août', 9: 'septembre', 10: 'octobre', 11: 'novembre', 12: 'décembre'}
MOIS_FR_CAP = {k: v.capitalize() for k, v in MOIS_FR.items()}  # noms d'onglets ('Mai', 'Septembre'...)

# Agents réguliers connus par le projet (cf. contexte). Lydie et Eloïse sont
# volontairement EXCLUS de cette liste de "planification automatique" :
# Lydie a quitté, Eloïse n'est jamais planifiée automatiquement.
AGENTS_CONNUS = [
    'Marie-France', 'Anne-Françoise', 'Christine', 'Léa', 'Chloé', 'Macha',
    'Delphine', 'Barbara', 'Stéphane', 'Stéphanie', 'Robin', 'Guillaume',
    'Agnès', 'Tiphaine',
]
AGENTS_A_IGNORER = {'lydie', 'eloïse', 'eloise'}  # jamais retenus comme agent planifié

INITIALES_ACCUEIL_CLASSE = {
    'SD': 'Stéphanie', 'TV': 'Tiphaine', 'GC': 'Guillaume',
    'RL': 'Robin', 'BP': 'Barbara', 'DR': 'Delphine', 'EG': 'Eloïse',
}

JAUNE = PatternFill('solid', fgColor='FFFFFF00')


def date_fr(d: date) -> str:
    """'jeudi 7 mai 2026' — format texte attendu par parse_date_flexible."""
    return f"{JOURS_FR[d.weekday()]} {d.day} {MOIS_FR[d.month]} {d.year}"


def normalize_agent(name):
    """Renvoie le nom canonique si c'est un agent connu, sinon None.
    Insensible à la casse/accents approximatifs simples. Exclut explicitement
    Lydie et Eloïse (cf. AGENTS_A_IGNORER)."""
    if not name:
        return None
    key = str(name).strip().lower()
    if key in AGENTS_A_IGNORER:
        return None
    for a in AGENTS_CONNUS:
        if a.lower() == key:
            return a
    return None


def detect_agents_in_text(text):
    """Repère les prénoms d'agents connus dans un texte libre
    ('Agnès et Tiphaine', 'Agnès Stéphanie', 'Agnès, Robin'...)."""
    if not text:
        return []
    found = []
    txt = str(text)
    for a in AGENTS_CONNUS:
        if re.search(re.escape(a), txt, re.IGNORECASE):
            found.append(a)
    return found


def parse_heure_range(raw):
    """'10h-11h', '14H-15H', '10h-10h30' → ('10h', '11h'). Renvoie (None, None)
    si non interprétable (case vide, format inattendu)."""
    if not raw:
        return None, None
    s = str(raw).strip().lower()
    if '-' not in s:
        return None, None
    a, b = s.split('-', 1)
    a, b = a.strip(), b.strip()
    if not a or not b:
        return None, None
    return a, b


def _event(d, debut, fin, nom, agents, alert=False, alert_reason=None, source=None):
    return {
        'date': d, 'debut': debut, 'fin': fin, 'nom': nom,
        'agents': [a for a in agents if a],
        'alert': alert, 'alert_reason': alert_reason, 'source': source,
    }


# ══════════════════════════════════════════════════════════════
#  1. CONGÉS ÉQUIPE
# ══════════════════════════════════════════════════════════════

def parse_conges(path, mois, annee):
    """Onglet nommé d'après le mois ('Mai', 'Septembre'...). Ligne d'en-tête
    = celle où colonne B == 'Nom de l'employé' (avec apostrophe typographique
    ou droite) ; les numéros de jour sont sur cette même ligne, à partir de la
    colonne C. Une lettre = congé journée complète (9h-19h). Un nombre < 1 =
    demi-journée, heure inconnue → alerte."""
    wb = load_workbook(path, data_only=True)
    sheet_name = MOIS_FR_CAP.get(mois)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Onglet '{sheet_name}' introuvable dans {path} "
                          f"(onglets disponibles : {wb.sheetnames})")
    ws = wb[sheet_name]

    header_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v and 'nom de l' in str(v).lower():
            header_row = r
            break
    if header_row is None:
        raise ValueError(f"Ligne d'en-tête ('Nom de l'employé') introuvable dans {path} / {sheet_name}")

    # Colonnes de jours : de C jusqu'à ce que la cellule d'en-tête ne soit
    # plus un nombre de jour valide (s'arrête à la colonne 'Total des jours').
    day_cols = {}
    for c in range(3, 40):
        v = ws.cell(row=header_row, column=c).value
        try:
            day_num = int(v)
            if 1 <= day_num <= 31:
                day_cols[c] = day_num
        except (TypeError, ValueError):
            continue

    conges_par_jour = {}   # date -> {'full': [agents], 'half': [agents]}
    r = header_row + 1
    while r <= ws.max_row:
        nom_agent = ws.cell(row=r, column=2).value
        if not nom_agent or not str(nom_agent).strip():
            r += 1
            continue
        agent = normalize_agent(nom_agent)
        if agent is None:
            # Lydie / Eloïse / nom non reconnu : on ignore silencieusement
            r += 1
            continue
        for c, day_num in day_cols.items():
            v = ws.cell(row=r, column=c).value
            if v is None or v == '':
                continue
            try:
                d = date(annee, mois, day_num)
            except ValueError:
                continue
            entry = conges_par_jour.setdefault(d, {'full': [], 'half': []})
            if isinstance(v, (int, float)):
                if v >= 1:
                    entry['full'].append(agent)
                else:
                    entry['half'].append(agent)
            else:
                # N'importe quelle lettre (C, CS, M, récup...) = congé journée complète
                entry['full'].append(agent)
        r += 1

    events = []
    for d, grp in sorted(conges_par_jour.items()):
        if grp['full']:
            events.append(_event(d, '9h', '19h', 'congé', sorted(set(grp['full'])),
                                  source='congés équipe'))
        for agent in sorted(set(grp['half'])):
            events.append(_event(
                d, None, None, 'congé (demi-journée)', [agent],
                alert=True,
                alert_reason="Demi-journée de congé : le fichier source ne précise pas "
                              "si c'est le matin ou l'après-midi — à compléter.",
                source='congés équipe',
            ))
    return events


# ══════════════════════════════════════════════════════════════
#  2. ACCUEIL CRÈCHES
# ══════════════════════════════════════════════════════════════

def parse_accueil_creche(path, mois, annee):
    """A1 = heure (ex: '10h -10h30'), colonne A = mois (rempli une seule fois
    par bloc), colonne B = jour ('jeudi 7' — cellule fusionnée sur plusieurs
    lignes quand plusieurs crèches viennent le même jour), colonne C = nom de
    la crèche (une ligne par crèche ; un jour peut avoir 0, 1 ou plusieurs
    crèches). Ce sont des visites en créneaux libres : AUCUN agent n'est
    jamais affecté à cet événement. On ne surligne donc jamais l'absence
    d'agent — c'est l'état normal. Seule une heure introuvable en A1
    déclenche une alerte."""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    heure_raw = ws['A1'].value
    debut, fin = parse_heure_range(heure_raw)
    heure_alert = (debut is None)

    events = []
    current_month_label = None
    current_day = None
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        if a:
            current_month_label = str(a).strip().lower()
            current_day = None
        if b:
            m = re.search(r'(\d+)', str(b))
            current_day = int(m.group(1)) if m else None
        if not c:
            continue  # pas de crèche indiquée sur cette ligne → pas d'accueil
        if current_month_label != MOIS_FR[mois] or current_day is None:
            continue
        try:
            d = date(annee, mois, current_day)
        except ValueError:
            continue
        events.append(_event(
            d, debut, fin, 'Accueil libre crèche', [],
            alert=heure_alert,
            alert_reason="Heure introuvable en cellule A1 du fichier source — à compléter."
            if heure_alert else None,
            source='accueil libre crèche',
        ))
    return events


# ══════════════════════════════════════════════════════════════
#  3. ACCUEIL DE CLASSE
# ══════════════════════════════════════════════════════════════

VISITE_LIBRE_MARQUEURS = ('visite libre',)  # texte(s) signalant une visite en accès libre


def _est_visite_libre(*valeurs):
    """True si l'une des cellules fournies contient une mention de type
    'uniquement en visite libre' (comparaison insensible à la casse)."""
    for v in valeurs:
        if v and any(m in str(v).strip().lower() for m in VISITE_LIBRE_MARQUEURS):
            return True
    return False


def _jour_si_mois_cible(texte_date, mois_cible):
    """Le fichier accueil de classe écrit la date en texte, sans année
    ('mardi 3 novembre', parfois 'jeudi 22 (vacances)'). On ne connaît pas
    l'année à partir du texte seul : on se contente de vérifier que le nom
    du mois demandé apparaît dans le texte, et on renvoie le numéro du jour
    si c'est le cas (sinon None — ligne d'un autre mois, ou pas une date)."""
    if not texte_date:
        return None
    s = str(texte_date).strip().lower()
    m = re.search(r'(\d{1,2})', s)
    if not m:
        return None
    if MOIS_FR[mois_cible] not in s:
        return None
    return int(m.group(1))


def parse_accueil_classe(path, mois, annee):
    """Colonne A = initiales de l'intervenant, colonne B = date en texte
    ('mardi 3 novembre' — remplie seulement sur la 1re ligne du jour, les
    lignes suivantes du même jour ont la colonne B vide), colonne D =
    créneau horaire ('10h-11h' / '14H-15H'), colonne E = nom de l'école
    (événement seulement si rempli). Le fichier empile plusieurs mois à la
    suite (blocs 'NOVEMBRE', 'DECEMBRE'...), d'où le filtrage par nom de
    mois dans le texte plutôt que par position dans le fichier.

    Cas particulier "visite libre" : si la colonne A, E, ou J porte la
    mention 'uniquement en visite libre' (ou une variante contenant 'visite
    libre'), aucun agent n'est affecté et aucune alerte n'est levée — c'est
    l'état normal pour ce type de visite. L'événement est alors nommé
    'Accueil libre école'. Sinon, comportement inchangé : l'agent vient des
    initiales en colonne A, et son absence déclenche une alerte."""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    events = []
    current_date = None
    for r in range(1, ws.max_row + 1):
        a_init = ws.cell(row=r, column=1).value
        b_date = ws.cell(row=r, column=2).value
        d_heure = ws.cell(row=r, column=4).value
        e_ecole = ws.cell(row=r, column=5).value
        j_note = ws.cell(row=r, column=10).value

        if b_date:
            jour = _jour_si_mois_cible(b_date, mois)
            if jour:
                try:
                    current_date = date(annee, mois, jour)
                except ValueError:
                    current_date = None
            else:
                current_date = None  # ligne d'un autre mois : on ne reporte pas sa date

        if not e_ecole or not str(e_ecole).strip() or current_date is None:
            continue  # pas d'école indiquée, ou pas dans le mois demandé

        debut, fin = parse_heure_range(d_heure)

        if _est_visite_libre(a_init, e_ecole, j_note):
            events.append(_event(
                current_date, debut, fin, 'Accueil libre école', [],
                alert=False,
                source='accueil de classe',
            ))
            continue

        agent = None
        if a_init:
            agent = INITIALES_ACCUEIL_CLASSE.get(str(a_init).strip().upper())
            agent = normalize_agent(agent) if agent else None

        alert = agent is None
        events.append(_event(
            current_date, debut, fin, 'Accueil classe', [agent] if agent else [],
            alert=alert,
            alert_reason="Intervenant non précisé dans le fichier source." if alert else None,
            source='accueil de classe',
        ))
    return events


# ══════════════════════════════════════════════════════════════
#  4. LECTURE DU JEUDI MATIN
# ══════════════════════════════════════════════════════════════

def parse_lecture_jeudi_matin(path, mois, annee):
    """A1 = heure fixe ('10h-10h30'). Une séance occupe un bloc de lignes
    fusionnées (une ligne par enfant en-dessous, non utilisée ici) : seule
    la 1re ligne du bloc porte une vraie date (colonne B, en tant que date
    Excel, année comprise) et les intervenant(e)s (colonne J, plusieurs noms
    séparés par ';'). Nom de l'événement : 'Lectures AssMat/AssPar'
    (validé avec Elo)."""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    heure_raw = ws['A1'].value
    debut, fin = parse_heure_range(heure_raw)

    events = []
    for r in range(2, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if not isinstance(b, datetime):
            continue
        if not (b.year == annee and b.month == mois):
            continue

        j_val = ws.cell(row=r, column=10).value
        if j_val:
            agents = [normalize_agent(a) or a.strip() for a in str(j_val).split(';') if a.strip()]
        else:
            agents = []

        alert = len(agents) == 0
        events.append(_event(
            b.date(), debut, fin, 'Lectures AssMat/AssPar', agents,
            alert=alert,
            alert_reason="Intervenant non renseigné dans le fichier source." if alert else None,
            source='lecture jeudi matin',
        ))
    return events


# ══════════════════════════════════════════════════════════════
#  5. CALENDRIER DES ÉVÉNEMENTS DÉJÀ SAISI (passthrough)
# ══════════════════════════════════════════════════════════════

def parse_calendrier_evenements(path, sheet_name):
    """Le fichier est déjà dans le format cible (Date | Début | Fin | Nom |
    Agents) — on le relit tel quel, sans interprétation, pour pouvoir le
    fusionner avec les autres sources sans rien perdre de ce qui a déjà été
    saisi à la main."""
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Onglet '{sheet_name}' introuvable dans {path} "
                          f"(onglets disponibles : {wb.sheetnames})")
    ws = wb[sheet_name]
    events = []
    for r in range(2, ws.max_row + 1):
        date_txt = ws.cell(row=r, column=1).value
        debut = ws.cell(row=r, column=2).value
        fin = ws.cell(row=r, column=3).value
        nom = ws.cell(row=r, column=4).value
        agents_raw = ws.cell(row=r, column=5).value
        if not date_txt or not nom:
            continue
        agents = [normalize_agent(a) or a.strip() for a in str(agents_raw).split(';')] if agents_raw else []
        events.append(_event(date_txt, debut, fin, str(nom).strip(), agents,
                              source='calendrier déjà saisi'))
    return events


# ══════════════════════════════════════════════════════════════
#  ÉCRITURE DE L'ONGLET ÉVÉNEMENTS
# ══════════════════════════════════════════════════════════════

def build_onglet_evenements(events, out_path, sheet_name='Événements'):
    """Écrit un classeur Excel avec un onglet Événements dans le format
    attendu par planning_engine.parse_evenements.

    Règle de surlignage (colonne Agents) : si la colonne Agents est vide ou
    contient un texte qui n'est manifestement pas un prénom ('à déterminer',
    '?', 'à définir'...), toute la ligne est surlignée en jaune avec un
    commentaire — SAUF si l'intitulé de l'événement contient 'libre'
    (Accueil libre crèche, Accueil libre école...), où l'absence d'agent est
    normale et n'est jamais signalée.

    Les autres cas incomplets détectés par les parseurs (ev['alert'], par
    exemple une heure introuvable) restent surlignés comme avant."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = ['Date', 'Début', 'Fin', 'Nom', 'Agents']
    ws.append(headers)
    for c in range(1, 6):
        ws.cell(row=1, column=c).font = Font(bold=True)

    # Tri chronologique pour une lecture confortable
    def _sort_key(ev):
        d = ev['date']
        if isinstance(d, date):
            return (d, ev.get('debut') or '')
        return (date.max, '')
    events_sorted = sorted(events, key=_sort_key)

    PLACEHOLDERS_AGENT = {
        '', '?', 'à déterminer', 'a déterminer', 'à determiner', 'a determiner',
        'à définir', 'a définir', 'à definir', 'a definir',
    }

    row = 2
    n_alerts = 0
    for ev in events_sorted:
        d = ev['date']
        date_str = date_fr(d) if isinstance(d, date) else str(d)
        agents_str = ';'.join(ev['agents'])
        ws.cell(row=row, column=1, value=date_str)
        ws.cell(row=row, column=2, value=ev['debut'] or '')
        ws.cell(row=row, column=3, value=ev['fin'] or '')
        ws.cell(row=row, column=4, value=ev['nom'])
        ws.cell(row=row, column=5, value=agents_str)

        nom_low = (ev['nom'] or '').lower()
        est_libre = 'libre' in nom_low
        agent_placeholder = agents_str.strip().lower() in PLACEHOLDERS_AGENT

        if agent_placeholder and not est_libre:
            n_alerts += 1
            reason = (ev.get('alert_reason') if ev.get('alert') else None) or (
                "Agent non renseigné (vide, « à déterminer », « ? » ou "
                "« à définir ») — à compléter."
            )
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = JAUNE
            ws.cell(row=row, column=5).comment = Comment(reason, "Assistant planning")
        elif ev.get('alert'):
            n_alerts += 1
            reason = ev.get('alert_reason') or 'À compléter.'
            # Heure manquante -> on surligne Début/Fin
            cells_to_flag = []
            if not ev['debut'] and not ev['fin']:
                cells_to_flag = [2, 3]
            if not cells_to_flag:
                cells_to_flag = [5]
            for col in cells_to_flag:
                cell = ws.cell(row=row, column=col)
                cell.fill = JAUNE
                cell.comment = Comment(reason, "Assistant planning")
        row += 1

    for col, width in zip('ABCDE', [26, 10, 10, 24, 40]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A2'

    wb.save(out_path)
    return {'total': len(events_sorted), 'alerts': n_alerts}


# ══════════════════════════════════════════════════════════════
#  ORCHESTRATION
# ══════════════════════════════════════════════════════════════

def generate_evenements(mois, annee, out_path, sources=None):
    """sources: dict optionnel avec les clés suivantes (toutes facultatives) :
        'conges':            chemin du fichier congés équipe
        'accueil_creche':    chemin, ou liste de chemins (plusieurs années scolaires)
        'accueil_classe':    chemin, ou liste de chemins
        'lecture_jeudi':      chemin, ou liste de chemins
        'calendrier': (chemin, nom_onglet)  -> réinjecté tel quel (passthrough)
    Renvoie (events, stats) où stats = {'total':…, 'alerts':…, 'par_source': {...}}.
    """
    sources = sources or {}
    all_events = []
    par_source = {}

    def _as_list(v):
        if v is None:
            return []
        return v if isinstance(v, list) else [v]

    if sources.get('conges'):
        evs = parse_conges(sources['conges'], mois, annee)
        all_events += evs
        par_source['congés équipe'] = len(evs)

    for path in _as_list(sources.get('accueil_creche')):
        evs = parse_accueil_creche(path, mois, annee)
        all_events += evs
        par_source['accueil libre crèche'] = par_source.get('accueil libre crèche', 0) + len(evs)

    for path in _as_list(sources.get('accueil_classe')):
        evs = parse_accueil_classe(path, mois, annee)
        all_events += evs
        par_source['accueil de classe'] = par_source.get('accueil de classe', 0) + len(evs)

    for path in _as_list(sources.get('lecture_jeudi')):
        evs = parse_lecture_jeudi_matin(path, mois, annee)
        all_events += evs
        par_source['lecture jeudi matin'] = par_source.get('lecture jeudi matin', 0) + len(evs)

    if sources.get('calendrier'):
        path, sheet_name = sources['calendrier']
        evs = parse_calendrier_evenements(path, sheet_name)
        all_events += evs
        par_source['calendrier déjà saisi'] = len(evs)

    write_stats = build_onglet_evenements(all_events, out_path)
    stats = {**write_stats, 'par_source': par_source}
    return all_events, stats


if __name__ == '__main__':
    import sys
    print("Ce module s'utilise via generate_evenements(...) — voir test_mai_2026.py pour un exemple.")

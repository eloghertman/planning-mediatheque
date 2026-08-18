# -*- coding: utf-8 -*-
"""
app.py — Assistant Planning Médiathèque
=========================================

Application en une seule page (défilement vertical), organisée en 3 blocs :

  1. Créer l'onglet Événements à partir des fichiers sources bruts
  2. Générer le planning mensuel (à partir du fichier Événements + du
     fichier de Préparation mensuelle)
  3. Vérifier un planning déjà rempli / modifié à la main

Seul le bloc 1 est fonctionnel pour l'instant. Les blocs 2 et 3 sont affichés
en aperçu (grisés) pour montrer la structure finale de la page, en attendant
d'être branchés au moteur CP-SAT.
"""

import io
import os
import tempfile

import streamlit as st

from sources_to_evenements import generate_evenements, MOIS_FR_CAP
from generate_planning_excel_septembre import generer

st.set_page_config(page_title="Planning Médiathèque", page_icon="📅", layout="centered")


# ══════════════════════════════════════════════════════════════
#  OUTILS
# ══════════════════════════════════════════════════════════════

def _fusionner_evenements_dans_preparation(prep_path, evenements_path, out_path):
    """Copie l'onglet Événements du fichier généré au bloc 1 dans le fichier
    de Préparation mensuelle (en remplaçant l'onglet existant s'il y en a
    déjà un), pour obtenir un seul fichier combiné à donner au moteur de
    calcul — qui, lui, n'accepte qu'un seul fichier en entrée."""
    import openpyxl as _openpyxl
    wb_prep = _openpyxl.load_workbook(prep_path)
    wb_evt = _openpyxl.load_workbook(evenements_path, data_only=True)

    if 'Événements' in wb_evt.sheetnames:
        src_name = 'Événements'
    elif 'Evenements' in wb_evt.sheetnames:
        src_name = 'Evenements'
    else:
        src_name = wb_evt.sheetnames[0]
    ws_src = wb_evt[src_name]

    for name in ('Événements', 'Evenements'):
        if name in wb_prep.sheetnames:
            del wb_prep[name]
    ws_dst = wb_prep.create_sheet('Événements')
    for row in ws_src.iter_rows(values_only=True):
        ws_dst.append(row)

    wb_prep.save(out_path)
    return out_path

def _save_uploaded(uploaded_file, tmp_dir):
    """Enregistre un fichier uploadé par Streamlit sur le disque (les
    fonctions de lecture Excel ont besoin d'un chemin de fichier, pas
    directement du fichier envoyé par le navigateur)."""
    if uploaded_file is None:
        return None
    path = os.path.join(tmp_dir, uploaded_file.name)
    with open(path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return path


def _save_uploaded_list(uploaded_files, tmp_dir):
    if not uploaded_files:
        return []
    return [_save_uploaded(f, tmp_dir) for f in uploaded_files]


st.title("📅 Assistant Planning Médiathèque")
st.caption(
    "Une page, trois étapes. Chaque étape est indépendante : tu génères un "
    "fichier, tu le télécharges, et tu le réutilises toi-même à l'étape "
    "suivante si besoin."
)

st.divider()

# ══════════════════════════════════════════════════════════════
#  BLOC 1 — CRÉER L'ONGLET ÉVÉNEMENTS
# ══════════════════════════════════════════════════════════════

st.header("1. Créer l'onglet Événements")
st.markdown(
    "**Ce que fait ce bloc :** tu déposes ici les fichiers bruts que tu tiens "
    "déjà au fil du mois (congés, accueil libre crèche, accueil de classe, "
    "Lectures AssMat/AssPar, fichier Excel Evenement). L'app les lit tous et "
    "construit automatiquement un fichier Excel avec l'onglet **Événements** "
    "prêt à l'emploi, en surlignant en jaune les cas où il manque une "
    "information (heure ou nom d'agent) pour que tu puisses les compléter "
    "en un coup d'œil. Tu n'es pas obligée de fournir tous les fichiers : "
    "ne dépose que ceux que tu as pour ce mois-ci."
)

with st.form("form_evenements"):
    col1, col2 = st.columns(2)
    with col1:
        mois_label = st.selectbox(
            "Mois concerné",
            options=list(MOIS_FR_CAP.values()),
            index=8,  # Septembre par défaut
            help="Le mois pour lequel on construit les événements.",
        )
        mois_num = {v: k for k, v in MOIS_FR_CAP.items()}[mois_label]
    with col2:
        annee = st.number_input("Année", min_value=2020, max_value=2100, value=2026, step=1)

    st.markdown("**Congés équipe**")
    st.caption(
        "Le fichier Excel des congés, avec un onglet par mois et une ligne "
        "par agent (le fichier où une lettre = une journée de congé)."
    )
    f_conges = st.file_uploader("Fichier congés", type=["xlsx"], key="f_conges")

    st.markdown("**Accueil libre crèche**")
    st.caption(
        "Le(s) fichier(s) de suivi des accueils de crèches. Tu peux en "
        "déposer plusieurs si tu as un fichier par année scolaire. Ce sont "
        "des créneaux en visite libre : aucun agent n'est jamais affecté à "
        "cet événement, et c'est normal — pas d'alerte pour ça."
    )
    f_creche = st.file_uploader(
        "Fichier(s) accueil libre crèche", type=["xlsx"], accept_multiple_files=True, key="f_creche"
    )

    st.markdown("**Accueil de classe**")
    st.caption(
        "Le(s) fichier(s) de suivi des accueils de classes scolaires. Quand "
        "une ligne indique 'uniquement en visite libre', l'app le détecte "
        "automatiquement : l'événement est noté 'Accueil libre école', sans "
        "agent affecté et sans alerte."
    )
    f_classe = st.file_uploader(
        "Fichier(s) accueil de classe", type=["xlsx"], accept_multiple_files=True, key="f_classe"
    )

    st.markdown("**Lectures AssMat/AssPar**")
    st.caption("Le(s) fichier(s) de suivi des séances de lecture du jeudi matin.")
    f_lecture = st.file_uploader(
        "Fichier(s) Lectures AssMat/AssPar", type=["xlsx"], accept_multiple_files=True, key="f_lecture"
    )

    st.markdown("**Fichier Excel Evenement**")
    st.caption(
        "Réinjecte le fichier Excel \"calendrier des événements\" et indique "
        "le nom exact de son onglet."
    )
    f_calendrier = st.file_uploader(
        "Fichier Excel Evenement", type=["xlsx"], key="f_calendrier"
    )
    nom_onglet_calendrier = st.text_input(
        "Nom de l'onglet à réinjecter", value="Événements", key="onglet_calendrier"
    )

    st.markdown("**Autres documents (facultatif)**")
    st.caption(
        "Dépose ici tout autre document utile pour ce mois, même plusieurs. "
        "Pour l'instant ces documents ne sont pas encore lus automatiquement "
        "par l'app — préviens-moi de leur contenu et j'ajouterai la lecture "
        "correspondante si besoin."
    )
    f_autres = st.file_uploader(
        "Autres documents", accept_multiple_files=True, key="f_autres"
    )

    st.warning(
        "⚠️ **Avant de générer l'onglet événement**, et en plus des infos "
        "fournies ci-dessus, vérifie que tu as bien pris en compte les "
        "dates de : **portage, hors les murs, séances d'éloquence, CAJ, "
        "formations, et réunions/rdv/événements exceptionnels.**"
    )

    submitted = st.form_submit_button("Générer l'onglet Événements", type="primary")

if submitted:
    if not any([f_conges, f_creche, f_classe, f_lecture, f_calendrier]):
        st.warning("Dépose au moins un fichier source avant de générer.")
    else:
        with st.spinner("Lecture des fichiers et construction de l'onglet Événements…"):
            tmp_dir = tempfile.mkdtemp()
            try:
                sources = {}
                p_conges = _save_uploaded(f_conges, tmp_dir)
                if p_conges:
                    sources["conges"] = p_conges

                p_creche = _save_uploaded_list(f_creche, tmp_dir)
                if p_creche:
                    sources["accueil_creche"] = p_creche

                p_classe = _save_uploaded_list(f_classe, tmp_dir)
                if p_classe:
                    sources["accueil_classe"] = p_classe

                p_lecture = _save_uploaded_list(f_lecture, tmp_dir)
                if p_lecture:
                    sources["lecture_jeudi"] = p_lecture

                p_calendrier = _save_uploaded(f_calendrier, tmp_dir)
                if p_calendrier and nom_onglet_calendrier.strip():
                    sources["calendrier"] = (p_calendrier, nom_onglet_calendrier.strip())

                # Les "autres documents" sont sauvegardés mais pas encore
                # interprétés automatiquement (aucun format connu pour eux) :
                # on se contente de les garder de côté pour l'instant.
                _save_uploaded_list(f_autres, tmp_dir)

                out_path = os.path.join(tmp_dir, f"Evenements_{mois_label}{annee}.xlsx")
                events, stats = generate_evenements(mois_num, int(annee), out_path, sources=sources)

            except Exception as e:
                st.error(
                    "Un fichier n'a pas pu être lu correctement. Vérifie que "
                    "c'est bien le bon fichier pour le bon mois, et que sa "
                    "mise en page n'a pas changé.\n\n"
                    f"Détail technique : {e}"
                )
                st.stop()

        # ── Résumé à l'écran ──
        st.success(f"{stats['total']} événement(s) trouvé(s).")
        if stats["alerts"]:
            st.warning(
                f"⚠️ {stats['alerts']} événement(s) incomplet(s), surligné(s) "
                "en jaune dans le fichier — à compléter à la main avant "
                "de l'utiliser pour générer le planning."
            )
        else:
            st.info("Aucune information manquante détectée.")

        if stats.get("par_source"):
            st.markdown("**Détail par source :**")
            for source_name, n in stats["par_source"].items():
                st.markdown(f"- {source_name} : {n} événement(s)")

        with open(out_path, "rb") as f:
            file_bytes = f.read()

        st.download_button(
            "⬇️ Télécharger le fichier Événements",
            data=file_bytes,
            file_name=os.path.basename(out_path),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

st.divider()

# ══════════════════════════════════════════════════════════════
#  BLOC 2 — GÉNÉRER LE PLANNING MENSUEL (aperçu, pas encore branché)
# ══════════════════════════════════════════════════════════════

st.header("2. Générer le planning mensuel")
st.markdown(
    "**Ce que fait ce bloc :** dépose ici deux fichiers — le fichier "
    "Événements (celui généré au bloc 1, ou un ancien si tu le réutilises) "
    "et le fichier de Préparation mensuelle. L'app les combine, calcule le "
    "planning complet du mois et te le rend en Excel, prêt à imprimer. "
    "S'il reste des créneaux impossibles à couvrir malgré le calcul, ils "
    "sont listés clairement ci-dessous, en plus d'être signalés dans le "
    "fichier lui-même."
)

with st.form("form_planning"):
    f_evenements_b2 = st.file_uploader(
        "Fichier Événements", type=["xlsx"], key="f_evenements_b2"
    )
    f_preparation_b2 = st.file_uploader(
        "Fichier Préparation mensuelle", type=["xlsx"], key="f_preparation_b2"
    )
    submitted_b2 = st.form_submit_button("Générer le planning", type="primary")

if submitted_b2:
    if not f_preparation_b2:
        st.warning("Le fichier de Préparation mensuelle est obligatoire.")
    else:
        with st.spinner(
            "Calcul du planning en cours — ça peut prendre une à quelques "
            "minutes selon la taille du mois, merci de patienter…"
        ):
            tmp_dir = tempfile.mkdtemp()
            try:
                p_prep = _save_uploaded(f_preparation_b2, tmp_dir)

                if f_evenements_b2:
                    p_evt = _save_uploaded(f_evenements_b2, tmp_dir)
                    p_fusionne = os.path.join(tmp_dir, "Preparation_fusionnee.xlsx")
                    _fusionner_evenements_dans_preparation(p_prep, p_evt, p_fusionne)
                    input_path = p_fusionne
                else:
                    input_path = p_prep

                output_path = os.path.join(tmp_dir, "Planning_genere.xlsx")
                output_path, weeks_data, metadata = generer(input_path, output_path)

            except Exception as e:
                st.error(
                    "Le calcul n'a pas pu aboutir. Vérifie que le fichier de "
                    "Préparation a bien tous ses onglets (Affectations, "
                    "Horaires_Des_Agents, Paramètres, Planning_type...) et "
                    "qu'il n'a pas été modifié dans sa structure.\n\n"
                    f"Détail technique : {e}"
                )
                st.stop()

        # ── Résumé des alertes à l'écran ──
        alertes_par_jour = []
        for w in weeks_data:
            for j in w["jours"]:
                if j.get("alertes"):
                    alertes_par_jour.append((j["date"], j["alertes"]))

        if alertes_par_jour:
            st.warning(
                f"⚠️ {sum(len(a) for _, a in alertes_par_jour)} créneau(x) "
                "n'ont pas pu être entièrement couverts — à traiter à la main."
            )
            for date_str, alertes in alertes_par_jour:
                for cren_idx, section, message in alertes:
                    st.markdown(f"- **{date_str}** — {section} : {message}")
        else:
            st.success("Aucune alerte : tous les créneaux ont été couverts.")

        # ── Nom de fichier dynamique : Planning_MoisAnnée.xlsx, déduit de la
        # première date réellement présente dans le planning calculé ──
        premiere_date = weeks_data[0]["jours"][0]["date"]  # 'YYYY-MM-DD'
        annee_planning = premiere_date[:4]
        mois_num_planning = int(premiere_date[5:7])
        mois_nom_planning = MOIS_FR_CAP[mois_num_planning]
        nom_fichier_final = f"Planning_{mois_nom_planning}{annee_planning}.xlsx"

        with open(output_path, "rb") as f:
            file_bytes = f.read()

        st.download_button(
            "⬇️ Télécharger le planning généré",
            data=file_bytes,
            file_name=nom_fichier_final,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_planning",
        )

st.divider()

# ══════════════════════════════════════════════════════════════
#  BLOC 3 — VÉRIFIER UN PLANNING (aperçu, pas encore branché)
# ══════════════════════════════════════════════════════════════

st.header("3. Vérifier un planning")
st.markdown(
    "**Ce que fera ce bloc :** tu déposeras un fichier planning déjà rempli "
    "(éventuellement modifié à la main) dans la mise en page habituelle. "
    "L'app relira chaque case et te listera les anomalies trouvées "
    "(quelqu'un affecté à deux endroits en même temps, un agent en congé "
    "mais quand même planifié, Eloïse planifiée par erreur, un effectif "
    "minimum non respecté, etc.), avec la date et le créneau concernés."
)
st.info("🚧 Ce bloc sera construit après les blocs 1 et 2.")
st.file_uploader("Fichier planning à vérifier", disabled=True, key="stub_verif")
st.button("Vérifier le planning", disabled=True, key="stub_verifier")

"""
regeneration_ecriture.py — Brique 3 de la "régénération partielle".

Prend le fichier original (bytes) + le résultat de la brique 2, et produit
un NOUVEAU fichier Excel (bytes) où :

- Seules les colonnes B à G (RDC/Adulte/M&F/Jeunesse 1-2-3) des jours
  régénérés sont réécrites, avec le même style que le générateur habituel
  (couleurs par agent, bordures) — réutilise agent_cell_style/fmt_agents/
  GREY_BORDER de generate_planning_excel_septembre.py, pas de logique de
  style dupliquée.
- Tout le reste du classeur (autres jours, autres semaines, colonnes H à Z,
  onglets cachés, mise en forme) est laissé strictement intact — on ne
  touche même pas aux formules des colonnes H/I/J : elles restent celles
  déjà présentes dans le fichier.
- Les conflits entre éléments fixes (ConflitFixe, brique 1) sont signalés
  par une bordure rouge très visible + un commentaire Excel sur CHACUNE
  des deux cellules en cause.

  ⚠️ Choix technique à connaître : on n'écrit PAS de texte d'avertissement
  directement DANS la cellule H/I/J elle-même, parce que ces cellules
  contiennent une formule (celle qui combine les notes W-Z) — écrire
  dedans supprimerait cette formule, et casserait la remontée automatique
  de futures notes agents sur ce créneau. À la place, on pose une bordure
  rouge épaisse (même mécanisme que les alertes de couverture déjà utilisées
  ailleurs dans l'outil) + un commentaire Excel, qui affiche un petit
  triangle rouge visible dans le coin de la cellule, cliquable pour lire le
  détail — repérable au premier coup d'œil sans abîmer le fichier.
"""

from io import BytesIO

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.comments import Comment

from generate_planning_excel_septembre import (
    agent_cell_style, fmt_agents, GREY_BORDER, generer_vue_agent, is_vacataire,
)
from planning_checker import JOURS_ORDRE, _trouver_onglet_insensible_casse

BORDURE_ALERTE = Border(*[Side(style='thick', color='FFE74C3C')] * 4)

COL_PAR_TYPE = {
    'Accueil/Animation': 8,   # H
    'Réunion': 9,             # I
    'Absence': 10,            # J
}


def _demerger_si_besoin(ws, row, col):
    """openpyxl refuse d'écrire dans une cellule fusionnée qui n'est pas la
    cellule d'ancrage. Si la cellule (row, col) fait partie d'une fusion, on
    la défait avant d'écrire (les colonnes B-G ne sont normalement JAMAIS
    fusionnées par le générateur — cette fonction est une sécurité, au cas
    où une fusion aurait été ajoutée à la main)."""
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            ws.unmerge_cells(str(rng))
            return


def _ecrire_cellule_bg(ws, row, col, agents_liste):
    """Écrit une cellule B-G (RDC/Adulte/M&F/Jeunesse X) avec le même style
    que le générateur normal (fond + texte colorés par agent)."""
    _demerger_si_besoin(ws, row, col)
    texte = fmt_agents(agents_liste)
    cell = ws.cell(row=row, column=col, value=texte)
    fill, text_color = agent_cell_style(agents_liste)
    cell.fill = fill
    from openpyxl.styles import Font, Alignment
    cell.font = Font(size=10, bold=True, color='FF' + text_color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = GREY_BORDER


def _trouver_row_evenement(jour_data, ev):
    """Retrouve la ligne Excel exacte d'un événement (dict avec 'cs','ce',
    'nom_affichage') à l'intérieur du bloc du jour. Le texte de la case
    combine parfois plusieurs événements séparés par '; ' — on cherche donc
    une case qui CONTIENT ce segment, pas une égalité stricte."""
    champ = {'Accueil/Animation': 'accueil', 'Réunion': 'reunion', 'Absence': 'absence'}[ev['type']]
    cible = ev['nom_affichage']
    for cren in jour_data['creneaux']:
        val = cren.get(champ)
        if val and cible in val and cren['debut'] <= ev['cs'] < cren['fin']:
            return cren['row']
    # Repli : première case qui contient ce segment, même si l'horaire exact
    # ne tombe pas dedans (texte combiné sur plusieurs créneaux fusionnés).
    for cren in jour_data['creneaux']:
        val = cren.get(champ)
        if val and cible in val:
            return cren['row']
    return None


def _poser_alerte(ws, jour_data, ev, texte_commentaire):
    """Pose la bordure rouge + le commentaire sur la cellule (ou tout le
    bloc fusionné) correspondant à l'événement `ev`."""
    row = _trouver_row_evenement(jour_data, ev)
    if row is None:
        return False  # rien trouvé — on ne casse rien, juste pas d'alerte posée
    col = COL_PAR_TYPE[ev['type']]

    # Si la cellule fait partie d'une fusion (cas fréquent pour H/I/J),
    # applique la bordure sur TOUTES les lignes de la fusion pour que le
    # rendu visuel soit cohérent (même logique que fusionner_cellules_identiques).
    lignes_a_marquer = [row]
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            lignes_a_marquer = list(range(rng.min_row, rng.max_row + 1))
            row = rng.min_row  # le commentaire va sur la cellule d'ancrage
            break

    for r in lignes_a_marquer:
        ws.cell(row=r, column=col).border = BORDURE_ALERTE

    cell = ws.cell(row=row, column=col)
    texte = '⚠ CONFLIT : ' + texte_commentaire
    if cell.comment:
        cell.comment.text += '\n' + texte
    else:
        cell.comment = Comment(texte, 'Régénération partielle')
    return True


def ecrire_regeneration(file_bytes, lecture_resultat, calcul_resultat):
    """Retourne les bytes du nouveau fichier Excel (le fichier original
    n'est jamais modifié en place)."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes))  # data_only=False : on garde les formules
    semaine_num = lecture_resultat['semaine_num']
    nom_onglet = _trouver_onglet_insensible_casse(wb, f'Semaine_{semaine_num}')
    ws = wb[nom_onglet]
    jours_dispo = lecture_resultat['jours_data_bruts']

    jours_infaisables = []

    # ── 1. Réécriture des colonnes B à G des jours régénérés ────────────
    for jour_res in calcul_resultat['jours']:
        jour_maj = jour_res['jour']
        jour_data = jours_dispo[jour_maj]
        solution = jour_res['solution']
        creneaux_ouverts = jour_res['creneaux']

        if solution is None:
            # Aucune solution trouvée pour ce jour : on NE TOUCHE PAS à son
            # contenu existant plutôt que d'écrire n'importe quoi — le jour
            # garde ses affectations manuelles précédentes, telles quelles.
            jours_infaisables.append(jour_maj)
            continue

        for cren in jour_data['creneaux']:
            # Retrouver l'indice du créneau régénéré correspondant à cette
            # ligne Excel (même bornes horaires).
            c_idx = next((i for i, (cs, ce) in enumerate(creneaux_ouverts)
                          if cs == cren['debut'] and ce == cren['fin']), None)
            if c_idx is None:
                continue  # grille horaire différente pour cette ligne — on ne touche pas
            sol_c = solution.get(c_idx, {})
            rdc_l = sol_c.get('RDC', [])
            adulte_l = sol_c.get('Adulte', [])
            mf_l = sol_c.get('MF', [])
            jeun_l = sol_c.get('Jeunesse', [])
            r = cren['row']
            _ecrire_cellule_bg(ws, r, 2, rdc_l)
            _ecrire_cellule_bg(ws, r, 3, adulte_l)
            _ecrire_cellule_bg(ws, r, 4, mf_l)
            _ecrire_cellule_bg(ws, r, 5, jeun_l[0:1])
            _ecrire_cellule_bg(ws, r, 6, jeun_l[1:2])
            _ecrire_cellule_bg(ws, r, 7, jeun_l[2:3])
            # Colonnes cachées L-Q : formules '=B{r}' etc déjà en place,
            # elles se recalculeront automatiquement à l'ouverture du
            # fichier dans Excel — rien à faire ici.

    # ── 2. Reconstruction de l'onglet "vue par agent" ("Semaine_N_Agent") ──
    # Cet onglet est reconstruit à CHAQUE régénération, même partielle
    # (certains jours de la semaine restent fixes) — corrigé le 22/08 :
    # avant, ce n'était fait que si toute la semaine était régénérée, sinon
    # l'onglet gardait l'affichage d'AVANT la régénération (le lien direct
    # avec les notes W-Z semblait "cassé"). La brique 1 fournit maintenant
    # les événements (congés/réunions/absences) de TOUS les jours de la
    # semaine, régénérés ou fixes, ce qui permet cette reconstruction
    # complète et correcte à chaque fois.
    prep = lecture_resultat['prep']
    affectations = prep.get('affectations', {})
    horaires_agents = prep.get('horaires_agents', {})
    pause_flex = prep.get('pause_flex', set())
    agents_recap_vue_agent = [a for a in affectations.keys() if not is_vacataire(a)]

    jours_ordonnes = sorted(jours_dispo.keys(),
                              key=lambda j: JOURS_ORDRE.index(j) if j in JOURS_ORDRE else 99)
    jours_arg = []
    row_lookup = {}
    for jour_maj in jours_ordonnes:
        jd = jours_dispo[jour_maj]
        jour_cap = jour_maj.capitalize()  # 'MERCREDI' -> 'Mercredi', suffit ici
        creneaux_jour = [(cren['debut'], cren['fin']) for cren in jd['creneaux']]
        jours_arg.append({'jour': jour_cap, 'date': jd.get('date_str'),
                           'creneaux': creneaux_jour})
        for cren in jd['creneaux']:
            row_lookup[(jour_cap, cren['debut'], cren['fin'])] = cren['row']

    nom_agent_sheet = f'Semaine_{semaine_num}_Agent'
    ancien_agent_sheet = _trouver_onglet_insensible_casse(wb, nom_agent_sheet)
    if ancien_agent_sheet is not None:
        ancien_index = wb.sheetnames.index(ancien_agent_sheet)
        del wb[ancien_agent_sheet]
    else:
        ancien_index = wb.sheetnames.index(nom_onglet) + 1

    generer_vue_agent(wb, semaine_num, jours_arg, row_lookup, agents_recap_vue_agent,
                       horaires_agents, pause_flex, lecture_resultat['evenements_tous_jours'])
    # generer_vue_agent recrée l'onglet à la FIN du classeur — on le
    # replace à sa position d'origine (juste après Semaine_N).
    nouvel_onglet = wb[nom_agent_sheet]
    wb._sheets.remove(nouvel_onglet)
    wb._sheets.insert(ancien_index, nouvel_onglet)
    agent_sheet_reconstruit = True

    # ── 3. Alertes visuelles sur les conflits entre éléments fixes ─────
    conflits = calcul_resultat.get('conflits_a_signaler', [])
    for c in conflits:
        jour_data = next((jd for jd in jours_dispo.values()
                           if jd.get('date_str') == c.date), None)
        if jour_data is None:
            continue
        msg1 = (f"{c.agent} est aussi noté·e « {c.evenement_2['nom_affichage']} » de "
                 f"{c.evenement_2['cs']//60}h{c.evenement_2['cs']%60:02d} à "
                 f"{c.evenement_2['ce']//60}h{c.evenement_2['ce']%60:02d} — "
                 f"à vérifier, non tranché automatiquement.")
        msg2 = (f"{c.agent} est aussi noté·e « {c.evenement_1['nom_affichage']} » de "
                 f"{c.evenement_1['cs']//60}h{c.evenement_1['cs']%60:02d} à "
                 f"{c.evenement_1['ce']//60}h{c.evenement_1['ce']%60:02d} — "
                 f"à vérifier, non tranché automatiquement.")
        _poser_alerte(ws, jour_data, c.evenement_1, msg1)
        _poser_alerte(ws, jour_data, c.evenement_2, msg2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), jours_infaisables, agent_sheet_reconstruit

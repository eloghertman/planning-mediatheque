"""
planning_checker.py — Vérification d'un planning déjà rempli / modifié à la main.

Contrairement à planning_engine_cpsat.py (qui CALCULE un planning), ce module
RELIT un planning déjà généré et modifié, et signale les contradictions avec
les contraintes dures. Un seul fichier en entrée : le classeur Excel du
planning (même structure que la sortie de generate_planning_excel_septembre.py).

Fonction principale : verifier_planning(file_bytes) -> list[Anomalie]
"""

import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from io import BytesIO

import openpyxl

from planning_engine_cpsat import (
    parse_parametres, parse_affectations, parse_horaires_agents,
    parse_roulement_samedi, agent_disponible, is_vacataire, _parse_fr_date,
    parse_planning_type, parse_besoins_jeunesse, parse_jours_speciaux,
    parse_creneau as parse_creneau_engine,
    parse_horaires_agents_grille, ONGLET_HORAIRES_GRILLE,
)

# Onglets de préparation recopiés (très masqués) par generate_planning_excel_septembre.py
# (fonction copier_onglets_preparation_caches). Préfixe '_prep_' + nom d'origine.
ONGLETS_PREP_PREFIXE = '_prep_'
# Onglets recopiés en VISIBLE (sans préfixe '_prep_') plutôt qu'en très masqué :
# 'Planning_type' et "horaires d'équipes" (consultables, verrouillés) depuis
# 09/2026 ; 'Paramètres' et 'Affectations' (modifiables) depuis 08/2026, pour
# permettre les mises à jour en cours de mois (ex. nouvelle habilitation
# d'un agent) suivies d'une régénération partielle (bloc 4). Repli automatique
# sur la version masquée '_prep_...' pour les fichiers générés par une version
# antérieure de l'outil (voir boucle ci-dessous).
ONGLETS_PREP_SANS_PREFIXE = {'Planning_type', 'Paramètres', 'Affectations'}
ONGLETS_PREP_NOMS = [
    'Paramètres', 'Horaires_Des_Agents', 'Affectations', 'Roulement_Samedi',
    # Ajoutés (09/2026) : nécessaires à la vérification de la couverture
    # RDC/Adulte/M&F/Jeunesse par rapport au planning type (§ ci-dessous,
    # R10). Un fichier généré avec une version antérieure de l'outil ne les
    # contiendra pas : ils apparaîtront alors dans 'manquants' et la
    # vérification correspondante sera simplement sautée (pas d'erreur).
    'Planning_type', 'Besoins_Jeunesse',
]
# Onglet facultatif : améliore la précision (jours de vacances scolaires
# ponctuels / fériés en dehors du réglage habituel par semaine) mais n'est
# pas indispensable — son absence n'est donc pas comptée dans 'manquants'.
ONGLET_JOURS_SPECIAUX = 'Jours_speciaux'


# ─────────────────────────────────────────────────────────────
#  RÉFÉRENTIEL MÉTIER (voir context_projet_mediatheque_v30.md §3, §7)
# ─────────────────────────────────────────────────────────────

HABILITATIONS = {
    'Marie-France':    ['RDC', 'Adulte', 'M & F'],
    'Anne-Françoise':  ['Adulte', 'Jeunesse', 'M & F', 'RDC'],
    'Christine':       ['Adulte', 'RDC'],
    'Léa':             ['Adulte', 'M & F', 'RDC', 'Jeunesse'],
    'Chloé':           ['Adulte', 'RDC', 'Jeunesse'],
    'Macha':           ['RDC', 'Adulte'],
    'Delphine':        ['M & F', 'RDC', 'Jeunesse', 'Adulte'],
    'Barbara':         ['M & F', 'Jeunesse'],
    'Stéphane':        ['M & F'],
    'Stéphanie':       ['Jeunesse', 'RDC'],
    'Robin':           ['Jeunesse', 'RDC'],
    'Guillaume':       ['Jeunesse', 'RDC'],
    'Agnès':           ['Jeunesse'],
    'Tiphaine':        ['Jeunesse', 'RDC', 'M & F', 'Adulte'],
}
ALL_AGENTS_CONNUS = list(HABILITATIONS.keys()) + ['Vacataire 1', 'Vacataire 2', 'Vacataire 3', 'Vacataire']

PAUSE_EXEMPTS = {'Delphine'}          # jamais de contrôle de pause déjeuner (§7 C3)
AGENTS_IGNORES = {'lydie'}            # a quitté l'équipe, ignorée partout

JOURS_ORDRE = ['LUNDI', 'MARDI', 'MERCREDI', 'JEUDI', 'VENDREDI', 'SAMEDI', 'DIMANCHE']
# planning_engine_cpsat.py utilise 'Mardi', 'Mercredi'... (1ère lettre capitale
# seulement) comme clé de jour dans Horaires_Des_Agents — ici on manipule les
# jours en MAJUSCULES (titres de bloc), d'où cette table de correspondance.
JOUR_CAPITALISE = {j: j.capitalize() for j in JOURS_ORDRE}

PAUSE_FENETRE = (12 * 60, 14 * 60)    # 12h-14h, en minutes
PAUSE_MIN_LIBRE = 60                  # minutes


# ─────────────────────────────────────────────────────────────
#  UTILITAIRES TEXTE / TEMPS
# ─────────────────────────────────────────────────────────────

def normalize(s):
    if not s:
        return ''
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s


def canon_section(s):
    """Normalise un nom de section pour comparaison, insensible aux variantes
    d'écriture ('M & F', 'M&F', 'MF', espaces, accents, casse...). Le fichier
    Affectations et les en-têtes du planning n'utilisent pas toujours
    exactement la même orthographe pour désigner la même section."""
    return re.sub(r'[^a-z]', '', normalize(s))


def est_vacataire(nom):
    return is_vacataire(nom or '')


def est_eloise(nom):
    return normalize(nom) in ('eloise', 'eloïse'.replace('ï', 'i'), normalize('Eloïse'))


def est_ignore(nom):
    return normalize(nom) in AGENTS_IGNORES


def fmt_min(m):
    if m is None:
        return '?'
    h, mn = divmod(int(m), 60)
    return f"{h}h{mn:02d}" if mn else f"{h}h"


def is_creneau(val):
    return isinstance(val, str) and re.match(r'^\d{1,2}:\d{2}-\d{1,2}:\d{2}$', val.strip()) is not None


def parse_creneau(val):
    a, b = val.strip().split('-')
    h1, m1 = a.split(':')
    h2, m2 = b.split(':')
    return int(h1) * 60 + int(m1), int(h2) * 60 + int(m2)


def parse_heure_texte(txt):
    """'8h30' / '17h' -> minutes. None si non trouvé."""
    m = re.search(r'(\d{1,2})h(\d{2})?', txt)
    if not m:
        return None
    return int(m.group(1)) * 60 + int(m.group(2) or 0)


def parenthese_finale(texte):
    m = re.search(r'\(([^)]*)\)\s*$', texte or '')
    return m.group(1) if m else None


def _extraire_un_segment(segment, defaut_debut, defaut_fin, agents_connus):
    """Version 'un seul événement' de l'extraction — reprend la logique
    d'origine, appliquée à UN SEUL segment de texte (pas toute la case)."""
    agents_trouves = []
    debut, fin = defaut_debut, defaut_fin
    inner = parenthese_finale(segment)
    if inner:
        for part in inner.split(','):
            part = part.strip()
            m = re.match(r'^(\d{1,2})h(\d{2})?-(\d{1,2})h(\d{2})?$', part)
            if m:
                h1, m1, h2, m2 = m.groups()
                debut = int(h1) * 60 + int(m1 or 0)
                fin = int(h2) * 60 + int(m2 or 0)
            else:
                for agent in agents_connus:
                    if normalize(agent) == normalize(part):
                        agents_trouves.append(agent)
                        break
    # Cas particulier "congé (Nom1, Nom2)" : le nom du champ précède la parenthèse.
    if segment and normalize(segment).startswith('conge') and not agents_trouves and inner:
        for part in inner.split(','):
            part = part.strip()
            for agent in agents_connus:
                if normalize(agent) == normalize(part):
                    agents_trouves.append(agent)
                    break
    return agents_trouves, debut, fin


def extraire_agents_et_fenetre(texte, defaut_debut, defaut_fin, agents_connus):
    """À partir d'un texte du type 'Accueil de classe (Stéphanie, 10h15-10h45)'
    ou 'congé (Marie-France, Christine)', retourne (liste_agents, debut, fin).

    Une case peut combiner PLUSIEURS événements distincts, ajoutés au fil de
    l'eau par différent·es agent·es, séparés par '; ' (cf. cascade des notes
    W-Z, §14 du contexte projet) — ex. 'absence (Anne-Françoise); congé
    (Stéphane)'. On découpe donc d'abord sur ce séparateur avant d'extraire,
    pour ne perdre aucun segment (l'ancienne version ne lisait que la toute
    dernière parenthèse de la case entière — bug découvert le 20/08 sur un
    fichier réel : un agent absent noté AVANT un autre événement dans la même
    case n'était jamais détecté).

    ⚠️ Conservée pour compatibilité (signature d'origine) : quand plusieurs
    segments ont des horaires DIFFÉRENTS, cette fonction renvoie la fenêtre
    du dernier segment traité et la liste cumulée des agents — imprécis dans
    ce cas de figure précis. Préférer `extraire_occurrences_multiples()`
    (ci-dessous) pour un résultat exact par agent — c'est elle qu'utilise
    désormais `construire_occurrences_jour`."""
    agents_trouves = []
    debut, fin = defaut_debut, defaut_fin
    for segment in re.split(r';\s*', texte or ''):
        segment = segment.strip()
        if not segment:
            continue
        a, d, f = _extraire_un_segment(segment, defaut_debut, defaut_fin, agents_connus)
        agents_trouves.extend(a)
        if a:
            debut, fin = d, f
    return agents_trouves, debut, fin


def extraire_occurrences_multiples(texte, defaut_debut, defaut_fin, agents_connus):
    """Comme extraire_agents_et_fenetre, mais renvoie une liste de
    (agent, debut, fin, segment) — un quadruplet par segment/agent trouvé,
    chacun avec SA PROPRE fenêtre horaire si elle est précisée dans son
    segment, ET son PROPRE texte source (pas le texte de la case entière —
    corrige un bug découvert le 20/08 : une case combinant plusieurs
    événements, ex. 'absence (Anne-Françoise); congé (Stéphane)', faisait
    porter le texte COMPLET à chaque agent, y compris à ceux qui n'étaient
    concernés que par un seul des deux segments — ça empêchait notamment de
    reconnaître un 'congé' comme tel dès que combiné avec autre chose dans
    la même case). C'est la version à utiliser pour construire les
    occurrences par agent, afin qu'un chevauchement (ex. absence 10h-11h
    d'un·e agent·e + réunion d'un·e autre dans la même case) soit détecté
    pour la bonne personne, au bon horaire, avec le bon texte — pas
    seulement pour le dernier segment de la case."""
    resultat = []
    for segment in re.split(r';\s*', texte or ''):
        segment = segment.strip()
        if not segment:
            continue
        agents, d, f = _extraire_un_segment(segment, defaut_debut, defaut_fin, agents_connus)
        for agent in agents:
            resultat.append((agent, d, f, segment))
    return resultat


# ─────────────────────────────────────────────────────────────
#  ANOMALIE
# ─────────────────────────────────────────────────────────────

@dataclass
class Anomalie:
    gravite: str          # 'rouge' (impossibilité) ou 'jaune' (suspect)
    semaine: str
    jour: str
    message: str
    regle: str = ''


# ─────────────────────────────────────────────────────────────
#  DONNÉES DE PRÉPARATION (onglets '_prep_...' très masqués, si présents)
# ─────────────────────────────────────────────────────────────

def _detecter_grille_horaires_dans_classeur(wb):
    """Retrouve, dans un classeur DÉJÀ GÉNÉRÉ (planning final), l'onglet
    grille "horaires d'équipes" par sa mise en page (case A6='ADULTES',
    H6='JEUNESSE') plutôt que par son nom d'onglet — même principe que la
    détection déjà faite côté fichier de préparation par
    planning_engine_cpsat.py (_detecter_onglet_horaires_grille), pour rester
    tolérant si l'onglet est renommé dans le planning généré (ex. Elo
    l'appelle "planning_agent"). Retourne le nom de l'onglet trouvé, ou None."""
    for nom in wb.sheetnames:
        ws = wb[nom]
        try:
            a6 = normalize(ws.cell(row=6, column=1).value or '')
            h6 = normalize(ws.cell(row=6, column=8).value or '')
        except Exception:
            continue
        if a6 == 'adultes' and h6 == 'jeunesse':
            return nom
    return None


def _trouver_onglet_insensible_casse(wb, nom):
    """Retrouve un onglet par son nom, insensible à la casse. Nécessaire pour
    'Planning_type' : le générateur écrit ce nom avec un P majuscule depuis
    une certaine version, mais des fichiers plus anciens (et donc encore en
    circulation) l'ont créé en 'planning_type' minuscule — confirmé le 24/08
    en test réel (Elo a eu 'Onglet manquant : Planning_type' alors qu'il
    était bien présent, juste minuscule). generate_planning_excel_septembre.py
    tolère déjà les deux casses en LECTURE du fichier de préparation
    (`raw.get('Planning_type') or raw.get('planning_type')`) ; cette fonction
    apporte la même tolérance ici, côté lecture d'un planning déjà généré."""
    for n in wb.sheetnames:
        if n.lower() == nom.lower():
            return n
    return None


def charger_donnees_preparation(wb):
    """Cherche les onglets de préparation dans le classeur, et si présents,
    retourne un dict avec toutes les données de préparation parsées via les
    mêmes fonctions que le moteur de calcul (planning_engine_cpsat.py) —
    même lecture, même vérité.
    La plupart de ces onglets sont recopiés en '_prep_...' très masqués
    (usage interne uniquement). Quatre exceptions, recopiées en onglet
    VISIBLE nommé directement sans préfixe '_prep_' : 'Planning_type' et
    "horaires d'équipes" (demande utilisatrice 09/2026, consultables mais
    verrouillés) ; 'Paramètres' et 'Affectations' (demande utilisatrice
    08/2026, librement modifiables — cf. régénération partielle, bloc 4).
    On les cherche donc sous leur nom tel quel en priorité, avec repli sur
    la version masquée '_prep_...' pour les fichiers générés par une
    version antérieure de l'outil.
    Retourne None si aucun onglet de préparation n'est présent (fichier
    généré avec une version antérieure de generate_planning_excel_septembre.py) :
    dans ce cas, verifier_planning() se rabat sur une vérification
    approximative à partir de la 'vue par agent' seule."""
    raw = {}
    for nom in ONGLETS_PREP_NOMS + [ONGLET_JOURS_SPECIAUX]:
        if nom == 'Horaires_Des_Agents':
            continue  # traité séparément ci-dessous (grille en priorité, repli liste à plat)
        # Onglet visible en priorité (Planning_type, Paramètres, Affectations) ;
        # recherche insensible à la casse (cf. _trouver_onglet_insensible_casse,
        # nécessaire pour 'Planning_type' — voir §25.9 du contexte projet) ;
        # sinon repli sur la version masquée '_prep_...' — couvre à la fois
        # les nouveaux fichiers (visibles) et les anciens déjà générés
        # (masqués uniquement, avant ce changement).
        if nom in ONGLETS_PREP_SANS_PREFIXE:
            trouve = _trouver_onglet_insensible_casse(wb, nom)
            if trouve is not None:
                raw[nom] = wb[trouve]
            else:
                trouve_prefixe = _trouver_onglet_insensible_casse(wb, ONGLETS_PREP_PREFIXE + nom)
                if trouve_prefixe is not None:
                    raw[nom] = wb[trouve_prefixe]
        else:
            trouve_prefixe = _trouver_onglet_insensible_casse(wb, ONGLETS_PREP_PREFIXE + nom)
            if trouve_prefixe is not None:
                raw[nom] = wb[trouve_prefixe]

    # Horaires agents : grille collaborative "horaires d'équipes" en priorité
    # (détectée par sa mise en page, peu importe le nom de l'onglet — cf.
    # _detecter_grille_horaires_dans_classeur ci-dessus, testé et confirmé
    # tolérant à tout nom d'onglet, ex. "horaires_des_agents" — demande
    # utilisatrice 24/08, voir §25.11), repli sur l'ancienne liste à plat
    # 'Horaires_Des_Agents' (visible ou '_prep_...', insensible à la casse)
    # si la grille est absente.
    horaires_source_trouvee = False
    nom_grille = _detecter_grille_horaires_dans_classeur(wb)
    if nom_grille is not None:
        raw[ONGLET_HORAIRES_GRILLE] = wb[nom_grille]
        horaires_source_trouvee = True
    else:
        trouve = (_trouver_onglet_insensible_casse(wb, 'Horaires_Des_Agents')
                  or _trouver_onglet_insensible_casse(wb, ONGLETS_PREP_PREFIXE + 'Horaires_Des_Agents'))
        if trouve is not None:
            raw['Horaires_Des_Agents'] = wb[trouve]
            horaires_source_trouvee = True

    if not raw:
        return None

    manquants = [n for n in ONGLETS_PREP_NOMS if n != 'Horaires_Des_Agents' and n not in raw]
    if not horaires_source_trouvee:
        manquants.append('Horaires_Des_Agents')
    donnees = {'manquants': manquants}
    try:
        if 'Paramètres' in raw:
            donnees['params'] = parse_parametres(raw)
        if 'Affectations' in raw:
            (donnees['affectations'], donnees['categories'], donnees['responsables'],
             donnees['pause_flex'], donnees['priorite_rdc']) = parse_affectations(raw)
        if ONGLET_HORAIRES_GRILLE in raw:
            donnees['horaires_agents'] = parse_horaires_agents_grille(raw)
        elif 'Horaires_Des_Agents' in raw:
            donnees['horaires_agents'] = parse_horaires_agents(raw)
        if 'Roulement_Samedi' in raw:
            donnees['roulement_type'], donnees['roulement_exceptions'] = parse_roulement_samedi(raw)
        if 'Planning_type' in raw:
            donnees['planning_type'] = parse_planning_type(raw)
        if 'Besoins_Jeunesse' in raw:
            donnees['besoins_jeunesse'] = parse_besoins_jeunesse(raw)
        if ONGLET_JOURS_SPECIAUX in raw:
            donnees['jours_speciaux'] = parse_jours_speciaux(raw)
    except Exception as e:
        # Onglet présent mais mal formé : on retombe en mode dégradé plutôt
        # que de faire planter toute la vérification.
        return {'erreur_lecture': str(e)}
    return donnees


# ─────────────────────────────────────────────────────────────
#  LECTURE — cellules / cellules fusionnées
# ─────────────────────────────────────────────────────────────

def build_merge_map(ws):
    merge_map = {}
    for mc in ws.merged_cells.ranges:
        top_val = ws.cell(row=mc.min_row, column=mc.min_col).value
        for row in range(mc.min_row, mc.max_row + 1):
            for col in range(mc.min_col, mc.max_col + 1):
                merge_map[(row, col)] = top_val
    return merge_map


def get_cell(ws, row, col, merge_map):
    if (row, col) in merge_map:
        return merge_map[(row, col)]
    return ws.cell(row=row, column=col).value


# ─────────────────────────────────────────────────────────────
#  LECTURE — feuille "Semaine_N" (grille principale)
# ─────────────────────────────────────────────────────────────

def lire_jours_semaine(ws):
    """Découpe la feuille en blocs 'journée' à partir des titres en colonne A
    ('  MARDI  1 Septembre 2026', éventuellement suivi de '— SAMEDI BLEU')."""
    merge_map = build_merge_map(ws)
    jours = []
    max_row = ws.max_row
    r = 1
    while r <= max_row:
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str):
            v = val.strip().upper()
            jour_trouve = next((j for j in JOURS_ORDRE if v.startswith(j)), None)
            if jour_trouve:
                samedi_type = None
                if 'BLEU' in v:
                    samedi_type = 'BLEU'
                elif 'ROUGE' in v:
                    samedi_type = 'ROUGE'
                # Ligne d'en-tête juste après (contient 'Créneau'), données ensuite.
                r_data = r + 2
                creneaux = []
                while r_data <= max_row and is_creneau(ws.cell(row=r_data, column=1).value):
                    debut, fin = parse_creneau(ws.cell(row=r_data, column=1).value)
                    rdc = get_cell(ws, r_data, 2, merge_map)
                    adulte = get_cell(ws, r_data, 3, merge_map)
                    mf = get_cell(ws, r_data, 4, merge_map)
                    jeun = [get_cell(ws, r_data, c, merge_map) for c in (5, 6, 7)]
                    accueil = get_cell(ws, r_data, 8, merge_map)
                    reunion = get_cell(ws, r_data, 9, merge_map)
                    absence = get_cell(ws, r_data, 10, merge_map)

                    def clean(v):
                        if v in (None, '—', ''):
                            return None
                        return v

                    creneaux.append({
                        'row': r_data, 'debut': debut, 'fin': fin,
                        'rdc': clean(rdc), 'adulte': clean(adulte), 'mf': clean(mf),
                        'jeunesse': [clean(x) for x in jeun],
                        'accueil': clean(accueil), 'reunion': clean(reunion),
                        'absence': clean(absence),
                    })
                    r_data += 1
                jours.append({
                    'jour': jour_trouve, 'titre': val.strip(),
                    'samedi_type': samedi_type,
                    'date_str': (lambda d: d.strftime('%Y-%m-%d') if d else None)(_parse_fr_date(val)),
                    'row_titre': r, 'row_debut_data': r + 2, 'row_fin_data': r_data - 1,
                    'creneaux': creneaux,
                })
                r = r_data
                continue
        r += 1
    return jours


def lire_notes_agents_jour(ws, row_debut, row_fin):
    """Lit les colonnes W/X (Nom/Événement) et Y/Z (Nom/Événement) pour les
    lignes d'un jour donné. Retourne une liste de (agent, texte_note)."""
    notes = []
    for r in range(row_debut, row_fin + 1):
        for col_nom, col_evt in ((23, 24), (25, 26)):  # W/X, Y/Z
            nom = ws.cell(row=r, column=col_nom).value
            evt = ws.cell(row=r, column=col_evt).value
            if nom and evt and str(evt).strip():
                notes.append((str(nom).strip(), str(evt).strip()))
    return notes


# ─────────────────────────────────────────────────────────────
#  LECTURE — feuille "Semaine_N_Agent" (vue par agent = horaires réels)
# ─────────────────────────────────────────────────────────────

def lire_vue_agent(ws):
    """Retourne dict agent -> jour -> {'arrivee':min|None,'depart':min|None,'conge':bool}."""
    result = {}
    max_row = ws.max_row
    max_col = ws.max_column
    header = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    jours_cols = {}
    for c, h in enumerate(header, start=1):
        if isinstance(h, str):
            v = h.strip().upper()
            jour_trouve = next((j for j in JOURS_ORDRE if v.startswith(j)), None)
            if jour_trouve:
                jours_cols[jour_trouve] = c

    current_agent = None
    r = 2
    while r <= max_row:
        colA = ws.cell(row=r, column=1).value
        if colA is None:
            current_agent = None
            r += 1
            continue
        if not is_creneau(colA):
            current_agent = str(colA).strip()
            result.setdefault(current_agent, {j: {'arrivee': None, 'depart': None, 'conge': False}
                                                for j in jours_cols})
            r += 1
            continue
        if current_agent:
            for jour, col in jours_cols.items():
                val = ws.cell(row=r, column=col).value
                if isinstance(val, str):
                    v = val.strip()
                    vlow = normalize(v)
                    if vlow.startswith('arriv'):
                        h = parse_heure_texte(v)
                        if h is not None:
                            # Certain·es agent·es (pause flexible) ont 2 lignes
                            # "Arrivée" le même jour (avant/après une pause
                            # variable) : on ne garde ici que la BORNE LA PLUS
                            # TÔT, pour ne pas déclencher de fausse alerte sur
                            # le retour de pause. La pause elle-même reste
                            # couverte par la règle "pause déjeuner" (§ R4).
                            prev = result[current_agent][jour]['arrivee']
                            result[current_agent][jour]['arrivee'] = h if prev is None else min(prev, h)
                    elif vlow.startswith('conge'):
                        result[current_agent][jour]['conge'] = True
                    elif vlow.startswith('depar'):
                        h = parse_heure_texte(v)
                        if h is not None:
                            prev = result[current_agent][jour]['depart']
                            result[current_agent][jour]['depart'] = h if prev is None else max(prev, h)
        r += 1
    return result


# ─────────────────────────────────────────────────────────────
#  CONSTRUCTION DES OCCURRENCES PAR AGENT / JOUR
# ─────────────────────────────────────────────────────────────

def construire_occurrences_jour(jour_data, agents_connus):
    occ = defaultdict(list)
    for cren in jour_data['creneaux']:
        cs, ce = cren['debut'], cren['fin']
        for label, val in (('RDC', cren['rdc']), ('Adulte', cren['adulte']), ('M & F', cren['mf'])):
            if val:
                occ[val].append({'debut': cs, 'fin': ce, 'type': label, 'detail': label})
        for val in cren['jeunesse']:
            if val:
                occ[val].append({'debut': cs, 'fin': ce, 'type': 'Jeunesse', 'detail': 'Jeunesse'})
        if cren['accueil']:
            for a, d, f, seg in extraire_occurrences_multiples(cren['accueil'], cs, ce, agents_connus):
                occ[a].append({'debut': d, 'fin': f, 'type': 'Accueil/Animation', 'detail': seg})
        if cren['reunion']:
            for a, d, f, seg in extraire_occurrences_multiples(cren['reunion'], cs, ce, agents_connus):
                occ[a].append({'debut': d, 'fin': f, 'type': 'Réunion', 'detail': seg})
        if cren['absence']:
            for a, d, f, seg in extraire_occurrences_multiples(cren['absence'], cs, ce, agents_connus):
                occ[a].append({'debut': d, 'fin': f, 'type': 'Absence', 'detail': seg})
    return occ


def fusionner_occurrences(liste):
    groups = defaultdict(list)
    for o in liste:
        groups[(o['type'], o['detail'])].append(o)
    fusion = []
    for items in groups.values():
        items.sort(key=lambda x: x['debut'])
        cur = None
        for it in items:
            if cur is None:
                cur = dict(it)
            elif it['debut'] <= cur['fin']:
                cur['fin'] = max(cur['fin'], it['fin'])
            else:
                fusion.append(cur)
                cur = dict(it)
        if cur:
            fusion.append(cur)
    return fusion


# ─────────────────────────────────────────────────────────────
#  RÈGLES DE VÉRIFICATION
# ─────────────────────────────────────────────────────────────

def verifier_jour(jour_data, semaine_label, semaine_num, vue_agent, agents_connus, prep, anomalies):
    jour = jour_data['jour']
    jour_cap = JOUR_CAPITALISE.get(jour, jour.capitalize())
    date_str = jour_data.get('date_str')
    occ_brutes = construire_occurrences_jour(jour_data, agents_connus)

    # bornes d'ouverture approximatives ce jour = 1er début / dernière fin des créneaux
    if jour_data['creneaux']:
        ouverture_debut = jour_data['creneaux'][0]['debut']
        ouverture_fin = jour_data['creneaux'][-1]['fin']
    else:
        ouverture_debut = ouverture_fin = None

    mode_complet = bool(prep and 'horaires_agents' in prep)
    horaires_agents = prep.get('horaires_agents', {}) if prep else {}
    pause_flex = prep.get('pause_flex', set()) if prep else set()
    affectations = prep.get('affectations', {}) if prep else {}
    presences_vac = prep.get('params', {}).get('presences_vac', {}) if prep else {}
    roulement_type = prep.get('roulement_type', {}) if prep else {}
    roulement_exceptions = prep.get('roulement_exceptions', {}) if prep else {}
    samedis_couleur = prep.get('params', {}).get('samedis', {}) if prep else {}

    for agent, liste in occ_brutes.items():
        if est_ignore(agent):
            continue
        occs = fusionner_occurrences(liste)
        occs_travail = [o for o in occs if o['type'] != 'Absence']

        # R8 — Eloïse ne doit jamais apparaître
        if est_eloise(agent):
            for o in occs:
                anomalies.append(Anomalie(
                    'rouge', semaine_label, jour,
                    f"Eloïse apparaît dans le planning ({o['type']}, {fmt_min(o['debut'])}-{fmt_min(o['fin'])}) "
                    f"— elle ne doit jamais être affectée.",
                    'Eloïse jamais planifiée'))
            continue

        # R1 + R4 — horaires contractuels ET pause déjeuner, en un seul
        # contrôle certain (🔴) si les onglets de préparation sont
        # disponibles : on réutilise directement agent_disponible(), la
        # fonction que le moteur de calcul utilise lui-même pour décider si
        # un agent peut être placé sur un créneau. Même règle, même vérité.
        if mode_complet and not est_vacataire(agent):
            h = horaires_agents.get(agent, {}).get(jour_cap)
            if h is None:
                for o in occs_travail:
                    anomalies.append(Anomalie(
                        'jaune', semaine_label, jour,
                        f"{agent} est planifié·e ({o['type']}, {fmt_min(o['debut'])}-{fmt_min(o['fin'])}) "
                        f"mais aucun horaire n'est défini pour {agent} ce jour dans Horaires_Des_Agents "
                        f"— agent normalement absent ce jour-là ?",
                        'Horaires contractuels'))
            else:
                for o in occs_travail:
                    if not agent_disponible(agent, jour_cap, o['debut'], o['fin'], horaires_agents,
                                             [], date_str, pause_flex):
                        anomalies.append(Anomalie(
                            'rouge', semaine_label, jour,
                            f"{agent} est indiqué·e en {o['type']} de {fmt_min(o['debut'])} à {fmt_min(o['fin'])}, "
                            f"ce qui sort de son horaire contractuel ce jour-là ou empiète sur sa pause "
                            f"déjeuner obligatoire.",
                            'Horaires contractuels / pause déjeuner'))
        elif mode_complet and est_vacataire(agent):
            # Vacataire : présence définie par le tableau "Présence Vacataire"
            # du Paramètres (prioritaire), sinon par Horaires_Des_Agents.
            for o in occs_travail:
                if not agent_disponible(agent, jour_cap, o['debut'], o['fin'], horaires_agents,
                                         [], date_str, pause_flex, presences_vac):
                    anomalies.append(Anomalie(
                        'rouge', semaine_label, jour,
                        f"{agent} (vacataire) est indiqué·e de {fmt_min(o['debut'])} à {fmt_min(o['fin'])}, "
                        f"en dehors de sa présence prévue ce jour-là (tableau Présence Vacataire / horaires).",
                        'Présence vacataire'))
        else:
            # Mode dégradé (pas d'onglets de préparation dans ce fichier) :
            # on se rabat sur la 'vue par agent' — moins précis, notamment
            # pour la pause déjeuner (cf. limites documentées).
            info_h = vue_agent.get(agent, {}).get(jour, {})
            arrivee, depart = info_h.get('arrivee'), info_h.get('depart')
            for o in occs_travail:
                if arrivee is not None and o['debut'] < arrivee:
                    anomalies.append(Anomalie(
                        'rouge', semaine_label, jour,
                        f"{agent} est indiqué·e en {o['type']} dès {fmt_min(o['debut'])}, "
                        f"mais son horaire indique une arrivée à {fmt_min(arrivee)}.",
                        'Horaires contractuels'))
                if depart is not None and o['fin'] > depart:
                    anomalies.append(Anomalie(
                        'rouge', semaine_label, jour,
                        f"{agent} est indiqué·e en {o['type']} jusqu'à {fmt_min(o['fin'])}, "
                        f"mais son horaire indique un départ à {fmt_min(depart)}.",
                        'Horaires contractuels'))
            if agent not in PAUSE_EXEMPTS and not est_vacataire(agent):
                pres_debut = arrivee if arrivee is not None else ouverture_debut
                pres_fin = depart if depart is not None else ouverture_fin
                if pres_debut is not None and pres_fin is not None:
                    fen_debut = max(pres_debut, PAUSE_FENETRE[0])
                    fen_fin = min(pres_fin, PAUSE_FENETRE[1])
                    if fen_fin - fen_debut >= PAUSE_MIN_LIBRE:
                        segs = sorted(
                            [(max(o['debut'], fen_debut), min(o['fin'], fen_fin))
                             for o in occs_travail if o['debut'] < fen_fin and o['fin'] > fen_debut],
                            key=lambda x: x[0])
                        libre_max = 0
                        curseur = fen_debut
                        for d, f in segs:
                            if d > curseur:
                                libre_max = max(libre_max, d - curseur)
                            curseur = max(curseur, f)
                        libre_max = max(libre_max, fen_fin - curseur)
                        if libre_max < PAUSE_MIN_LIBRE:
                            anomalies.append(Anomalie(
                                'jaune', semaine_label, jour,
                                f"{agent} ne semble pas avoir au moins 1h vraiment libre entre 12h et 14h "
                                f"(sur la plage {fmt_min(fen_debut)}-{fmt_min(fen_fin)} où il/elle est présent·e). "
                                f"À vérifier — peut être normal si son contrat prévoit une présence continue. "
                                f"(Vérification approximative : les onglets de préparation ne sont pas présents "
                                f"dans ce fichier.)",
                                'Pause déjeuner'))

        # R2/R3 — chevauchements (y compris congé/absence vs travail)
        for i in range(len(occs)):
            for j in range(i + 1, len(occs)):
                a, b = occs[i], occs[j]
                if a['debut'] < b['fin'] and b['debut'] < a['fin']:
                    if a['type'] == 'Absence' or b['type'] == 'Absence':
                        autre = b if a['type'] == 'Absence' else a
                        anomalies.append(Anomalie(
                            'rouge', semaine_label, jour,
                            f"{agent} est en congé/absence mais apparaît aussi en {autre['type']} "
                            f"({autre['detail']}) de {fmt_min(autre['debut'])} à {fmt_min(autre['fin'])}.",
                            'Congé = jamais planifié'))
                    else:
                        anomalies.append(Anomalie(
                            'rouge', semaine_label, jour,
                            f"{agent} est indiqué·e en {a['type']} ({a['detail']}) de "
                            f"{fmt_min(a['debut'])} à {fmt_min(a['fin'])} ET en {b['type']} ({b['detail']}) "
                            f"de {fmt_min(b['debut'])} à {fmt_min(b['fin'])} — ces deux horaires se chevauchent.",
                            'Un agent à un seul endroit à la fois'))

        # R5 — habilitations (Affectations si disponible, sinon liste codée en dur)
        table_habilitations = affectations if mode_complet and affectations else HABILITATIONS
        if not est_vacataire(agent) and agent in table_habilitations:
            sections_ok = {canon_section(x) for x in table_habilitations[agent]}
            for o in occs_travail:
                if o['type'] in ('RDC', 'Adulte', 'M & F', 'Jeunesse') and canon_section(o['type']) not in sections_ok:
                    anomalies.append(Anomalie(
                        'rouge', semaine_label, jour,
                        f"{agent} est affecté·e en {o['type']} de {fmt_min(o['debut'])} à {fmt_min(o['fin'])}, "
                        f"section non habilitée (habilitations : {', '.join(table_habilitations[agent])}).",
                        'Habilitations par section'))
        elif not est_vacataire(agent) and agent not in table_habilitations:
            anomalies.append(Anomalie(
                'jaune', semaine_label, jour,
                f"'{agent}' n'est pas reconnu·e dans la liste habituelle des agents — vérifier l'orthographe "
                f"ou une éventuelle nouvelle recrue non encore répertoriée.",
                'Agent inconnu'))

        # R6 — vacataires jamais au RDC
        if est_vacataire(agent):
            for o in occs_travail:
                if o['type'] == 'RDC':
                    anomalies.append(Anomalie(
                        'rouge', semaine_label, jour,
                        f"{agent} (vacataire) est affecté·e au RDC de {fmt_min(o['debut'])} à {fmt_min(o['fin'])} "
                        f"— un vacataire ne doit jamais être au RDC.",
                        'Vacataires jamais au RDC'))

        # R9 — roulement samedi Bleu/Rouge (nécessite les onglets de préparation)
        if (mode_complet and jour == 'SAMEDI' and jour_data.get('samedi_type')
                and agent in roulement_type and occs_travail):
            couleur_effective = roulement_exceptions.get(semaine_num, {}).get(agent, roulement_type[agent])
            if couleur_effective != jour_data['samedi_type']:
                anomalies.append(Anomalie(
                    'rouge', semaine_label, jour,
                    f"{agent} est planifié·e ce samedi {jour_data['samedi_type'].lower()}, mais son roulement "
                    f"(éventuelles exceptions incluses) l'affecte au samedi {couleur_effective.lower() if couleur_effective else '?'}.",
                    'Roulement samedi'))

    # R10 — couverture RDC / Adulte / M&F / Jeunesse par rapport au planning
    # type — CONTRAINTE DURE (demande utilisatrice 09/2026) : à chaque
    # créneau, le nombre d'agent·es affecté·es dans chaque section doit
    # correspondre EXACTEMENT à ce que prévoit le planning type :
    #   - RDC / Adulte / M & F : toujours calé sur le planning type (que la
    #     semaine soit "vacances scolaires" ou non — seule la Jeunesse
    #     change de référence pendant les vacances, cf. moteur de calcul).
    #   - Jeunesse : le planning type hors vacances scolaires, ou le nombre
    #     donné par l'onglet Besoins_Jeunesse pendant les vacances scolaires
    #     (le jour effectif "vacances" vient du réglage Semaine_N, sauf
    #     override ponctuel via Jours_speciaux si l'onglet est présent).
    # Un agent manquant par rapport au planning type = 🔴 "trou". Un agent en
    # trop par rapport au planning type = 🔴 aussi (le moteur de calcul ne
    # dépasse jamais ce nombre, donc un dépassement en main est une anomalie
    # réelle, pas juste une préférence).
    planning_type = prep.get('planning_type') if prep else None
    besoins_jeunesse_data = prep.get('besoins_jeunesse') if prep else None
    if planning_type:
        periode_semaine = (prep.get('params', {}) or {}).get('semaines', {}).get(
            semaine_num, 'Hors Vacances scolaires')
        js_info = (prep.get('jours_speciaux', {}) or {}).get(date_str) if date_str else None
        periode_effective = 'Vacances Scolaires' if (js_info and js_info.get('vacances')) else periode_semaine
        est_vacances = 'Hors' not in str(periode_effective)

        if jour == 'SAMEDI' and jour_data.get('samedi_type'):
            pt_jour_key = f"Samedi_{jour_data['samedi_type']}"
        else:
            pt_jour_key = jour_cap
        pt_jour = planning_type.get(pt_jour_key, {})

        pt_blocs = []
        for cren_str, sections_agents in pt_jour.items():
            parsed = parse_creneau_engine(cren_str)
            if parsed:
                pt_blocs.append((parsed[0], parsed[1], sections_agents))

        def _pt_agents(section, cs, ce):
            """Agents prévus par le planning type pour ce créneau/section.
            None si aucun bloc du PT ne couvre ce créneau (pas de contrainte)."""
            for bcs, bce, sections_agents in pt_blocs:
                if cs >= bcs and ce <= bce:
                    return [a for a in sections_agents.get(section, []) if a and a.strip()]
            return None

        # Besoins Jeunesse (uniquement utile en période de vacances scolaires)
        besoins_jour = {}
        if est_vacances and besoins_jeunesse_data:
            periode_key = next((k for k in besoins_jeunesse_data if 'Hors' not in k), None)
            jour_key_besoin = jour_cap
            if jour == 'SAMEDI' and jour_data.get('samedi_type'):
                def _norm(s):
                    return s.lower().replace('_', ' ').replace('-', ' ').strip()
                cible = _norm(f"samedi {jour_data['samedi_type']}")
                jours_dispo = besoins_jeunesse_data.get(periode_key, {}) if periode_key else {}
                jour_key_besoin = next((k for k in jours_dispo if _norm(k) == cible), None)
            if periode_key and jour_key_besoin:
                besoins_jour = besoins_jeunesse_data.get(periode_key, {}).get(jour_key_besoin, {})
        besoins_ranges = []
        for cren_str, besoin in besoins_jour.items():
            parsed = parse_creneau_engine(cren_str)
            if parsed:
                besoins_ranges.append((parsed[0], parsed[1], besoin))

        for cren in jour_data['creneaux']:
            cs, ce = cren['debut'], cren['fin']

            # RDC / Adulte / M&F
            for section, champ_label, val in (
                ('RDC', 'RDC', cren['rdc']),
                ('Adulte', 'Adulte', cren['adulte']),
                ('MF', 'M & F', cren['mf']),
            ):
                pt_agents = _pt_agents(section, cs, ce)
                if pt_agents is None:
                    continue  # créneau hors planning type : pas de contrainte vérifiable
                requis = len(pt_agents)
                present = 1 if val else 0
                if requis > present:
                    anomalies.append(Anomalie(
                        'rouge', semaine_label, jour,
                        f"{champ_label} {fmt_min(cs)}-{fmt_min(ce)} : aucun·e agent·e affecté·e alors que "
                        f"le planning type y prévoit {', '.join(pt_agents)} — trou par rapport au planning type.",
                        'Couverture planning type'))
                elif requis == 0 and present > 0:
                    anomalies.append(Anomalie(
                        'rouge', semaine_label, jour,
                        f"{champ_label} {fmt_min(cs)}-{fmt_min(ce)} : {val} est affecté·e alors que le "
                        f"planning type ne prévoit personne dans cette section à ce créneau.",
                        'Couverture planning type'))

            # Jeunesse
            jeunesse_presents = [a for a in cren['jeunesse'] if a and not est_ignore(a)]
            if est_vacances:
                sous_tranches = [b for (bcs, bce, b) in besoins_ranges if bcs >= cs and bce <= ce]
                if sous_tranches:
                    requis_j = min(sous_tranches)
                else:
                    cren_str_exact = f'{cs//60:02d}:{cs%60:02d}-{ce//60:02d}:{ce%60:02d}'
                    requis_j = besoins_jour.get(cren_str_exact, 0)
                reference = 'les besoins Jeunesse en période de vacances scolaires (onglet Besoins_Jeunesse)'
            else:
                pt_agents_j = _pt_agents('Jeunesse', cs, ce)
                requis_j = len(pt_agents_j) if pt_agents_j is not None else None
                reference = 'le planning type'

            if requis_j is not None:
                if len(jeunesse_presents) < requis_j:
                    anomalies.append(Anomalie(
                        'rouge', semaine_label, jour,
                        f"Jeunesse {fmt_min(cs)}-{fmt_min(ce)} : {len(jeunesse_presents)} agent(s) affecté(s) "
                        f"({', '.join(jeunesse_presents) or 'aucun'}) alors que {reference} en prévoit "
                        f"{requis_j} — trou en Jeunesse.",
                        'Couverture Jeunesse'))
                elif len(jeunesse_presents) > requis_j:
                    anomalies.append(Anomalie(
                        'rouge', semaine_label, jour,
                        f"Jeunesse {fmt_min(cs)}-{fmt_min(ce)} : {len(jeunesse_presents)} agent(s) affecté(s) "
                        f"({', '.join(jeunesse_presents)}) alors que {reference} n'en prévoit que {requis_j}.",
                        'Couverture Jeunesse'))

    # R7 — vacataire seul en Jeunesse hors 12h-14h
    for cren in jour_data['creneaux']:
        jeunesse_agents = [a for a in cren['jeunesse'] if a and not est_ignore(a)]
        if jeunesse_agents and all(est_vacataire(a) for a in jeunesse_agents):
            if not (cren['debut'] >= PAUSE_FENETRE[0] and cren['fin'] <= PAUSE_FENETRE[1]):
                anomalies.append(Anomalie(
                    'rouge', semaine_label, jour,
                    f"Jeunesse {fmt_min(cren['debut'])}-{fmt_min(cren['fin'])} : uniquement des vacataires "
                    f"({', '.join(jeunesse_agents)}) — autorisé seulement sur 12h-14h.",
                    'Vacataire seul en Jeunesse'))


# ─────────────────────────────────────────────────────────────
#  FONCTION PRINCIPALE
# ─────────────────────────────────────────────────────────────

def verifier_planning(file_bytes):
    """file_bytes : bytes du classeur Excel du planning déjà rempli.
    Retourne une liste d'Anomalie."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    anomalies = []

    prep = charger_donnees_preparation(wb)
    if prep is None:
        anomalies.append(Anomalie(
            'jaune', '', '',
            "Ce fichier ne contient pas les onglets de préparation (Paramètres, Horaires_Des_Agents, "
            "Affectations, Roulement_Samedi) — probablement généré avec une version antérieure de l'outil. "
            "La vérification se fait donc en mode approximatif (habilitations et horaires partiellement "
            "devinés, pause déjeuner incertaine, roulement samedi et présence vacataire non vérifiables). "
            "Régénérez le planning avec la version à jour pour une vérification complète.",
            'Mode dégradé'))
        prep = {}
    elif 'erreur_lecture' in prep:
        anomalies.append(Anomalie(
            'jaune', '', '',
            f"Les onglets de préparation sont présents mais n'ont pas pu être lus correctement "
            f"({prep['erreur_lecture']}) — vérification en mode approximatif.",
            'Mode dégradé'))
        prep = {}
    elif prep.get('manquants'):
        anomalies.append(Anomalie(
            'jaune', '', '',
            f"Onglet(s) de préparation manquant(s) dans ce fichier : {', '.join(prep['manquants'])} "
            f"— les vérifications correspondantes sont faites en mode approximatif ou ignorées.",
            'Mode dégradé partiel'))

    semaine_sheets = sorted(
        [n for n in wb.sheetnames if re.match(r'^Semaine_\d+$', n, re.IGNORECASE)],
        key=lambda n: int(re.search(r'\d+', n).group())
    )

    for sn in semaine_sheets:
        ws = wb[sn]
        semaine_num = int(re.search(r'\d+', sn).group())
        agent_sheet_name = _trouver_onglet_insensible_casse(wb, f"{sn}_Agent")
        vue_agent = {}
        if agent_sheet_name is not None:
            vue_agent = lire_vue_agent(wb[agent_sheet_name])
        elif not prep.get('horaires_agents'):
            anomalies.append(Anomalie(
                'jaune', sn, '',
                f"L'onglet '{sn}_Agent' est introuvable : les horaires contractuels "
                f"(règle 'arrivée/départ') n'ont pas pu être vérifiés pour cette semaine.",
                'Fichier incomplet'))

        jours = lire_jours_semaine(ws)
        for jour_data in jours:
            verifier_jour(jour_data, sn, semaine_num, vue_agent, ALL_AGENTS_CONNUS, prep, anomalies)

            # Garde-fou : rien trouvé (ni H/I/J, ni notes W-Z) ce jour-là
            rien_dans_grille = all(
                not c['accueil'] and not c['reunion'] and not c['absence']
                for c in jour_data['creneaux']
            )
            notes = lire_notes_agents_jour(ws, jour_data['row_debut_data'], jour_data['row_fin_data'])
            if rien_dans_grille and not notes:
                anomalies.append(Anomalie(
                    'jaune', sn, jour_data['jour'],
                    f"Aucun événement noté ce jour (ni dans les colonnes Accueil/Animation/Réunion/Absence, "
                    f"ni dans les notes agents W-Z) — à vérifier si c'est normal.",
                    'Garde-fou : rien trouvé'))

            # Cohérence notes agents (W-Z) <-> ce qui apparaît dans H/I/J
            for agent, texte in notes:
                if est_ignore(agent):
                    continue
                fragment = texte.split(' ', 1)[-1] if re.match(r'^\d{1,2}h', texte) else texte
                fragment_norm = normalize(fragment)[:20]
                trouve = False
                for c in jour_data['creneaux']:
                    for champ in (c['accueil'], c['reunion'], c['absence']):
                        if champ and fragment_norm and fragment_norm in normalize(champ):
                            trouve = True
                if not trouve and fragment_norm:
                    anomalies.append(Anomalie(
                        'jaune', sn, jour_data['jour'],
                        f"Note ajoutée par {agent} (« {texte} ») ne semble pas se retrouver dans le planning "
                        f"(colonnes Accueil/Animation, Réunion ou Absence) — à vérifier manuellement.",
                        'Note non répercutée'))

    return anomalies


# ─────────────────────────────────────────────────────────────
#  AFFICHAGE (utilisable directement en dehors de Streamlit)
# ─────────────────────────────────────────────────────────────

def resumer(anomalies):
    n_rouge = sum(1 for a in anomalies if a.gravite == 'rouge')
    n_jaune = sum(1 for a in anomalies if a.gravite == 'jaune')
    return n_rouge, n_jaune


if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else 'Planning_Semaine1_avec_notes_agents.xlsx'
    with open(path, 'rb') as f:
        data = f.read()
    anomalies = verifier_planning(data)
    n_rouge, n_jaune = resumer(anomalies)
    print(f"{len(anomalies)} anomalies détectées ({n_rouge} rouge, {n_jaune} jaune)\n")
    for a in anomalies:
        marqueur = '🔴' if a.gravite == 'rouge' else '🟡'
        print(f"{marqueur} [{a.semaine} — {a.jour}] {a.message}")

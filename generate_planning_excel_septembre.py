"""
Génère le planning Excel de septembre 2026 (format visuel identique au modèle mai)
à partir du moteur CP-SAT (planning_engine_cpsat.py), avec récap d'heures de
service public dynamique (formules Excel) par semaine.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.formula import ArrayFormula
from copy import copy
import re

from planning_engine_cpsat import (
    compute_full_planning, load_excel_data, parse_jours_speciaux,
    parse_horaires_agents_grille, ONGLET_HORAIRES_GRILLE,
    parse_evenements, parse_horaires_ouverture, hhmm_to_min, parse_affectations,
    parse_parametres, parse_horaires_agents
)

INPUT_PREP = '/mnt/user-data/uploads/SEPTEMBRE2026_Preparation_Planning_Mediatheque.xlsx'
OUTPUT_PATH = '/mnt/user-data/outputs/Planning_Septembre_2026_CPSAT.xlsx'
MOIS_LABEL = 'Septembre 2026'

# ── Couleurs / styles (repris du modèle fourni) ─────────────────────
COL_TITLE_FILL = 'FF2C3E50'
COL_TITLE_FONT = 'FFFFFFFF'
COL_DAY_FILL = 'FF34495E'
COL_DAY_FONT = 'FFFFFFFF'
COL_SAMEDI_BLEU_FILL = 'FFAED6F1'   # bandeau Samedi Bleu (08/2026)
COL_SAMEDI_ROUGE_FILL = 'FFE6534A'  # bandeau Samedi Rouge (08/2026)
COL_FERIE_FILL = 'FFF9EBEA'
COL_FERIE_MSG_FILL = 'FFFADBD8'
COL_HEADER_CREN = 'FFCCCCCC'
HEADERS = ['Créneau', 'RDC', 'Adulte', 'M & F', 'Jeunesse 1', 'Jeunesse 2', 'Jeunesse 3',
           'Accueil / Animation', 'Réunion', 'Absence']
# 1/ (demande utilisatrice 08/2026) : même fond gris que "Créneau" pour TOUS
# les en-têtes de colonne — on n'affiche plus une couleur différente par
# intitulé de section.
HEADER_FILLS = {h: 'FFCCCCCC' for h in HEADERS}
DATA_FILLS_OPEN = {
    'Créneau': 'FFF8F9FA',
    # RDC/Adulte/M&F/Jeunesse 1-3 : non utilisées (fond déterminé par
    # l'agent, cf. agent_cell_style) — gardées pour compatibilité.
    'RDC': 'FFD6E8F7', 'Adulte': 'FFD4EDD4', 'M & F': 'FFFFF0CC',
    'Jeunesse 1': 'FFFFE0E0', 'Jeunesse 2': 'FFFFE0E0', 'Jeunesse 3': 'FFFFE0E0',
    'Accueil / Animation': 'FFE8E8F2', 'Réunion': 'FFE8E8F2', 'Absence': 'FFE8E8F2',
}
DATA_FILLS_CLOSED = {k: 'FFF5F5F5' for k in HEADERS}
DATA_FILLS_CLOSED['Créneau'] = 'FFF5F5F5'
for k in HEADERS[1:]:
    DATA_FILLS_CLOSED[k] = 'FFEBEBEB'

# 2/ (demande utilisatrice 08/2026) : Jeunesse pouvait accueillir jusqu'à 3
# agents sur un même créneau → 3 colonnes dédiées (une par agent), plus
# lisible que plusieurs couleurs superposées dans une seule case. En
# contrepartie, Accueil et Animation sont fusionnés en une seule colonne
# (moins de colonnes au total). Une alerte "Jeunesse" (besoin non couvert)
# s'affiche sur les 3 colonnes, faute de savoir laquelle en particulier.
SECTION_COL = {'RDC': 'RDC', 'Adulte': 'Adulte', 'MF': 'M & F',
               'Jeunesse': ['Jeunesse 1', 'Jeunesse 2', 'Jeunesse 3']}
COL_WIDTHS = [14, 18, 18, 18, 13, 13, 13, 22, 18, 22]
ALERT_BORDER = Border(*[Side(style='thick', color='FFE74C3C')] * 4)
# Bordure grise fine (08/2026, demande utilisatrice) sur toutes les cellules du
# planning, pour mieux séparer visuellement les créneaux/colonnes.
THIN_GREY = Side(style='thin', color='FFBFBFBF')
GREY_BORDER = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

# ── Surlignage jaune — événement sans agent ou sans horaire (09/2026,
# demande utilisatrice) : même jaune que celui déjà utilisé dans
# sources_to_evenements.py (bloc 1), pour rester cohérent visuellement.
JAUNE_EVENEMENT_INCOMPLET = PatternFill('solid', fgColor='FFFFFF00')


def evenement_incomplet(ev):
    """Vrai si l'événement n'a ni agent, ni horaire renseigné — sauf les
    événements 'libre' (Accueil libre crèche/école), où l'absence d'agent est
    normale et ne doit jamais être signalée (même règle qu'au bloc 1)."""
    if 'libre' in ev.get('nom', '').lower():
        return False
    sans_agent = not ev.get('agents')
    sans_horaire = ev.get('cs') is None or ev.get('ce') is None
    return sans_agent or sans_horaire

# ── Couleur de police par agent (08/2026, demande utilisatrice) ────────────
# Conservée pour le bandeau nom d'agent (vue par agent) et comme couleur de
# repli pour les agents non couverts par la capture d'écran d'Elo.
AGENT_COLORS = {
    'Marie-France':    'C00000',  # Rouge foncé
    'Anne-Françoise':  'FF0000',  # Rouge
    'Christine':       'ED7D31',  # Orange
    'Léa':             'BF9000',  # Jaune foncé / Or
    'Chloé':           '548235',  # Vert
    'Macha':           '00B050',  # Vert clair
    'Delphine':        '006100',  # Vert foncé
    'Barbara':         '008080',  # Sarcelle
    'Stéphane':        '00B0F0',  # Bleu clair
    'Stéphanie':       '0070C0',  # Bleu
    'Robin':           '002060',  # Bleu foncé
    'Guillaume':       '7030A0',  # Violet
    'Agnès':           'C00090',  # Magenta
    'Tiphaine':        '833C00',  # Marron
    'Vacataire 1':     '808080',  # Gris
    'Vacataire 2':     '404040',  # Gris foncé
    'Vacataire 3':     '44546A',  # Bleu-gris
}

# ── Fond par agent — ESSAI 08/2026, couleurs reprises TELLES QUELLES de la
# capture d'écran fournie par Elo (référence de couleurs qu'elle utilise déjà
# à la main). Certains agents de son tableau (Virginie, Lydie, Gabriel, David)
# ne font pas partie de l'équipe actuelle et ne sont pas repris. À l'inverse,
# Léa et Barbara n'apparaissaient pas dans sa capture : couleurs provisoires
# choisies pour rester dans le même esprit, à valider avec elle.
AGENT_FILL_COLORS = {
    'Delphine':        '23E8B7',  # vert d'eau (capture)
    'Stéphanie':       'ED7D31',  # orange (capture)
    'Christine':       'FF3F3F',  # rouge (capture)
    'Guillaume':       '92D050',  # vert clair (capture)
    'Macha':           'C65911',  # brun/orange (capture)
    'Stéphane':        'FCE4D6',  # pêche très clair (capture)
    'Tiphaine':        '7030A0',  # violet (capture)
    'Chloé':           'FFF2CC',  # crème (capture)
    'Robin':           'BF8F00',  # or foncé (capture)
    'Anne-Françoise':  'FF99FF',  # rose (capture)
    'Marie-France':    '0070C0',  # bleu (capture)
    'Agnès':           '66FF33',  # vert vif (capture)
    # ⚠️ Provisoires (absents de la capture) — à valider avec Elo :
    'Léa':             'FFD966',  # or clair
    'Barbara':         'B4A7D6',  # mauve clair
    # Vacataires : gris neutres (inchangé, cohérent avec AGENT_COLORS).
    # Clara-Jade et Vacataire Marie (vues dans le fichier de référence,
    # absentes des Affectations) : même gris que les autres vacataires,
    # demande explicite d'Elo — pas besoin de les distinguer entre elles.
    'Vacataire 1':     '808080',
    'Vacataire 2':     '404040',
    'Vacataire 3':     '44546A',
    'Clara-Jade':      '808080',
    'Vacataire Marie': '808080',
}


def _luminance(hexcolor):
    r = int(hexcolor[0:2], 16)
    g = int(hexcolor[2:4], 16)
    b = int(hexcolor[4:6], 16)
    return 0.299 * r + 0.587 * g + 0.114 * b


def _texte_lisible(hexcolor):
    """Noir sur fond clair, blanc sur fond foncé — même logique que la
    capture d'écran d'Elo (elle alterne déjà noir/blanc selon la couleur)."""
    return '000000' if _luminance(hexcolor) > 150 else 'FFFFFF'


def is_vacataire(agent):
    return 'vacataire' in str(agent).lower()


def agent_cell_style(names):
    """Retourne (PatternFill, couleur_texte_hex) pour une cellule RDC/Adulte/
    M&F/Jeunesse X contenant 0 ou 1 agent (depuis l'ESSAI 08/2026 : chaque
    colonne Jeunesse 1/2/3 ne porte plus qu'un seul agent, donc plus besoin
    de rayures). Le cas 2+ agents est conservé par sécurité (ne devrait plus
    se produire avec la nouvelle structure à 3 colonnes)."""
    if not names:
        return PatternFill('solid', fgColor='FFF8F9FA'), '000000'
    cols = ['FF' + AGENT_FILL_COLORS.get(n, '808080') for n in names]
    if len(cols) == 1:
        return PatternFill('solid', fgColor=cols[0]), _texte_lisible(cols[0][2:])
    return PatternFill(patternType='lightHorizontal', fgColor=cols[0], bgColor=cols[1]), '000000'


# ── Récap heures dynamique ──────────────────────────────────────────
# Colonnes qui comptent comme "service public" (même périmètre que ce que le
# moteur compare au planning-type) : RDC, Adulte, M&F, Jeunesse 1-3 -> colonnes
# B, C, D, E, F, G. Accueil/Animation/Réunion/Absence ne comptent PAS.
RECAP_SECTION_COLS = ['B', 'C', 'D', 'E', 'F', 'G']
# Colonnes techniques cachées : copie systématique de B-G, ligne par ligne,
# JAMAIS fusionnées — sert de source au récap d'heures dynamique.
RECAP_SOURCE_COLS = {'B': 'L', 'C': 'M', 'D': 'N', 'E': 'O', 'F': 'P', 'G': 'Q'}
# Idem pour Accueil/Animation (H) et Réunion (I) : copies stables, utilisées
# par la vue par agent pour afficher les événements même si H-I sont fusionnées.
EVENT_SOURCE_COLS = {'H': 'R', 'I': 'S'}
# Colonnes supplémentaires cachées T/U : versions sans prénom des mêmes
# événements (vue par agent uniquement) — pas une simple copie de H/I, donc
# en dehors du mapping EVENT_SOURCE_COLS, mais à cacher comme les autres.
EVENT_SOURCE_COLS_SANS_NOMS = ['T', 'U']
COL_DUREE = 'K'  # colonne cachée : durée du créneau en heures (formule)
COL_RECAP_FILL = 'FFEFEFEF'
COL_RECAP_HEADER_FILL = 'FF2C3E50'
JOURS_DISCRETS = {'Mardi', 'Jeudi', 'Vendredi'}  # créneaux fermés affichés en discret


def classer_evenement(nom):
    """Devine la colonne (Accueil / Animation fusionnées, ou Réunion) selon
    le nom de l'événement — même logique de mots-clés qu'avant, juste
    regroupée en 2 catégories au lieu de 3 (ESSAI 08/2026, demande
    utilisatrice : une seule colonne 'Accueil / Animation')."""
    n = nom.lower()
    if 'réunion' in n or 'reunion' in n:
        return 'Réunion'
    return 'Accueil / Animation'  # accueil, portage, caj, et tout le reste (programmation...)


def is_open_fixed(jour, cs, ce, hor_ouv):
    """Fusionne les plages adjacentes pour éviter les faux '—' (ex: Mercredi 12h-13h)."""
    ranges = sorted(hor_ouv.get(jour, []))
    if not ranges:
        return False
    merged = [list(ranges[0])]
    for s, e in ranges[1:]:
        if s <= merged[-1][1]:
            merged[-1][1] = max(merged[-1][1], e)
        else:
            merged.append([s, e])
    return any(cs >= s and ce <= e for s, e in merged)


def fmt_agents(lst):
    return ' / '.join(lst) if lst else None


def rich_agents(lst):
    """Construit le texte enrichi (une couleur de police par agent, cf.
    AGENT_COLORS, en gras) pour une cellule RDC/Adulte/M&F/Jeunesse pouvant
    contenir plusieurs agents séparés par ' / '. Retourne None si vide."""
    if not lst:
        return None
    blocks = []
    for i, n in enumerate(lst):
        if i > 0:
            blocks.append(' / ')
        color = AGENT_COLORS.get(n, '000000')
        blocks.append(TextBlock(InlineFont(sz=10, color=color, b=True), n))
    return CellRichText(*blocks)


def fmt_hhmm(m):
    h, mn = divmod(int(m), 60)
    return f'{h}h{mn:02d}' if mn else f'{h}h'


def label_evenement(ev, cs, ce):
    """Nom de l'événement, avec les précisions entre parenthèses : agents
    concernés et/ou horaires exacts si l'événement ne correspond pas pile au
    créneau affiché (ex: "Bébés se livrent (10h15-10h45)" dans le créneau
    10h-11h). Documenté dans le contexte projet (§ règles d'affichage)."""
    nom = ev['nom']
    agents_ev = ev.get('agents', [])
    precisions = list(agents_ev)
    if ev['cs'] != cs or ev['ce'] != ce:
        precisions.append(f"{fmt_hhmm(ev['cs'])}-{fmt_hhmm(ev['ce'])}")
    return f"{nom} ({', '.join(precisions)})" if precisions else nom


def label_evenement_sans_noms(ev, cs, ce):
    """Version SANS AUCUN prénom (demande utilisatrice) — utilisée uniquement
    dans la vue par agent : juste le nom de l'événement, et l'horaire exact
    entre parenthèses si l'événement ne correspond pas pile au créneau
    affiché (ex: "Réunion pôle (10h-11h30)"). Jamais de prénom, ni celui de
    l'agent concerné ni ceux des autres — on est déjà dans le tableau de
    l'agent, les prénoms n'apportent rien."""
    nom = ev['nom']
    if ev['cs'] != cs or ev['ce'] != ce:
        return f"{nom} ({fmt_hhmm(ev['cs'])}-{fmt_hhmm(ev['ce'])})"
    return nom


def write_row(ws, r, values, fills, bold=False, font_size=10, alert_headers=None,
               alert_msgs=None, discret=False, agent_fill_cols=None,
               alerte_jaune_headers=None, alerte_jaune_msgs=None):
    alert_headers = alert_headers or set()
    alert_msgs = alert_msgs or {}
    # Surlignage jaune "événement incomplet" (09/2026) : indépendant des
    # alertes de couverture (bordure rouge) ci-dessus — les deux peuvent
    # coexister sur une même cellule si les deux cas se présentent.
    alerte_jaune_headers = alerte_jaune_headers or set()
    alerte_jaune_msgs = alerte_jaune_msgs or {}
    # Colonnes (ex: 'RDC') où le fond ET la couleur de texte sont déterminés
    # par le(s) agent(s) présent(s), plutôt que par la section (ESSAI 08/2026,
    # demande utilisatrice — couleurs reprises de sa capture d'écran).
    agent_fill_cols = agent_fill_cols or set()
    for ci, (h, val) in enumerate(zip(HEADERS, values), start=1):
        is_agent_col = h in agent_fill_cols and isinstance(val, list)
        write_val = fmt_agents(val) if is_agent_col else val
        cell = ws.cell(row=r, column=ci, value=write_val)
        if is_agent_col:
            fill, text_color = agent_cell_style(val)
            cell.fill = fill
            cell.font = Font(size=font_size, bold=True, color='FF' + text_color)
        elif h in alerte_jaune_headers:
            cell.fill = JAUNE_EVENEMENT_INCOMPLET
        else:
            cell.fill = PatternFill('solid', fgColor=fills[h])
        cell.border = GREY_BORDER
        if discret:
            # Créneaux sans service public (mar/jeu/ven) : discret, gris, italique
            cell.font = Font(size=8, italic=True, color='FF999999')
        elif not is_agent_col:
            italic = h in ('Accueil / Animation', 'Réunion', 'Absence')
            # Essai (08/2026) : taille 9 dédiée pour les colonnes d'événements
            # (Accueil/Animation/Réunion/Absence), au lieu de suivre font_size.
            size = 9 if italic else font_size
            cell.font = Font(size=size, bold=bold, italic=italic)
        cell.alignment = Alignment(horizontal='center' if ci > 1 else 'left',
                                    vertical='center', wrap_text=True)
        if h in alert_headers:
            cell.border = Border(*[Side(style='thick', color='FFE74C3C')] * 4)
            if h in alert_msgs:
                from openpyxl.comments import Comment
                cell.comment = Comment('ALERTE : ' + alert_msgs[h], 'Moteur CP-SAT')
        if h in alerte_jaune_headers and h in alerte_jaune_msgs:
            from openpyxl.comments import Comment
            # Si une bordure d'alerte structurelle a déjà posé un commentaire,
            # on ne l'écrase pas : on l'étend plutôt.
            texte = 'INCOMPLET : ' + alerte_jaune_msgs[h]
            if cell.comment:
                cell.comment.text += '\n' + texte
            else:
                cell.comment = Comment(texte, 'Moteur CP-SAT')


def fusionner_cellules_identiques(ws, lignes, valeurs_brutes, colonnes=range(2, 10),
                                    hidden_map=None):
    """Fusionne verticalement les cellules consécutives ayant le même contenu
    visible, colonne par colonne, sur l'ensemble des lignes d'une même journée.
    Ignore les valeurs vides/'—' (rien à fusionner). N'affecte que l'affichage :
    voir RECAP_SOURCE_COLS pour le calcul d'heures, qui reste indépendant.
    `valeurs_brutes[(row, col)]` donne le texte PLAT (pas le texte enrichi) pour
    décider des runs identiques — plus fiable qu'une comparaison sur les objets
    de texte enrichi eux-mêmes.

    `hidden_map` (optionnel) : {col_visible: lettre_colonne_cachée}. Pour CES
    colonnes, écrit aussi la formule de la colonne cachée en fonction du
    résultat RÉEL de la fusion (⚠️ 08/2026 — corrige un bug : l'ancienne
    formule "si vide, recopier la ligne du dessus" supposait que toute cellule
    vide était une continuation de fusion, ce qui est vrai pour RDC/Adulte/M&F/
    Jeunesse (toujours remplies sur un créneau ouvert) mais FAUX pour Accueil/
    Animation/Réunion (vides la plupart du temps) : un événement se retrouvait
    recopié sur toutes les heures suivantes jusqu'au prochain événement. Toutes
    les lignes d'une même fusion référencent désormais directement la ligne du
    HAUT de la fusion — plus de supposition, que la fusion existe ou non.
    Les colonnes SANS fusion (B-E depuis 08/2026) utilisent aussi ce mécanisme
    quand `hidden_map` les couvre : référence directe à soi-même."""
    hidden_map = hidden_map or {}
    for col in colonnes:
        i = 0
        while i < len(lignes):
            val = valeurs_brutes.get((lignes[i], col))
            j = i
            while (j + 1 < len(lignes)
                   and valeurs_brutes.get((lignes[j + 1], col)) == val):
                j += 1
            fusionne = val not in (None, '—') and j > i
            if fusionne:
                ws.merge_cells(start_row=lignes[i], start_column=col,
                                end_row=lignes[j], end_column=col)
                top = ws.cell(row=lignes[i], column=col)
                top.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)
                for rr in range(lignes[i], lignes[j] + 1):
                    ws.cell(row=rr, column=col).border = GREY_BORDER
            if col in hidden_map:
                col_letter = get_column_letter(col)
                hcol = hidden_map[col]
                top_row = lignes[i]
                for k in range(i, j + 1):
                    ws[f'{hcol}{lignes[k]}'] = f'={col_letter}{top_row}'
            i = j + 1


# ── Zone de notes agents (09/2026, demande utilisatrice) ────────────
# Petit tableau "Nom / Événement" à droite de CHAQUE journée (pas un seul
# tableau pour toute la semaine) : les agents notent au fil de l'eau ce qui
# change (réunion, absence, accueil...), et la note remonte automatiquement
# dans la bonne colonne du planning (H=Accueil/Animation, I=Réunion,
# J=Absence), au bon créneau horaire si l'heure est précisée en premier.
NOTES_HEADER_LABEL = "Événements à ajouter (ex. format : 14h-15h Accueil classe)"
NOTES_NOM1_COL, NOTES_NOTE1_COL = 23, 24   # W, X
NOTES_NOM2_COL, NOTES_NOTE2_COL = 25, 26   # Y, Z
NOTES_HELPER_START = {NOTES_NOTE1_COL: 28, NOTES_NOTE2_COL: 33}  # X->AB.., Z->AG..
NOTES_H_COL, NOTES_I_COL, NOTES_J_COL = 8, 9, 10
NOTES_R_COL, NOTES_S_COL, NOTES_T_COL, NOTES_U_COL = 18, 19, 20, 21
NOTES_CAT_BY_COL = {NOTES_H_COL: 3, NOTES_I_COL: 1, NOTES_J_COL: 2}


def _notes_ft(ref):
    return f'LEFT({ref},IFERROR(FIND(" ",{ref})-1,LEN({ref})))'


def _notes_convert(expr):
    hpos = f'FIND("h",{expr})'
    hour = f'VALUE(LEFT({expr},{hpos}-1))'
    mintxt = f'MID({expr},{hpos}+1,LEN({expr})-{hpos})'
    minute = f'IF({mintxt}="",0,VALUE({mintxt}))'
    return f'({hour}+{minute}/60)'


def _notes_esc(s):
    return '' if s is None else str(s).replace('"', '""')


def _notes_rng(col, r1, r2):
    return f'{get_column_letter(col)}${r1}:{get_column_letter(col)}${r2}'


def _notes_blocks_for(merges_by_col, col, first_cren, last_cren):
    """Regroupe les lignes d'une journée en blocs "écrivables" : soit une
    fusion existante (plusieurs lignes -> une seule case visible), soit une
    ligne isolée (non fusionnée) qui référence sa propre ligne. Reflète les
    fusions RÉELLES (posées par fusionner_cellules_identiques), pas une
    supposition sur le contenu."""
    day_merges = [(s, e) for (s, e) in merges_by_col.get(col, [])
                  if s >= first_cren and e <= last_cren]
    merged_rows = set()
    for s, e in day_merges:
        merged_rows.update(range(s, e + 1))
    blocks = list(day_merges)
    for rr in range(first_cren, last_cren + 1):
        if rr not in merged_rows:
            blocks.append((rr, rr))
    return blocks


def ajouter_zone_notes_jour(ws, header_row, first_cren, last_cren, agents_14):
    """Ajoute, pour UNE journée, le petit tableau Nom/Événement (2 groupes de
    colonnes W/X et Y/Z), ses colonnes cachées d'analyse, et fait remonter
    automatiquement les notes vers H/I/J — ainsi que vers les colonnes
    cachées R/S/T/U utilisées par la vue par agent (Accueil/Animation et
    Réunion uniquement ; l'Absence de la vue par agent est calculée à part,
    directement depuis les événements du fichier de préparation, donc une
    note d'absence tapée ici ne remonte QUE dans le planning principal, pas
    dans la vue par agent — limite connue, documentée dans le contexte
    projet)."""
    mid = (len(agents_14) + 1) // 2
    group1, group2 = agents_14[:mid], agents_14[mid:]
    thin = Side(style='thin', color='FFBFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    groups = [(NOTES_NOM1_COL, NOTES_NOTE1_COL, group1),
              (NOTES_NOM2_COL, NOTES_NOTE2_COL, group2)]

    # ---- 1) tableau visible : noms pré-remplis (fond agent), case note vide
    for nom_col, note_col, group in groups:
        h1 = ws.cell(row=header_row, column=nom_col, value='Nom')
        h2 = ws.cell(row=header_row, column=note_col, value=NOTES_HEADER_LABEL)
        for hc in (h1, h2):
            hc.font = Font(bold=True, size=9)
            hc.fill = PatternFill('solid', fgColor='FFCCCCCC')
            hc.border = border
        for i, name in enumerate(group):
            rr = header_row + 1 + i
            fond_hex = AGENT_FILL_COLORS.get(name, 'F4F4F4')
            texte_hex = _texte_lisible(fond_hex)
            nc = ws.cell(row=rr, column=nom_col, value=name)
            nc.font = Font(bold=True, size=9, color='FF' + texte_hex)
            nc.fill = PatternFill('solid', fgColor='FF' + fond_hex)
            nc.border = border
            nc.alignment = Alignment(vertical='center')
            ec = ws.cell(row=rr, column=note_col)
            ec.border = border
            ec.font = Font(size=9)
            ec.alignment = Alignment(wrap_text=True, vertical='top')
        ws.column_dimensions[get_column_letter(nom_col)].width = 14
        ws.column_dimensions[get_column_letter(note_col)].width = 26

    # ---- 2) colonnes cachées d'analyse (catégorie / texte / début / fin)
    for nom_col, note_col, group in groups:
        hstart = NOTES_HELPER_START[note_col]
        cat_col, txt_col, deb_col, fin_col = hstart, hstart + 1, hstart + 2, hstart + 3
        for i in range(len(group)):
            rr = header_row + 1 + i
            ref = f'{get_column_letter(note_col)}{rr}'
            ft = _notes_ft(ref)
            is_horaire = f'ISNUMBER(SEARCH("h",{ft}))'
            has_dash = f'ISNUMBER(FIND("-",{ft}))'
            start_part = f'LEFT({ft},FIND("-",{ft})-1)'
            end_part = f'MID({ft},FIND("-",{ft})+1,50)'

            ws.cell(row=rr, column=cat_col, value=(
                f'=IF({ref}="",0,'
                f'IF(OR(ISNUMBER(SEARCH("réunion",{ref})),ISNUMBER(SEARCH("reunion",{ref})),ISNUMBER(SEARCH("rdv",{ref}))),1,'
                f'IF(OR(ISNUMBER(SEARCH("congé",{ref})),ISNUMBER(SEARCH("conge",{ref})),ISNUMBER(SEARCH("absen",{ref})),ISNUMBER(SEARCH("part",{ref}))),2,'
                f'3)))'
            ))
            ws.cell(row=rr, column=txt_col, value=(
                f'=IF({ref}="","",'
                f'IF({is_horaire},TRIM(MID({ref},LEN({ft})+2,300)),TRIM({ref})))'
            ))
            ws.cell(row=rr, column=deb_col, value=(
                f'=IF({ref}="",0,'
                f'IF(NOT({is_horaire}),0,'
                f'IF({has_dash},{_notes_convert(start_part)},{_notes_convert(ft)})))'
            ))
            ws.cell(row=rr, column=fin_col, value=(
                f'=IF({ref}="",0,'
                f'IF(NOT({is_horaire}),24,'
                f'IF({has_dash},{_notes_convert(end_part)},{_notes_convert(ft)}+0.0167)))'
            ))
        for c in (cat_col, txt_col, deb_col, fin_col):
            ws.column_dimensions[get_column_letter(c)].hidden = True

    def group_new_part(cat_code, block_start_row, block_end_row, name_included):
        parts = []
        for nom_col, note_col, group in groups:
            hstart = NOTES_HELPER_START[note_col]
            cat_col, txt_col, deb_col, fin_col = hstart, hstart + 1, hstart + 2, hstart + 3
            r1, r2 = header_row + 1, header_row + len(group)
            cat_r = _notes_rng(cat_col, r1, r2)
            txt_r = _notes_rng(txt_col, r1, r2)
            deb_r = _notes_rng(deb_col, r1, r2)
            fin_r = _notes_rng(fin_col, r1, r2)
            nom_r = _notes_rng(nom_col, r1, r2)
            row_start = f'(TIMEVALUE(LEFT(A{block_start_row},5))*24)'
            row_end = f'(TIMEVALUE(MID(A{block_end_row},7,5))*24)'
            cond = f'(({cat_r}={cat_code})*({fin_r}>{row_start})*({deb_r}<{row_end}))'
            item = f'{txt_r}&" ("&{nom_r}&")"' if name_included else f'{txt_r}'
            parts.append(f'_xlfn.TEXTJOIN("; ",TRUE,IF({cond},{item},""))')
        a, b = parts
        return f'IF(({a})="",({b}),IF(({b})="",({a}),({a})&"; "&({b})))'

    # ---- 3) fusions réellement posées sur H/I/J pour CETTE journée
    merges_by_col = {NOTES_H_COL: [], NOTES_I_COL: [], NOTES_J_COL: []}
    for mc in list(ws.merged_cells.ranges):
        if mc.min_col in merges_by_col:
            merges_by_col[mc.min_col].append((mc.min_row, mc.max_row))

    # ---- 4) cascade dans H / I / J (texte déjà généré + nouvelles notes)
    for vcol in (NOTES_H_COL, NOTES_I_COL, NOTES_J_COL):
        cat_code = NOTES_CAT_BY_COL[vcol]
        for (bs, be) in _notes_blocks_for(merges_by_col, vcol, first_cren, last_cren):
            anchor = ws.cell(row=bs, column=vcol)
            baked_literal = f'"{_notes_esc(anchor.value)}"' if anchor.value else '""'
            npf = group_new_part(cat_code, bs, be, name_included=True)
            formula = (
                f'=TRIM({baked_literal}&IF(({npf})="","",'
                f'IF({baked_literal}="","","; ")&({npf})))'
            )
            anchor.value = ArrayFormula(ref=anchor.coordinate, text=formula)

    # ---- 5) T (Accueil/Animation sans prénom) aux ancres de H
    for (bs, be) in _notes_blocks_for(merges_by_col, NOTES_H_COL, first_cren, last_cren):
        t_cell = ws.cell(row=bs, column=NOTES_T_COL)
        baked_literal = f'"{_notes_esc(t_cell.value)}"' if t_cell.value else '""'
        npf = group_new_part(3, bs, be, name_included=False)
        formula = (
            f'=TRIM({baked_literal}&IF(({npf})="","",'
            f'IF({baked_literal}="","","; ")&({npf})))'
        )
        t_cell.value = ArrayFormula(ref=t_cell.coordinate, text=formula)

    # ---- 6) U (Réunion sans prénom) aux ancres de I
    for (bs, be) in _notes_blocks_for(merges_by_col, NOTES_I_COL, first_cren, last_cren):
        u_cell = ws.cell(row=bs, column=NOTES_U_COL)
        baked_literal = f'"{_notes_esc(u_cell.value)}"' if u_cell.value else '""'
        npf = group_new_part(1, bs, be, name_included=False)
        formula = (
            f'=TRIM({baked_literal}&IF(({npf})="","",'
            f'IF({baked_literal}="","","; ")&({npf})))'
        )
        u_cell.value = ArrayFormula(ref=u_cell.coordinate, text=formula)

    # ---- 7) R (miroir de H) et S (miroir de I), ligne par ligne, chacune
    #      référençant l'ancre RÉELLE de son bloc (corrige un défaut de
    #      fusionner_cellules_identiques : les lignes vides consécutives, non
    #      fusionnées visuellement, étaient quand même regroupées comme si
    #      elles l'étaient, cf. contexte projet).
    h_blocks = _notes_blocks_for(merges_by_col, NOTES_H_COL, first_cren, last_cren)
    i_blocks = _notes_blocks_for(merges_by_col, NOTES_I_COL, first_cren, last_cren)
    for rr in range(first_cren, last_cren + 1):
        h_anchor = next(bs for (bs, be) in h_blocks if bs <= rr <= be)
        i_anchor = next(bs for (bs, be) in i_blocks if bs <= rr <= be)
        ws.cell(row=rr, column=NOTES_R_COL, value=f'=H{h_anchor}')
        ws.cell(row=rr, column=NOTES_S_COL, value=f'=I{i_anchor}')


ONGLETS_PREP_A_EMBARQUER = [
    'Paramètres', 'Horaires_Des_Agents', 'Affectations', 'Roulement_Samedi',
    'Besoins_Jeunesse', 'Jours_speciaux',
    # 'Planning_type' n'est pas ici : recopié à part, visible et verrouillé
    # (cf. embarquer_planning_type_visible dans ce même fichier).
]


def embarquer_onglets_preparation(wb, input_path):
    """Copie, en 'très masqué' (invisibles pour un agent qui ouvre le fichier —
    même le clic droit > Afficher ne les propose pas), les onglets de
    préparation utiles à la vérification ultérieure du planning (bloc 3 :
    horaires contractuels exacts, habilitations, roulement samedi...).

    L'onglet Événements n'est volontairement PAS copié (demande utilisatrice
    09/2026) : une fois le planning ajusté à la main, il n'est plus à jour —
    la référence devient les colonnes H/I/J et les notes agents (W-Z) du
    planning généré lui-même, pas le fichier de préparation d'origine."""
    wb_src = openpyxl.load_workbook(input_path, data_only=True)
    for nom in ONGLETS_PREP_A_EMBARQUER:
        if nom not in wb_src.sheetnames:
            continue
        ws_src = wb_src[nom]
        nom_cible = nom if nom not in wb.sheetnames else f'_prep_{nom}'
        ws_dst = wb.create_sheet(nom_cible)
        for row in ws_src.iter_rows(values_only=True):
            ws_dst.append(row)
        ws_dst.sheet_state = 'veryHidden'


def verrouiller_cellules_formules(wb):
    """Verrouille TOUTES les cellules contenant une formule, sur tous les
    onglets, dès la génération (demande utilisatrice 09/2026) : évite les
    écrasements accidentels de formule en tapant directement dans Excel. Les
    cellules SANS formule (dont la nouvelle zone de notes agents, colonne
    Événement) restent éditables. Pas de mot de passe — le but est d'éviter
    les fausses manipulations, pas de bloquer un usage volontaire."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                is_formula = isinstance(cell.value, ArrayFormula) or (
                    isinstance(cell.value, str) and cell.value.startswith('=')
                )
                cell.protection = Protection(locked=is_formula)
        ws.protection.sheet = True
        ws.protection.formatCells = False
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        ws.protection.sort = False
        ws.protection.autoFilter = False


def generer(input_path=None, output_path=None):
    input_path = input_path or INPUT_PREP
    output_path = output_path or OUTPUT_PATH

    raw = load_excel_data(input_path)
    jours_speciaux = parse_jours_speciaux(raw)
    params_mois = parse_parametres(raw)
    evenements = parse_evenements(raw, annee_defaut=params_mois.get('annee'))
    hor_ouv = parse_horaires_ouverture(raw)

    weeks_data, metadata = compute_full_planning(input_path)

    # Liste des agents pour le récap heures : tous les agents habilités
    # (réguliers + vacataires, dans l'ordre du fichier Affectations), hors
    # Eloïse (jamais dans ce tableau, cf. parse_affectations).
    affectations, categories, responsables, pause_flex, priorite_rdc = parse_affectations(raw)
    agents_recap = list(affectations.keys())
    # Lecture directe de la grille collaborative "horaires d'équipes" ; repli
    # sur l'ancienne liste à plat "Horaires_Des_Agents" si le fichier de
    # préparation n'a pas encore été mis à jour avec le nouvel onglet.
    if ONGLET_HORAIRES_GRILLE in raw:
        horaires_agents = parse_horaires_agents_grille(raw)
    else:
        horaires_agents = parse_horaires_agents(raw)
    # Agents réguliers hors vacataires, dans l'ordre du fichier Affectations —
    # sert de base à la zone de notes agents (09/2026) : 2 groupes de colonnes
    # (7 + 7, ou moins si l'équipe est plus petite / plus grande) à côté de
    # chaque journée.
    agents_recap_vue_agent = [a for a in agents_recap if not is_vacataire(a)]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for w in weeks_data:
        week_num = w['week_num']
        jours = w['jours']
        ws = wb.create_sheet(f'Semaine_{week_num}')
        row_lookup = {}  # (jour, cs, ce) -> numéro de ligne source, pour la vue par agent
        for i, width in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        # 3/ (demande utilisatrice 08/2026) : colonne A (Créneau) figée, pour
        # rester visible en défilant horizontalement dans le planning.
        ws.freeze_panes = 'B1'

        # Titre
        first_date = jours[0]['date']
        last_date = jours[-1]['date']
        d1 = int(first_date[-2:])
        d2 = int(last_date[-2:])
        a1 = int(first_date[:4])
        a2 = int(last_date[:4])
        mois_jour_fr = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',6:'Juin',
                         7:'Juillet',8:'Août',9:'Septembre',10:'Octobre',
                         11:'Novembre',12:'Décembre'}
        m1 = mois_jour_fr[int(first_date[5:7])]
        m2 = mois_jour_fr[int(last_date[5:7])]
        periode_txt = metadata.get('periode_semaine', {}).get(week_num, '')
        if m1 == m2 and a1 == a2:
            titre = f'PLANNING SP — Semaine {week_num}  |  {d1} au {d2} {m1} {a1}'
        else:
            titre = f'PLANNING SP — Semaine {week_num}  |  {d1} {m1} {a1} au {d2} {m2} {a2}'
        ws.merge_cells('A1:J1')
        c = ws.cell(row=1, column=1, value=titre)
        c.fill = PatternFill('solid', fgColor=COL_TITLE_FILL)
        c.font = Font(size=13, bold=True, color=COL_TITLE_FONT)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 28

        ws.merge_cells('A2:J2')
        c2 = ws.cell(row=2, column=1,
                      value='  Bordure rouge épaisse = ALERTE (besoin non entièrement couvert, voir commentaire de la cellule)')
        c2.font = Font(size=9, italic=True, color='FFE74C3C')
        c2.alignment = Alignment(horizontal='left', vertical='center')

        r = 3
        for j in jours:
            date_str = j['date']
            jour = j['jour']
            dnum = int(date_str[-2:])
            sam_type = j.get('sam_type')
            js_info = jours_speciaux.get(date_str, {})
            est_ferie = js_info.get('ferie', False)

            mois_jour_fr = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',6:'Juin',
                             7:'Juillet',8:'Août',9:'Septembre',10:'Octobre',
                             11:'Novembre',12:'Décembre'}[int(date_str[5:7])]
            annee_jour = int(date_str[:4])
            libelle_jour = f'  {jour.upper()}  {dnum} {mois_jour_fr} {annee_jour}'
            if sam_type:
                libelle_jour += f'  —  SAMEDI {sam_type}'
            if est_ferie:
                libelle_jour += '  —  JOUR FÉRIÉ'

            ws.merge_cells(f'A{r}:J{r}')
            c = ws.cell(row=r, column=1, value=libelle_jour)
            if sam_type == 'BLEU':
                bandeau_fill = COL_SAMEDI_BLEU_FILL
                bandeau_font_color = 'FF1B4F72'  # texte foncé lisible sur fond clair
            elif sam_type == 'ROUGE':
                bandeau_fill = COL_SAMEDI_ROUGE_FILL
                bandeau_font_color = COL_DAY_FONT
            else:
                bandeau_fill = COL_FERIE_FILL if est_ferie else COL_DAY_FILL
                bandeau_font_color = 'FF000000' if est_ferie else COL_DAY_FONT
            c.fill = PatternFill('solid', fgColor=bandeau_fill)
            c.font = Font(size=12, bold=True, color=bandeau_font_color)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.row_dimensions[r].height = 22
            r += 1

            if est_ferie:
                ws.merge_cells(f'A{r}:J{r}')
                c = ws.cell(row=r, column=1, value='🎉  Médiathèque fermée — Jour Férié')
                c.fill = PatternFill('solid', fgColor=COL_FERIE_MSG_FILL)
                c.font = Font(size=10)
                c.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[r].height = 18
                r += 1
                r += 1  # ligne vide
                continue

            # En-têtes de colonnes
            header_row = r
            write_row(ws, r, HEADERS, HEADER_FILLS, bold=True, font_size=9)
            # (demande utilisatrice 08/2026) : les 3 colonnes Jeunesse
            # restent séparées dans le CONTENU (une ligne par créneau, pas de
            # fusion verticale même si le même agent enchaîne plusieurs
            # créneaux) — seule la ligne d'EN-TÊTE fusionne les 3 cases en
            # une seule case "Jeunesse", pour indiquer que ce sont 3
            # sous-colonnes d'une même section.
            ws.cell(row=r, column=5, value='Jeunesse')
            ws.cell(row=r, column=6, value=None)
            ws.cell(row=r, column=7, value=None)
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
            ws.cell(row=r, column=5).alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[r].height = 14
            r += 1

            creneaux = j['creneaux']
            solution = j['solution']
            alertes_jour = j.get('alertes', [])
            lignes_jour = []  # lignes de créneaux (hors titre/en-tête) pour la fusion
            valeurs_brutes = {}  # (row, col) -> texte plat, pour décider des fusions
            for c_idx, (cs, ce) in enumerate(creneaux):
                cren_str = f'{cs//60:02d}:{cs%60:02d}-{ce//60:02d}:{ce%60:02d}'
                ouvert = is_open_fixed(jour, cs, ce, hor_ouv)
                sol_c = solution[c_idx] if solution else {}
                alertes_ici = [(sec, msg) for (ci2, sec, msg) in alertes_jour if ci2 == c_idx]
                # Jeunesse peut correspondre à 3 en-têtes (Jeunesse 1/2/3) —
                # on aplatit la correspondance section -> en-tête(s).
                alert_headers = set()
                alert_msgs = {}
                for sec, msg in alertes_ici:
                    mapped = SECTION_COL.get(sec, sec)
                    headers_sec = mapped if isinstance(mapped, list) else [mapped]
                    for h in headers_sec:
                        alert_headers.add(h)
                        alert_msgs[h] = msg

                # Événements chevauchant ce créneau
                accueil_animation = reunion = None
                accueil_animation_sn = reunion_sn = None  # versions SANS prénoms (vue par agent)
                absence = []
                # Surlignage jaune (09/2026, demande utilisatrice) : un
                # événement Accueil/Animation ou Réunion sans agent ni
                # horaire renseigné est signalé directement sur la cellule
                # du planning final, pas seulement dans l'onglet Événements
                # source (bloc 1) — plus facile à repérer d'un coup d'œil.
                accueil_incomplet = reunion_incomplet = False
                accueil_incomplet_msg = reunion_incomplet_msg = None
                for ev in evenements:
                    if ev['date'] != date_str:
                        continue
                    if not (cs < ev['ce'] and ce > ev['cs']):
                        continue
                    nom = ev['nom']
                    agents_ev = ev.get('agents', [])
                    if nom.strip().lower() == 'congé':
                        absence.extend(agents_ev)
                        continue
                    label = label_evenement(ev, cs, ce)
                    label_sn = label_evenement_sans_noms(ev, cs, ce)
                    cat = classer_evenement(nom)
                    incomplet = evenement_incomplet(ev)
                    if incomplet:
                        raisons = []
                        if not ev.get('agents'):
                            raisons.append('aucun agent renseigné')
                        if ev.get('cs') is None or ev.get('ce') is None:
                            raisons.append('horaire non renseigné')
                        msg_incomplet = f"« {nom} » : {', '.join(raisons)}"
                    if cat == 'Réunion':
                        reunion, reunion_sn = label, label_sn
                        reunion_incomplet = incomplet
                        if incomplet:
                            reunion_incomplet_msg = msg_incomplet
                    else:
                        accueil_animation, accueil_animation_sn = label, label_sn
                        accueil_incomplet = incomplet
                        if incomplet:
                            accueil_incomplet_msg = msg_incomplet
                absence_txt = f"congé ({', '.join(sorted(set(absence)))})" if absence else None

                if not ouvert:
                    rdc_l = adulte_l = mf_l = jeun1_l = jeun2_l = jeun3_l = []
                    values = [cren_str, '—', '—', '—', '—', '—', '—', None, None, absence_txt]
                    write_row(ws, r, values, DATA_FILLS_CLOSED,
                              discret=(jour in JOURS_DISCRETS))
                    rdc = adulte = mf = jeun1 = jeun2 = jeun3 = '—'
                else:
                    rdc_l = sol_c.get('RDC', [])
                    adulte_l = sol_c.get('Adulte', [])
                    mf_l = sol_c.get('MF', [])
                    jeun_l = sol_c.get('Jeunesse', [])
                    # 3 colonnes Jeunesse : un agent par colonne (ESSAI 08/2026).
                    jeun1_l = jeun_l[0:1]
                    jeun2_l = jeun_l[1:2]
                    jeun3_l = jeun_l[2:3]
                    values = [cren_str, rdc_l, adulte_l, mf_l, jeun1_l, jeun2_l, jeun3_l,
                              accueil_animation, reunion, absence_txt]
                    alerte_jaune_headers = set()
                    alerte_jaune_msgs = {}
                    if accueil_incomplet:
                        alerte_jaune_headers.add('Accueil / Animation')
                        alerte_jaune_msgs['Accueil / Animation'] = accueil_incomplet_msg
                    if reunion_incomplet:
                        alerte_jaune_headers.add('Réunion')
                        alerte_jaune_msgs['Réunion'] = reunion_incomplet_msg
                    write_row(ws, r, values, DATA_FILLS_OPEN,
                              alert_headers=alert_headers, alert_msgs=alert_msgs,
                              agent_fill_cols={'RDC', 'Adulte', 'M & F',
                                               'Jeunesse 1', 'Jeunesse 2', 'Jeunesse 3'},
                              alerte_jaune_headers=alerte_jaune_headers,
                              alerte_jaune_msgs=alerte_jaune_msgs)
                    rdc, adulte, mf = fmt_agents(rdc_l), fmt_agents(adulte_l), fmt_agents(mf_l)
                    jeun1, jeun2, jeun3 = fmt_agents(jeun1_l), fmt_agents(jeun2_l), fmt_agents(jeun3_l)
                # Texte plat (pour la fusion) : colonnes B-G depuis rdc/adulte/mf/jeun1-3,
                # H-J depuis les valeurs déjà écrites (accueil_animation/réunion/absence).
                valeurs_brutes[(r, 2)] = rdc
                valeurs_brutes[(r, 3)] = adulte
                valeurs_brutes[(r, 4)] = mf
                valeurs_brutes[(r, 5)] = jeun1
                valeurs_brutes[(r, 6)] = jeun2
                valeurs_brutes[(r, 7)] = jeun3
                valeurs_brutes[(r, 8)] = accueil_animation if ouvert else None
                valeurs_brutes[(r, 9)] = reunion if ouvert else None
                valeurs_brutes[(r, 10)] = absence_txt
                # Colonne cachée : durée du créneau en heures, calculée depuis le
                # texte "HH:MM-HH:MM" de la colonne A. IFERROR->0 pour les lignes
                # qui ne sont pas des créneaux (titres, en-têtes). Sert au récap
                # d'heures dynamique en bas de la feuille.
                ws[f'{COL_DUREE}{r}'] = f'=IFERROR((TIMEVALUE(MID(A{r},7,5))-TIMEVALUE(LEFT(A{r},5)))*24,0)'
                # Colonnes techniques cachées L-S : valeur par défaut = référence
                # directe à la cellule visible correspondante. Réécrites juste
                # après par fusionner_cellules_identiques() pour les colonnes
                # fusionnées (référence au haut de la fusion) — donc correctes
                # que la cellule soit fusionnée ou non, sans supposition sur ce
                # que signifie une cellule vide (cf. bug corrigé §13.15).
                ws[f'L{r}'] = f'=B{r}'
                ws[f'M{r}'] = f'=C{r}'
                ws[f'N{r}'] = f'=D{r}'
                ws[f'O{r}'] = f'=E{r}'
                ws[f'P{r}'] = f'=F{r}'
                ws[f'Q{r}'] = f'=G{r}'
                ws[f'R{r}'] = f'=H{r}'
                ws[f'S{r}'] = f'=I{r}'
                # Colonnes cachées T/U : versions SANS AUCUN prénom des
                # événements Accueil/Animation et Réunion (demande
                # utilisatrice), utilisées uniquement par la vue par agent —
                # valeur écrite directement (pas une formule miroir de H/I,
                # puisque le texte diffère : jamais de prénom ici).
                ws[f'T{r}'] = accueil_animation_sn if ouvert else None
                ws[f'U{r}'] = reunion_sn if ouvert else None
                ws.row_dimensions[r].height = 20
                lignes_jour.append(r)
                row_lookup[(jour, cs, ce)] = r
                r += 1

            fusionner_cellules_identiques(ws, lignes_jour, valeurs_brutes, colonnes=range(8, 11),
                                           hidden_map={8: 'R', 9: 'S'})
            if lignes_jour:
                ajouter_zone_notes_jour(ws, header_row, lignes_jour[0], lignes_jour[-1],
                                         agents_recap_vue_agent)
            r += 1  # ligne vide entre jours

        # ── Récap heures de service public (dynamique) ──────────────────
        # RDC + Adulte + M&F + Jeunesse uniquement (même périmètre que ce
        # que le moteur compare au planning-type) — Accueil/Animation/
        # Réunion/Absence ne comptent pas. Recalcule automatiquement si
        # les cellules du planning sont modifiées à la main dans Excel.
        premiere_ligne_data = 3
        derniere_ligne_data = r - 1  # dernière ligne écrite pour cette semaine
        r += 1  # ligne d'espacement avant le récap

        ws.merge_cells(f'A{r}:J{r}')
        c = ws.cell(row=r, column=1, value='  RÉCAP HEURES DE SERVICE PUBLIC (RDC + Adulte + M&F + Jeunesse)')
        c.fill = PatternFill('solid', fgColor=COL_RECAP_HEADER_FILL)
        c.font = Font(size=11, bold=True, color='FFFFFFFF')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[r].height = 20
        r += 1

        agent_row = ws.cell(row=r, column=1, value='Agent')
        heures_row = ws.cell(row=r, column=2, value='Heures')
        for cell in (agent_row, heures_row):
            cell.fill = PatternFill('solid', fgColor='FFCCCCCC')
            cell.font = Font(size=9, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
        r += 1

        premiere_ligne_recap = r
        for agent in agents_recap:
            ws.cell(row=r, column=1, value=agent).font = Font(size=10)
            # Somme des durées de créneau (col K) où le nom de l'agent apparaît
            # dans l'une des 4 colonnes techniques L/M/N/O (copies stables de
            # B/C/D/E, jamais affectées par la fusion visuelle des cellules).
            termes = '+'.join(
                f'ISNUMBER(SEARCH($A{r},${col}${premiere_ligne_data}:${col}${derniere_ligne_data}))'
                for col in RECAP_SOURCE_COLS.values()
            )
            formule = (f'=SUMPRODUCT(({termes})*'
                       f'${COL_DUREE}${premiere_ligne_data}:${COL_DUREE}${derniere_ligne_data})')
            cell_h = ws.cell(row=r, column=2, value=formule)
            cell_h.font = Font(size=10)
            cell_h.number_format = '0.0" h"'
            for ci in (1, 2):
                ws.cell(row=r, column=ci).fill = PatternFill('solid', fgColor=COL_RECAP_FILL)
            r += 1
        derniere_ligne_recap = r - 1

        ws.cell(row=r, column=1, value='TOTAL').font = Font(size=10, bold=True)
        cell_tot = ws.cell(row=r, column=2,
                            value=f'=SUM(B{premiere_ligne_recap}:B{derniere_ligne_recap})')
        cell_tot.font = Font(size=10, bold=True)
        cell_tot.number_format = '0.0" h"'
        for ci in (1, 2):
            ws.cell(row=r, column=ci).fill = PatternFill('solid', fgColor='FFD9D9D9')
        r += 1

        # Colonnes techniques (durée par créneau + copies B-E et F-H non fusionnées) : cachées
        ws.column_dimensions[COL_DUREE].hidden = True
        for col in list(RECAP_SOURCE_COLS.values()) + list(EVENT_SOURCE_COLS.values()) + EVENT_SOURCE_COLS_SANS_NOMS:
            ws.column_dimensions[col].hidden = True

        # Vacataires exclus de la vue par agent (demande utilisatrice) — ils
        # restent bien présents dans le planning global (Semaine_X) et dans
        # le récap heures ci-dessus.
        generer_vue_agent(wb, week_num, jours, row_lookup, agents_recap_vue_agent,
                           horaires_agents, pause_flex, evenements)

    copier_onglets_preparation_caches(wb, raw)

    verrouiller_cellules_formules(wb)
    # Appelé APRÈS verrouiller_cellules_formules (voir docstring) : sinon ses
    # cellules seraient déverrouillées par la passe générique ci-dessus.
    embarquer_planning_type_visible(wb, raw)
    embarquer_horaires_agents_visible(wb, raw)
    embarquer_parametres_visible(wb, raw)
    embarquer_affectations_visible(wb, raw)
    wb.save(output_path)
    print('Fichier genere:', output_path)
    return output_path, weeks_data, metadata


# ── Onglets de préparation recopiés en "très masqué" (09/2026, demande
# utilisatrice) ──────────────────────────────────────────────────────────
# But : permettre à planning_checker.py (bloc 3, vérification) de relire les
# horaires contractuels, habilitations et roulement samedi EXACTS, sans avoir
# à redéposer le fichier de Préparation en plus du planning généré. Les
# onglets sont recopiés tels quels (valeurs uniquement), puis rendus
# "très masqués" (veryHidden) : invisibles dans Excel, y compris via le menu
# clic droit > Afficher — seul un programme peut les relire.
#
# L'onglet "Événements" n'est volontairement PAS recopié : une fois le
# planning modifié à la main, il n'est plus à jour ; la référence devient les
# colonnes Accueil/Animation/Réunion/Absence (H/I/J) et les notes agents
# (colonnes W-Z) du planning lui-même.
ONGLETS_PREPARATION_A_RECOPIER = [
    'Roulement_Samedi',
    # Ajoutés (09/2026, demande utilisatrice) : nécessaires à planning_checker.py
    # pour vérifier la contrainte dure "nombre d'agents Jeunesse = planning type
    # (hors vacances) / Besoins_Jeunesse (vacances scolaires)", ainsi que la
    # couverture RDC/Adulte/M&F par rapport au planning type.
    # NB : ni 'Planning_type', ni "horaires d'équipes", ni 'Paramètres', ni
    # 'Affectations' ne sont dans cette liste — ils sont recopiés à part, en
    # onglet VISIBLE (cf. embarquer_planning_type_visible /
    # embarquer_horaires_agents_visible / embarquer_parametres_visible /
    # embarquer_affectations_visible), pas en très masqué, car les agents
    # doivent pouvoir les consulter (Planning_type, horaires d'équipes) ou
    # même les modifier en cours de mois (Paramètres, Affectations —
    # demande utilisatrice 08/2026, cf. régénération partielle).
    'Besoins_Jeunesse', 'Jours_speciaux',
]

# Ancien onglet "Horaires_Des_Agents" (liste à plat) : recopié en masqué
# uniquement en repli, pour les fichiers de préparation pas encore migrés
# vers la grille collaborative "horaires d'équipes".
ONGLET_HORAIRES_ANCIEN = 'Horaires_Des_Agents'


def copier_onglets_preparation_caches(wb, raw):
    for nom in ONGLETS_PREPARATION_A_RECOPIER:
        ws_src = raw.get(nom)
        if ws_src is None:
            continue
        ws_dst = wb.create_sheet(f'_prep_{nom}')
        for row in ws_src.iter_rows(values_only=True):
            ws_dst.append(row)
        ws_dst.sheet_state = 'veryHidden'

    # Repli (voir plus haut) : uniquement si le fichier de préparation n'a pas
    # encore l'onglet grille "horaires d'équipes".
    if ONGLET_HORAIRES_GRILLE not in raw:
        ws_src = raw.get(ONGLET_HORAIRES_ANCIEN)
        if ws_src is not None:
            ws_dst = wb.create_sheet(f'_prep_{ONGLET_HORAIRES_ANCIEN}')
            for row in ws_src.iter_rows(values_only=True):
                ws_dst.append(row)
            ws_dst.sheet_state = 'veryHidden'


def _copier_feuille_avec_mise_en_forme(ws_src, ws_dst):
    """Copie une feuille source vers une feuille destination en conservant
    TOUT le formatage d'origine (police, couleurs de remplissage, bordures,
    alignement, format numérique), les cellules fusionnées, les largeurs de
    colonnes et les hauteurs de lignes — pas seulement les valeurs.
    (openpyxl ne sait copier une feuille que dans le MÊME classeur via
    wb.copy_worksheet ; ici la source et la destination sont deux classeurs
    différents, d'où cette copie manuelle cellule par cellule.)"""
    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    for merged_range in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merged_range))

    for col_letter, dim in ws_src.column_dimensions.items():
        if dim.width is not None:
            ws_dst.column_dimensions[col_letter].width = dim.width
        if dim.hidden:
            ws_dst.column_dimensions[col_letter].hidden = dim.hidden

    for row_idx, dim in ws_src.row_dimensions.items():
        if dim.height is not None:
            ws_dst.row_dimensions[row_idx].height = dim.height

    if ws_src.freeze_panes:
        ws_dst.freeze_panes = ws_src.freeze_panes
    if ws_src.sheet_view:
        ws_dst.sheet_view.showGridLines = ws_src.sheet_view.showGridLines


def embarquer_planning_type_visible(wb, raw):
    """Recopie l'onglet Planning_type en onglet VISIBLE (demande utilisatrice
    09/2026) : contrairement aux autres onglets de préparation (masqués, à
    usage interne uniquement), le planning type doit rester consultable par
    tous les agents qui ouvrent le fichier — c'est leur référence pour
    comprendre l'organisation habituelle. Il est en revanche verrouillé
    (comme une vitre : on voit à travers, mais on n'y touche pas) pour éviter
    toute modification accidentelle. Pas de mot de passe, dans le même esprit
    que verrouiller_cellules_formules : l'objectif est d'éviter la fausse
    manipulation, pas de bloquer un usage volontaire (Excel permet d'ôter la
    protection en un clic, sans mot de passe).

    Le formatage d'origine (couleurs, polices, bordures, colonnes fusionnées
    Jeunesse 1/2/3, largeurs de colonnes...) est conservé à l'identique — ce
    n'est pas une recopie de simples valeurs (correctif 09/2026, demande
    utilisatrice : la première version perdait toutes les couleurs).

    IMPORTANT : cette fonction doit être appelée APRÈS
    verrouiller_cellules_formules(wb), sinon cette dernière — qui reparcourt
    tous les onglets du classeur — déverrouillerait les cellules sans formule
    de ce nouvel onglet (comportement voulu partout ailleurs, mais pas ici)."""
    ws_src = raw.get('Planning_type') or raw.get('planning_type')
    if ws_src is None:
        return
    nom_cible = 'Planning_type' if 'Planning_type' not in wb.sheetnames else 'Planning_type (référence)'
    ws_dst = wb.create_sheet(nom_cible)
    _copier_feuille_avec_mise_en_forme(ws_src, ws_dst)
    for row in ws_dst.iter_rows():
        for cell in row:
            # On préserve le style (police/couleurs/bordures) déjà posé par
            # _copier_feuille_avec_mise_en_forme ci-dessus, on ajoute juste
            # le verrouillage par-dessus (la protection est un attribut de
            # cellule indépendant du style visuel, donc ceci ne touche pas
            # aux couleurs).
            cell.protection = Protection(locked=True)
    ws_dst.protection.sheet = True
    ws_dst.protection.formatCells = False
    ws_dst.protection.formatColumns = False
    ws_dst.protection.formatRows = False
    ws_dst.protection.sort = False
    ws_dst.protection.autoFilter = False
    ws_dst.sheet_state = 'visible'


def embarquer_horaires_agents_visible(wb, raw):
    """Recopie l'onglet "horaires d'équipes" (grille collaborative) en onglet
    VISIBLE et MODIFIABLE (demande utilisatrice 08/2026 : Elo doit pouvoir
    corriger un horaire d'agent en cours de mois — ex. changement de
    contrat — directement dans le planning généré, puis relancer une
    régénération partielle (bloc 4) qui relira ces valeurs telles quelles.
    Avant cette date, cet onglet était recopié verrouillé, en lecture seule
    — comme Planning_type ; ce n'est plus le cas, voir
    _embarquer_prep_visible_modifiable ci-dessous, réutilisée ici).

    Le formatage d'origine (couleurs, hachures des demi-journées non
    travaillées, polices, colonnes fusionnées, largeurs...) est conservé à
    l'identique.

    `verrouiller_formules=True` (demande utilisatrice 24/08) : contrairement
    à Paramètres/Affectations (aucune formule), cette grille contient ~105
    formules de totaux jour/semaine — les laisser déverrouillées comme le
    reste exposerait à un écrasement accidentel en tapant à côté d'une case
    de saisie. Seules CES cellules-là sont reverrouillées après coup ; les
    cases de saisie des horaires restent, elles, librement modifiables.

    IMPORTANT : comme pour Planning_type, doit être appelée APRÈS
    verrouiller_cellules_formules(wb) — sinon celle-ci verrouillerait les
    cellules de formule (totaux jour/semaine) de ce nouvel onglet."""
    _embarquer_prep_visible_modifiable(wb, raw, ONGLET_HORAIRES_GRILLE, verrouiller_formules=True)


def _normaliser_dates_visibles(ws):
    """Corrige un défaut d'affichage découvert le 24/08 (demande
    utilisatrice) : une cellule contenant une vraie date Excel, mais dont le
    format numérique inclut l'heure (ex. 'yyyy-mm-dd h:mm:ss', 19
    caractères) tient rarement dans une largeur de colonne raisonnable et
    s'affiche en '#######' — invisible tant que l'onglet était très masqué,
    devenu gênant maintenant qu'il est visible (cas rencontré : tableau
    Présence Vacataire de Paramètres). Uniforme avec le format déjà utilisé
    par Elo dans son fichier de préparation ('d-mmm-yy', ex. '5-sept-26').
    Ne touche PAS aux cellules d'heure pure (datetime.time) ni aux durées
    (timedelta) — seulement aux vraies dates (datetime.date/datetime), donc
    sans effet sur les colonnes d'horaires de la grille "horaires
    d'équipes"."""
    import datetime as _dt
    colonnes_a_elargir = set()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, _dt.date) and not isinstance(cell.value, _dt.time):
                cell.number_format = 'd-mmm-yy'
                colonnes_a_elargir.add(cell.column_letter)
    for col_letter in colonnes_a_elargir:
        dim = ws.column_dimensions[col_letter]
        if dim.width is None or dim.width < 16:
            dim.width = 16


def _embarquer_prep_visible_modifiable(wb, raw, nom_source, verrouiller_formules=False):
    """Recopie un onglet de préparation en onglet VISIBLE et MODIFIABLE
    (demande utilisatrice 08/2026, cas Paramètres/Affectations) : contrairement
    à embarquer_planning_type_visible (qui verrouille TOUTES les cellules —
    'vitre', on regarde sans toucher), ici on protège seulement la STRUCTURE
    de la feuille (pas d'insertion/suppression de ligne ou colonne
    intempestive) mais on laisse les cellules déverrouillées : la personne
    doit pouvoir taper directement dedans en cours de mois (ex. ajouter une
    nouvelle section habilitée pour un agent nouvellement formé) et relancer
    une régénération partielle (bloc 4) qui relira ces valeurs telles
    quelles. Pas de mot de passe, même esprit que verrouiller_cellules_formules :
    on protège des fausses manipulations de structure, pas d'un usage
    volontaire de saisie.

    `verrouiller_formules=True` (demande utilisatrice 24/08) : après avoir
    tout déverrouillé, reverrouille spécifiquement les cellules qui
    contiennent une formule (ex. totaux calculés dans "horaires d'équipes")
    — même logique que verrouiller_cellules_formules(wb), appliquée
    localement à cet onglet puisqu'il est créé après elle. Sans effet pour
    Paramètres/Affectations, qui n'ont aucune formule.

    Le formatage d'origine est conservé (mêmes largeurs de colonnes, etc.),
    hormis la correction d'affichage des dates (voir
    _normaliser_dates_visibles).

    IMPORTANT : comme pour Planning_type, à appeler APRÈS
    verrouiller_cellules_formules(wb) — sinon celle-ci verrouillerait certaines
    cellules par erreur si l'onglet source contenait des formules."""
    ws_src = raw.get(nom_source)
    if ws_src is None:
        return
    nom_cible = nom_source if nom_source not in wb.sheetnames else f'{nom_source} (référence)'
    ws_dst = wb.create_sheet(nom_cible)
    _copier_feuille_avec_mise_en_forme(ws_src, ws_dst)
    for row in ws_dst.iter_rows():
        for cell in row:
            if verrouiller_formules:
                is_formula = isinstance(cell.value, ArrayFormula) or (
                    isinstance(cell.value, str) and cell.value.startswith('=')
                )
                cell.protection = Protection(locked=is_formula)
            else:
                cell.protection = Protection(locked=False)
    _normaliser_dates_visibles(ws_dst)
    ws_dst.protection.sheet = True
    ws_dst.protection.formatCells = False
    ws_dst.protection.formatColumns = False
    ws_dst.protection.formatRows = False
    ws_dst.protection.insertRows = False
    ws_dst.protection.insertColumns = False
    ws_dst.protection.deleteRows = False
    ws_dst.protection.deleteColumns = False
    ws_dst.protection.sort = False
    ws_dst.protection.autoFilter = False
    ws_dst.sheet_state = 'visible'


def embarquer_parametres_visible(wb, raw):
    """Paramètres : voir _embarquer_prep_visible_modifiable — onglet visible,
    structure protégée, contenu librement modifiable (mois/année, créneaux,
    couleurs samedi, présences vacataires...)."""
    _embarquer_prep_visible_modifiable(wb, raw, 'Paramètres')


def embarquer_affectations_visible(wb, raw):
    """Affectations : voir _embarquer_prep_visible_modifiable — onglet visible,
    structure protégée, contenu librement modifiable (sections habilitées par
    agent, catégorie, responsable, pause flexible...). Cas d'usage typique :
    un agent nouvellement formé sur une section en cours de mois, à répercuter
    via une régénération partielle (bloc 4) sur les jours restants."""
    _embarquer_prep_visible_modifiable(wb, raw, 'Affectations')


def grille_fine_commune(jours):
    """Construit la grille horaire la plus fine (union de toutes les bornes de
    créneaux de la semaine, tous jours confondus), pour aligner sur les mêmes
    lignes des jours aux découpages différents (ex: mardi/jeudi/vendredi ont un
    seul bloc 10h-12h30, mercredi/samedi sont découpés heure par heure)."""
    bornes = set()
    for j in jours:
        for cs, ce in j['creneaux']:
            bornes.add(cs)
            bornes.add(ce)
    bornes = sorted(bornes)
    return list(zip(bornes[:-1], bornes[1:]))


# ── Hachures grises : agent non disponible sur ce créneau (pause déjeuner
# ou hors de ses horaires contractuels) — ESSAI 08/2026, demande utilisatrice.
HATCH_FILL = PatternFill(patternType='lightDown', fgColor='FFBFBFBF', bgColor='FFF2F2F2')
# 7/ (demande utilisatrice 08/2026) : plus de fond "fermé" séparé — voir
# HATCH_FILL, seul code visuel désormais pour "agent pas au travail".
# Congé posé (journée ou demi-journée) — gris plus soutenu que les hachures,
# pour bien distinguer "en congé" de "hors de son contrat habituel".
CONGE_FILL = PatternFill('solid', fgColor='FFD0D0D0')


def _dans_horaires_agent(agent, jour, cs, ce, horaires_agents, pause_flex):
    """True si le créneau (cs, ce) tombe dans les horaires contractuels de
    l'agent ce jour-là (matin, après-midi, ou journée continue s'il n'y a pas
    de vraie coupure, fm == da).

    ⚠️ (correctif 08/2026, demande utilisatrice) : la "pause flexible"
    (colonne Affectations) autorise le SOLVEUR à placer exceptionnellement un
    agent sur son créneau de pause si besoin — mais ça ne veut pas dire que
    cet agent n'a PAS de pause déjeuner. Pour l'AFFICHAGE (cette fonction
    uniquement, jamais le moteur de résolution), la pause nominale
    (Horaires_Des_Agents, l'écart entre fm et da) est donc désormais toujours
    représentée par des hachures, pause flexible ou non — comme pour tous les
    agents ("tous les agents ont une heure de pause sauf s'ils terminent tôt,
    14h/15h"). Le paramètre `pause_flex` n'est plus utilisé ici, conservé
    uniquement pour ne pas casser les appels existants."""
    h = horaires_agents.get(agent, {}).get(jour)
    if not h:
        return False  # pas de contrat ce jour-là → hors horaires
    dm, fm, da, fa = h
    dans_matin = (dm is not None and fm is not None and cs >= dm and ce <= fm)
    dans_apm = (da is not None and fa is not None and cs >= da and ce <= fa)
    dans_global = (dm is not None and fa is not None and cs >= dm and ce <= fa and fm == da)
    return dans_matin or dans_apm or dans_global


# Arrivée/départ décalés par rapport au créneau (ESSAI 08/2026, demande
# utilisatrice) : pas de fond dédié (demande utilisatrice, 2e essai) — le
# texte "Arrivée/Départ HHhMM" en gras suffit à se repérer, sur le fond
# habituel (couleur de l'agent, cf. appelant).


def _arrivee_depart_label(agent, jour, cs, ce, horaires_agents):
    """Si une arrivée ou un départ de l'agent tombe EN PLEIN MILIEU de ce
    créneau (pas pile sur son bord), retourne un texte "Arrivée 9h15" /
    "Départ 17h15" à afficher à la place du contenu habituel. Retourne None
    si l'agent commence/termine pile sur le bord du créneau (rien à signaler)."""
    h = horaires_agents.get(agent, {}).get(jour)
    if not h:
        return None
    dm, fm, da, fa = h
    labels = []
    for t in (dm, da):
        if t is not None and cs < t < ce:
            labels.append(f'Arrivée {fmt_hhmm(t)}')
    for t in (fm, fa):
        if t is not None and cs < t < ce:
            labels.append(f'Départ {fmt_hhmm(t)}')
    return ' / '.join(labels) if labels else None


def generer_vue_agent(wb, week_num, jours, row_lookup, agents_recap,
                       horaires_agents, pause_flex, evenements):
    """Crée l'onglet 'Semaine_X_Agent' : un planning par agent (blocs empilés
    verticalement), avec une colonne par jour et une ligne par créneau fin.
    Chaque cellule est une FORMULE qui va chercher l'agent dans les colonnes
    techniques L-S de l'onglet 'Semaine_X' correspondant → toute modification
    du planning global se répercute automatiquement ici. En-tête (jours) figé
    en haut ; fond coloré PAR AGENT (couleurs de la capture d'écran d'Elo) ;
    grille étendue à partir de 8h (ESSAI 08/2026) pour voir les horaires
    d'arrivée même avant l'ouverture au public ; hachures grises quand
    l'agent n'est pas censé travailler sur ce créneau (pause déjeuner / hors
    de ses horaires) ; gris plein quand l'agent est en congé ; libellé
    "Arrivée HHhMM" / "Départ HHhMM" quand son horaire tombe en plein milieu
    d'un créneau plutôt que pile sur son bord. Vacataires exclus en amont
    (cf. appelant)."""
    sheet_src = f'Semaine_{week_num}'
    ws = wb.create_sheet(f'Semaine_{week_num}_Agent')
    ws.column_dimensions['A'].width = 14
    jours_semaine = [j['jour'] for j in jours]
    n_jours = len(jours_semaine)
    for i in range(n_jours):
        ws.column_dimensions[get_column_letter(i + 2)].width = 20

    # 08/2026 (demande utilisatrice) : grille étendue avec 2 créneaux fixes
    # 8h-9h et 9h-10h avant les créneaux réels du planning global, pour voir
    # les horaires d'arrivée des agents d'un coup d'œil, même avant
    # l'ouverture au public.
    fine_base = grille_fine_commune(jours)
    premier_cs = fine_base[0][0] if fine_base else 600
    fine_matinale = [seg for seg in [(480, 540), (540, 600)] if seg[1] <= premier_cs]
    fine = fine_matinale + fine_base
    # 1/ (demande utilisatrice 08/2026) : les créneaux de plus d'une heure
    # (ex: 17h-19h) sont découpés en blocs d'1h, pour repérer plus finement
    # les horaires des agents.
    fine_decoupee = []
    for cs, ce in fine:
        t = cs
        while ce - t > 60:
            fine_decoupee.append((t, t + 60))
            t += 60
        fine_decoupee.append((t, ce))
    fine = fine_decoupee

    # Jeunesse 1/2/3 (colonnes cachées O/P/Q) pointent toutes vers le même
    # libellé "Jeunesse" — peu importe la sous-colonne, l'agent doit juste
    # voir "je suis en Jeunesse ce créneau-là".
    SECTIONS_SRC = [('L', 'RDC'), ('M', 'Adulte'), ('N', 'M & F'),
                     ('O', 'Jeunesse'), ('P', 'Jeunesse'), ('Q', 'Jeunesse')]
    # Accueil/Animation et Réunion : détection via R/S (texte complet, avec
    # prénoms, utilisé seulement pour repérer si CET agent est concerné) mais
    # affichage via T/U (même événement, texte SANS AUCUN prénom — demande
    # utilisatrice : jamais de prénom dans la vue par agent, ni le sien ni
    # ceux des autres, juste le nom de l'événement + son horaire exact si
    # besoin).
    # ⚠️ EVENTS_SRC_COLS n'est plus utilisé pour la DÉTECTION (cf. correctif
    # ci-dessous : la détection se fait maintenant en Python, sur l'horaire
    # réel de l'événement, pas sur le gros bloc fusionné du planning
    # principal — voir _evenement_pour_agent_creneau). Gardé uniquement pour
    # mémoire de la correspondance de colonnes.
    EVENTS_SRC_COLS = [('R', 'T'), ('S', 'U')]

    date_par_jour = {j['jour']: j['date'] for j in jours}

    # ── CORRECTIF (09/2026, demande utilisatrice) : remontée EN DIRECT des
    # notes W-Z (Réunion/Accueil uniquement, cf. limite Absence documentée)
    # vers la vue par agent. Jusqu'ici, seuls les événements déjà présents
    # dans l'onglet Événements AU MOMENT DE LA GÉNÉRATION apparaissaient ici
    # (texte figé, cf. _evenement_pour_agent_creneau) : une note tapée à la
    # main APRÈS coup dans le classeur déjà généré ne remontait donc jamais
    # dans cet onglet, alors qu'elle remonte bien dans le planning principal
    # (colonnes H/I via formule). Cette section restaure ce comportement
    # attendu, SOUS FORME DE FORMULE (donc live), en plus du mécanisme
    # existant ci-dessus (qui reste prioritaire et gère la précision de durée
    # réelle des événements déjà connus à la génération, cf. §17.5 — pas de
    # régression sur ce point).
    # On retrouve, pour chaque (jour, agent), la case exacte où son nom
    # apparaît dans le petit tableau de notes (colonne W ou Y) et la case de
    # note en face (X ou Z), en répliquant EXACTEMENT la répartition faite
    # par ajouter_zone_notes_jour (même liste d'agents, même découpage en 2
    # groupes, mêmes lignes).
    notes_pos = {}  # (jour, agent) -> (helper_start_col, row)
    mid_notes = (len(agents_recap) + 1) // 2
    groupe1_notes, groupe2_notes = agents_recap[:mid_notes], agents_recap[mid_notes:]
    for jour in jours_semaine:
        lignes_ce_jour = [rr for (jj, _cs, _ce), rr in row_lookup.items() if jj == jour]
        if not lignes_ce_jour:
            continue
        header_row_jour = min(lignes_ce_jour) - 1
        for i, name in enumerate(groupe1_notes):
            notes_pos[(jour, name)] = (NOTES_HELPER_START[NOTES_NOTE1_COL], header_row_jour + 1 + i)
        for i, name in enumerate(groupe2_notes):
            notes_pos[(jour, name)] = (NOTES_HELPER_START[NOTES_NOTE2_COL], header_row_jour + 1 + i)

    def _evenement_pour_agent_creneau(agent, jour, cs, ce):
        """Cherche un événement (hors congé) où CET agent est concerné et
        dont l'horaire RÉEL (ev['cs']/ev['ce']) chevauche ce créneau fin
        précis — corrige le bug où un événement plus court que le gros bloc
        fusionné du planning principal (ex: accueil de classe 10h-11h dans
        un bloc 10h-12h30) se retrouvait affiché sur tout le bloc dans la
        vue par agent, et où un événement démarrant avant l'ouverture (ex:
        portage 9h-12h) n'apparaissait qu'à partir de l'heure d'ouverture."""
        date_str = date_par_jour.get(jour)
        if date_str is None:
            return None
        for ev in evenements:
            if ev['date'] != date_str:
                continue
            if ev['nom'].strip().lower() == 'congé':
                continue  # déjà géré séparément par _en_conge
            if agent not in ev.get('agents', []):
                continue
            if cs < ev['ce'] and ce > ev['cs']:  # chevauchement réel
                return ev
        return None

    # ── Congés par agent/jour : {(agent, jour): [(cs, ce), ...]} ────────
    # Un créneau qui chevauche un de ces intervalles est grisé "Congé",
    # quelle que soit la durée du congé (journée complète ou partielle).
    conge_par_agent_jour = {}
    for j in jours:
        date_str, jour = j['date'], j['jour']
        for ev in evenements:
            if ev['date'] != date_str or ev['nom'].strip().lower() != 'congé':
                continue
            for ag in ev.get('agents', []):
                conge_par_agent_jour.setdefault((ag, jour), []).append((ev['cs'], ev['ce']))

    def _en_conge(agent, jour, cs, ce):
        for ivs, ive in conge_par_agent_jour.get((agent, jour), []):
            if cs < ive and ce > ivs:  # chevauchement
                return True
        return False

    # ── En-tête unique, figé (ne se répète plus par agent) ──────────────
    hcell = ws.cell(row=1, column=1, value='Créneau')
    hcell.font = Font(size=9, bold=True)
    hcell.fill = PatternFill('solid', fgColor='FFCCCCCC')
    hcell.border = GREY_BORDER
    for ci, jour in enumerate(jours_semaine, start=2):
        hc = ws.cell(row=1, column=ci, value=jour)
        hc.font = Font(size=9, bold=True)
        hc.fill = PatternFill('solid', fgColor='FFCCCCCC')
        hc.border = GREY_BORDER
        hc.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'B2'

    r = 2
    for agent_idx, agent in enumerate(agents_recap):
        color = AGENT_COLORS.get(agent, '000000')
        fond_hex = AGENT_FILL_COLORS.get(agent, 'F4F4F4')
        fond_agent = 'FF' + fond_hex
        texte_agent = _texte_lisible(fond_hex)
        ws.merge_cells(f'A{r}:{get_column_letter(1 + n_jours)}{r}')
        c = ws.cell(row=r, column=1, value=agent)
        # Bandeau nom d'agent : gris clair, texte dans la couleur propre à
        # l'agent (inchangé — bonne lisibilité, cf. §13.17).
        c.fill = PatternFill('solid', fgColor='FFD9D9D9')
        c.font = Font(size=11, bold=True, color=color)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[r].height = 20
        r += 1

        for cs, ce in fine:
            cren_str = f'{cs//60:02d}:{cs%60:02d}-{ce//60:02d}:{ce%60:02d}'
            acell = ws.cell(row=r, column=1, value=cren_str)
            acell.font = Font(size=9)
            acell.fill = PatternFill('solid', fgColor=fond_agent)
            acell.border = GREY_BORDER
            for ci, jour in enumerate(jours_semaine, start=2):
                src_row = None
                for (jj, cs_src, ce_src), rr in row_lookup.items():
                    if jj == jour and cs_src <= cs and ce_src >= ce:
                        src_row = rr
                        break
                cell = ws.cell(row=r, column=ci)
                cell.border = GREY_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')

                if _en_conge(agent, jour, cs, ce):
                    # Agent en congé sur ce créneau → grisé (demande
                    # utilisatrice), quelle que soit la durée du congé.
                    cell.value = 'Congé'
                    cell.font = Font(size=9, italic=True, color='FF666666')
                    cell.fill = CONGE_FILL
                    continue

                ev_ici = _evenement_pour_agent_creneau(agent, jour, cs, ce)
                if ev_ici:
                    # Événement réel (accueil, portage, réunion...) chevauchant
                    # CE créneau fin précis — écrit en valeur littérale (pas en
                    # formule) car basé sur l'horaire réel de l'événement, pas
                    # sur le gros bloc fusionné du planning principal. Prend le
                    # pas sur la case "travaillé"/RDC/Adulte/etc. habituelle,
                    # y compris hors des horaires d'ouverture (ex: portage
                    # avant l'ouverture) — corrige le double bug signalé
                    # 09/2026 (durée d'événement dupliquée sur tout le bloc, et
                    # événement avant ouverture invisible).
                    cell.value = label_evenement_sans_noms(ev_ici, cs, ce)
                    cell.font = Font(size=9, color='FF' + texte_agent, bold=True)
                    cell.fill = PatternFill('solid', fgColor=fond_agent)
                    continue

                label_ad = _arrivee_depart_label(agent, jour, cs, ce, horaires_agents)
                if label_ad:
                    # Arrivée/départ en plein milieu du créneau — prioritaire
                    # sur tout le reste : c'est l'info la plus utile à voir
                    # ici. 2/ (demande utilisatrice) : pas de fond dédié,
                    # juste le texte en gras sur le fond habituel de l'agent.
                    cell.value = label_ad
                    cell.font = Font(size=8, italic=True, bold=True, color='FF7F4A00')
                    cell.fill = PatternFill('solid', fgColor=fond_agent)
                    continue

                if src_row is None:
                    # Pas de créneau réel du planning global ici (avant
                    # l'ouverture, après la fermeture, ou jour non ouvert).
                    if _dans_horaires_agent(agent, jour, cs, ce, horaires_agents, pause_flex):
                        # L'agent est pourtant censé être là (ex: préparation
                        # avant l'ouverture) → cellule "travaillée" neutre,
                        # dans sa couleur, sans texte (rien à afficher de plus
                        # précis puisque ce n'est pas suivi par le planning).
                        cell.value = None
                        cell.fill = PatternFill('solid', fgColor=fond_agent)
                    else:
                        # 7/ (demande utilisatrice 08/2026) : une seule façon
                        # de représenter "pas au travail" dans ce tableau —
                        # les hachures grises, qu'il s'agisse d'un horaire
                        # personnel hors contrat ou d'un créneau où la
                        # médiathèque n'est pas ouverte. Plus de gris à tiret
                        # séparé : un seul code visuel, plus simple à faire
                        # évoluer.
                        cell.value = None
                        cell.fill = HATCH_FILL
                    continue

                if not _dans_horaires_agent(agent, jour, cs, ce, horaires_agents, pause_flex):
                    # Agent pas censé travailler ici (pause déjeuner ou hors
                    # de ses horaires contractuels) → hachures grises.
                    cell.value = None
                    cell.fill = HATCH_FILL
                    continue

                agent_q = agent.replace('"', '""')
                inner = '""'
                # Accueil/Animation et Réunion CONNUS DÈS LA GÉNÉRATION ne
                # passent plus par ici : ils sont déjà traités plus haut
                # (_evenement_pour_agent_creneau), avec l'horaire réel de
                # l'événement plutôt que celui du gros bloc fusionné. Ne
                # reste ici que RDC/Adulte/M&F/Jeunesse, qui s'appliquent
                # légitimement à tout le bloc.
                for col_src, label in reversed(SECTIONS_SRC):
                    inner = (f'IF(ISNUMBER(SEARCH("{agent_q}",\'{sheet_src}\'!${col_src}${src_row})),'
                              f'"{label}",{inner})')
                # CORRECTIF 09/2026 : si une note W-Z (Réunion/Accueil) existe
                # pour CET agent CE jour-là et que son horaire chevauche ce
                # créneau fin, elle prend le pas sur RDC/Adulte/M&F/Jeunesse —
                # en formule, donc mise à jour automatique si la note est
                # tapée/modifiée après la génération. Absence (catégorie 2)
                # volontairement exclue (limite documentée, inchangée).
                note_pos = notes_pos.get((jour, agent))
                if note_pos:
                    hstart, note_row = note_pos
                    cat_c = get_column_letter(hstart)
                    txt_c = get_column_letter(hstart + 1)
                    deb_c = get_column_letter(hstart + 2)
                    fin_c = get_column_letter(hstart + 3)
                    cs_h, ce_h = cs / 60, ce / 60
                    note_cond = (f"AND(OR('{sheet_src}'!${cat_c}${note_row}=1,"
                                 f"'{sheet_src}'!${cat_c}${note_row}=3),"
                                 f"'{sheet_src}'!${deb_c}${note_row}<{ce_h},"
                                 f"'{sheet_src}'!${fin_c}${note_row}>{cs_h})")
                    inner = f"IF({note_cond},'{sheet_src}'!${txt_c}${note_row},{inner})"
                cell.value = f'=IFERROR({inner},"")'
                cell.font = Font(size=9, color='FF' + texte_agent, bold=True)
                cell.fill = PatternFill('solid', fgColor=fond_agent)
            r += 1
        r += 1  # ligne vide entre agents


if __name__ == '__main__':
    generer()

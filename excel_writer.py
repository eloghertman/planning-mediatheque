"""
Excel Writer v12 — Médiathèque
════════════════════════════════════════════════════════════════
Nouveautés v12 :
  - Planning_Agent : lignes 8h30-9h et 9h-10h (heure d'arrivée)
  - Nouvelles couleurs : lavande Jeunesse, hachures hors-horaires,
    rouge doux congés
  - Totaux SP dynamiques (formules IF) dans Planning_Agent
  - Récap Semaine_N dynamique (références cross-sheet)
════════════════════════════════════════════════════════════════
"""

import io
import re
import zipfile
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from collections import defaultdict

SECTIONS   = ['RDC', 'Adulte', 'MF', 'Jeunesse']
JOURS_SP   = ['Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
MOIS_FR    = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',6:'Juin',
              7:'Juillet',8:'Août',9:'Septembre',10:'Octobre',
              11:'Novembre',12:'Décembre'}

# ── Palette ────────────────────────────────────────────────────
C = {
    'hdr_dark':   '1A2F4A', 'hdr_rouge':  'C0392B', 'hdr_bleu':   '2471A3',
    'rdc':        'D6E8F7', 'adulte':     'D4EDD4', 'mf':         'FFF0CC',
    # Jeunesse : lavande douce (remplace rose/rouge)
    'jeunesse':   'FFE0E0',          # onglet Semaine (inchangé, repère visuel)
    'jeunesse_ag':'E8DAEF',          # Planning_Agent : lavande
    'closed':     'EBEBEB',
    'event_bg':   'EDE7F6', 'agent_txt': '2C3E50',
    'gray':       'AAAAAA', 'alt1':      'FDFEFE', 'alt2':        'EBF5FB',
    'dark':       '2C3E50', 'sp_ok':     'D5F5E3', 'sp_warn':     'FFF3CD',
    'sp_alert':   'FFCCCC', 'sp_ok_t':   '1E8449', 'sp_warn_t':   'E67E22',
    'sp_alert_t': '8B0000',
    'hdr_rdc':    '2980B9', 'hdr_adulte':'27AE60', 'hdr_mf':      'E67E22',
    'hdr_j':      'E74C3C', 'hdr_purple':'8E44AD', 'hdr_dark2':   '34495E',
    'oos_bg':     'FFCDD2', 'oos_txt':   '8B0000',
    'vac_bg':     'F0E6FF', 'vac_txt':   '6A0DAD',
    # Congé : rouge très doux (remplace FFCCCC/8B0000)
    'conge_bg':   'FADBD8', 'conge_txt': '922B21',
    # Hors-horaires : hachures (défini via _hatch())
    'bureau_bg':  'EDE7F6',          # au bureau (présent, hors SP) → violet doux
    'bureau_txt': '6A1B9A',          # texte violet bureau
    'off_bg':     'EBEBEB',          # off / hors horaires → même gris que médiathèque fermée
    'off_txt':    '999999',          # texte tiret gris
    # Arrivée
    'arrival_bg': 'EAF2FF',
    'sp_dyn_bg':  'E8DAEF',   # lavande neutre pour cellules SP dynamiques
    'arrival_txt':'1F4E79',
}
SEC_BG  = {'RDC': C['rdc'],      'Adulte': C['adulte'],
           'MF':  C['mf'],       'Jeunesse': C['jeunesse']}
SEC_BG_AG = {'RDC': C['rdc'],    'Adulte': C['adulte'],
             'MF':  C['mf'],     'Jeunesse': C['jeunesse_ag']}
SEC_HDR = {'RDC': C['hdr_rdc'],  'Adulte': C['hdr_adulte'],
           'MF':  C['hdr_mf'],   'Jeunesse': C['hdr_j']}

# Couleurs de police distinctes par agent (planning semaine)
AGENT_COLORS = {
    'Christine':        '000000',
    'Macha':            '8B0000',
    'Marie-France':     '006400',
    'Léa':              'CC5500',
    'Chloé':            '880E4F',
    'Anne-Françoise':   '01579B',
    'Delphine':         '6D4C41',
    'Stéphane':         '00008B',
    'Barbara':          '005F73',
    'Stéphanie':        'AD1457',
    'Robin':            '8B4513',
    'Guillaume':        '556B2F',
    'Tiphaine':         '4B0082',
    'Agnès':            '1A6B3A',
    'Vacataire 1':      '6A0DAD',
    'Vacataire 2':      '5D4E8C',
}

def agent_color(agent):
    return AGENT_COLORS.get(agent, '2C3E50')


# ── Helpers styles ─────────────────────────────────────────────
def _fill(h):
    return PatternFill('solid', fgColor=h)

def _hatch(fg='BBBBBB', bg='F5F5F5'):
    """Motif hachuré diagonal — représente hors-horaires."""
    return PatternFill(patternType='lightDown', fgColor=fg, bgColor=bg)

def _font(bold=False, size=10, color='000000', italic=False, name='Calibri'):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def _aln(h='center', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _brd(color='CCCCCC', style='thin'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _set(cell, value=None, bg=None, fnt=None, aln=None, brd=None, hatch=False):
    if value is not None:
        cell.value = value
    if hatch:
        cell.fill = _hatch()
    elif bg:
        cell.fill = _fill(bg)
    if fnt:
        cell.font = fnt
    if aln:
        cell.alignment = aln
    if brd:
        cell.border = brd


def is_vacataire(agent):
    return 'vacataire' in str(agent).lower()

def min_to_hhmm(m):
    return f"{int(m)//60}h{int(m)%60:02d}" if m else '0h00'

def min_to_dec(m):
    return round(m / 60, 2) if m else 0.0


# ══════════════════════════════════════════════════════════════
#  ONGLET SEMAINE
# ══════════════════════════════════════════════════════════════


def _cell_text(v):
    """Extrait le texte d'une cellule pour comparaison (gère CellRichText)."""
    if v is None:
        return None
    try:
        from openpyxl.cell.rich_text import CellRichText
        if isinstance(v, CellRichText):
            return str(v)
    except Exception:
        pass
    return str(v) if v != '' else None


def _merge_consecutive_cells(ws, col, row_start, row_end):
    """Fusionne les cellules consécutives identiques (même contenu textuel) dans une colonne."""
    if row_start >= row_end:
        return
    i = row_start
    while i <= row_end:
        val_i = _cell_text(ws.cell(row=i, column=col).value)
        if val_i is not None:
            j = i + 1
            while j <= row_end and _cell_text(ws.cell(row=j, column=col).value) == val_i:
                j += 1
            if j > i + 1:
                src = ws.cell(row=i, column=col)
                src_fill  = copy(src.fill)
                src_font  = copy(src.font)
                src_align = copy(src.alignment)
                src_bord  = copy(src.border)
                ws.merge_cells(start_row=i, start_column=col,
                               end_row=j-1, end_column=col)
                mc = ws.cell(row=i, column=col)
                mc.fill      = src_fill
                mc.font      = src_font
                mc.alignment = Alignment(horizontal=src_align.horizontal,
                                         vertical='center', wrap_text=True)
                mc.border    = src_bord
            i = j
        else:
            i += 1


def _merge_consecutive_agent_cells(ws, col, row_start, row_end):
    """Fusionne les cellules consécutives dans une colonne Planning_Agent :
    - Même valeur non-vide (congé, événement, SP section)
    - OU valeur vide/None avec même patternType de remplissage (bureau, off/hachures)
    """
    if row_start >= row_end:
        return

    def _cell_key(r):
        c = ws.cell(row=r, column=col)
        v = c.value
        v_str = _cell_text(v)
        # La sentinelle '·' (hors-horaires) est traitée comme vide pour la fusion
        if v_str == '·':
            v_str = None
        if v_str is not None:
            return ('val', v_str)
        # Cellule vide : clé = type de remplissage
        pt = c.fill.patternType if c.fill else None
        fg = ''
        try:
            fg = c.fill.fgColor.rgb if c.fill and c.fill.fgColor else ''
        except Exception:
            pass
        return ('empty', pt, fg)

    i = row_start
    while i <= row_end:
        key_i = _cell_key(i)
        j = i + 1
        while j <= row_end and _cell_key(j) == key_i:
            j += 1
        if j > i + 1:
            src = ws.cell(row=i, column=col)
            src_fill  = copy(src.fill)
            src_font  = copy(src.font)
            src_align = copy(src.alignment)
            src_bord  = copy(src.border)
            ws.merge_cells(start_row=i, start_column=col,
                           end_row=j-1, end_column=col)
            mc = ws.cell(row=i, column=col)
            mc.fill      = src_fill
            mc.font      = src_font
            mc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            mc.border    = src_bord
        i = j

# ── Catégorisation des événements en colonnes G/H/I/J ──────────────────────
# G = accueil public (classes, assistantes maternelles...)
# H = animations/événements (café, conférence, stage, bébé se livre, etc.)
# I = réunions/RDV internes (réunion, RMH, RDV RH, brainstorming...)
# J = absences (congé, RTT, formation, absence...)
_EV_KEYWORDS = {
    'G': ['accueil', 'assistante'],
    'H': ['café', 'club', 'conférence', 'conference', 'stage', 'bébé', 'bebe',
          'caj', 'cml', 'rmh', 'forum', 'hors les murs', 'sellier', 'saison', 'défi',
          'tournée', 'diplôme', 'éloquence', 'eloquence', 'atelier', 'animation',
          'visite'],
    'I': ['réunion', 'reunion', 'rdv', 'rmh', 'brainstorm', 'pôle', 'pole'],
    'J': ['congé', 'conge', 'rtt', 'vacation', 'absence', 'formation'],
}

def _ev_txt_color():
    """Couleur de texte pour les événements sur fond violet."""
    return '4A148C'


def _rich_agents(agents, oos=False):
    """CellRichText : agents avec couleurs distinctes, séparateur gris."""
    parts = []
    sep_font = InlineFont(color='AAAAAA', b=False, sz=9)
    for i, agent in enumerate(agents):
        if i > 0:
            parts.append(TextBlock(sep_font, '  /  '))
        color = C['oos_txt'] if oos else agent_color(agent)
        af = InlineFont(color=color, b=True, sz=10)
        parts.append(TextBlock(af, agent))
    return CellRichText(*parts) if parts else None

    if cren_range:
        parts.append(TextBlock(cren_font, f'\n{cren_range}'))

    return CellRichText(*parts) if parts else None


def _set_rich(cell, rich_val, bg, brd=None):
    """Set une cellule avec CellRichText + fond + alignement."""
    cell.value = rich_val
    cell.fill = _fill(bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if brd:
        cell.border = brd


def _ev_categorie(nom):
    """Retourne la colonne (G/H/I/J) d'un événement selon son nom."""
    n = nom.lower()
    for col, kws in _EV_KEYWORDS.items():
        if any(kw in n for kw in kws):
            return col
    return 'H'  # par défaut : animation

def _format_ev(ev, cs=None, ce=None):
    """Formate un événement avec horaire exact si non aligné sur le créneau.
    Ne montre pas l'horaire pour les événements journée entière (9h-19h)."""
    time_suffix = ''
    FULL_DAY_START, FULL_DAY_END = 540, 1140  # 9h-19h
    is_full_day = (ev['debut'] <= FULL_DAY_START and ev['fin'] >= FULL_DAY_END)
    if not is_full_day and cs is not None and ce is not None:
        if ev['debut'] != cs or ev['fin'] != ce:
            h1 = f"{ev['debut']//60}h{ev['debut']%60:02d}" if ev['debut']%60 else f"{ev['debut']//60}h"
            h2 = f"{ev['fin']//60}h{ev['fin']%60:02d}"    if ev['fin']%60   else f"{ev['fin']//60}h"
            time_suffix = f" [{h1}-{h2}]"
    return f"{ev['nom']}{time_suffix} ({', '.join(ev['agents'])})"

def _dispatch_events(events, cs=None, ce=None):
    """Répartit les événements dans G/H/I/J avec horaire exact si non aligné."""
    buckets = {'G': [], 'H': [], 'I': [], 'J': []}
    for ev in events:
        buckets[_ev_categorie(ev['nom'])].append(_format_ev(ev, cs, ce))
    return {col: '\n'.join(lines) if lines else None
            for col, lines in buckets.items()}

def write_week_sheet(wb, week_data, metadata, agent_sp_cells=None):
    """
    agent_sp_cells : dict {agent: {jour: 'C42'}} — références dans
    Planning_Agent_Semaine_N pour le récap dynamique.
    Si None : utilise les valeurs statiques du moteur.
    """
    week_num       = week_data['week_num']
    week_dates     = week_data['week_dates']
    plan           = week_data['plan']
    sp_jour_cum    = week_data.get('sp_jour_cum', {})
    sp_alerts      = week_data.get('sp_alerts', {})
    creneaux       = metadata['creneaux']
    sp_minmax_all  = metadata.get('sp_minmax_all', {})
    sp_minmax_week = sp_minmax_all.get(week_num, sp_minmax_all.get(2, {}))
    affectations   = metadata.get('affectations', {})
    evenements     = metadata.get('evenements', {})
    is_sam_week    = 'Samedi' in week_dates

    mois_num = {'janvier':1,'février':2,'mars':3,'avril':4,'mai':5,'juin':6,
                'juillet':7,'août':8,'septembre':9,'octobre':10,
                'novembre':11,'décembre':12
                }.get(metadata['mois'].lower(), 1)
    mois_cap  = MOIS_FR.get(mois_num, metadata['mois'].capitalize())
    annee     = metadata['annee']
    sem_type  = week_data.get('semaine_type', '')

    dates_list = sorted([d for d in week_dates.values() if d])
    if dates_list:
        d0, d1 = dates_list[0], dates_list[-1]
        m2 = MOIS_FR.get(d1.month, '')
        date_range = (f"{d0.day} au {d1.day} {mois_cap} {annee}"
                      if d0.month == d1.month
                      else f"{d0.day} {mois_cap} au {d1.day} {m2} {annee}")
    else:
        date_range = f"Semaine {week_num}"

    sname = f"Semaine_{week_num}"
    if sname in wb.sheetnames:
        wb.remove(wb[sname])
    ws = wb.create_sheet(sname)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'B1'  # Fige la colonne A (Créneau) pour navigation horizontale
    for col, w in {'A':14,'B':20,'C':20,'D':20,'E':26,'F':20,'G':22,'H':22,'I':26}.items():
        ws.column_dimensions[col].width = w

    # Mapping jour → (row_cren_start, row_cren_end) dans cet onglet
    # Utilisé par Planning_Agent pour les formules cross-sheet
    jour_cren_rows = {}

    row = 1
    ws.merge_cells(f'A{row}:I{row}')
    _set(ws.cell(row=row, column=1,
                 value=f"PLANNING SERVICE PUBLIC — Semaine {week_num}  |  {date_range}  |  {sem_type}"),
         bg=C['hdr_dark'], fnt=_font(bold=True, size=13, color='FFFFFF'), aln=_aln('center'))
    ws.row_dimensions[row].height = 30
    row += 1

    ws.merge_cells(f'A{row}:I{row}')
    _set(ws.cell(row=row, column=1,
                 value="  RDC     Adulte     Musique & Films     Jeunesse     "
                       "Fermé     Événement     ALERTE     Rouge=hors section"),
         bg=C['dark'], fnt=_font(size=9, color='FFFFFF', italic=True), aln=_aln('left'))
    ws.row_dimensions[row].height = 15
    row += 2

    for jour in JOURS_SP:
        date = week_dates.get(jour)
        if date is None:
            continue
        jour_plan   = plan.get(jour, {})
        samedi_type = jour_plan.get('_samedi_type')
        is_sam      = jour == 'Samedi'

        j_label = f"  {jour.upper()}  {date.day} {MOIS_FR.get(date.month,'')} {date.year}"
        if is_sam and samedi_type:
            j_label += f"  —  SAMEDI {samedi_type}"
        j_color = (C['hdr_rouge'] if (is_sam and samedi_type == 'ROUGE')
                   else C['hdr_bleu']  if (is_sam and samedi_type == 'BLEU')
                   else C['hdr_dark'])

        ws.merge_cells(f'A{row}:I{row}')
        _set(ws.cell(row=row, column=1, value=j_label),
             bg=j_color, fnt=_font(bold=True, size=12, color='FFFFFF'), aln=_aln('left'))
        ws.row_dimensions[row].height = 26
        row += 1

        # ── Jour férié ──────────────────────────────────────────
        if jour_plan.get('_ferie'):
            ws.merge_cells(f'A{row}:I{row}')
            _set(ws.cell(row=row, column=1, value='🎉  JOUR FÉRIÉ  —  Médiathèque fermée'),
                 bg='F4ECF7', fnt=_font(bold=True, size=11, color='6C3483'),
                 aln=_aln('center'))
            ws.row_dimensions[row].height = 28
            row += 1
            jour_cren_rows[jour] = (row, row - 1)  # plage vide
            for ci in range(1, 10):
                ws.cell(row=row, column=ci).fill = _fill('F0F3F4')
            ws.row_dimensions[row].height = 4
            row += 1
            continue

        # Colonnes : A=Créneau, B=RDC, C=Adulte, D=MF, E=Jeunesse, F=Accueil, G=Animation, H=Réunion, I=Absence
        hdrs = [('Créneau',C['dark']),('RDC',SEC_HDR['RDC']),
                ('Adulte',SEC_HDR['Adulte']),('M & F',SEC_HDR['MF']),
                ('Jeunesse',SEC_HDR['Jeunesse']),('Accueil',C['hdr_purple']),
                ('Animation',C['hdr_purple']),('Réunion',C['hdr_purple']),
                ('Absence',C['hdr_purple'])]
        for ci,(h_txt,hc) in enumerate(hdrs, 1):
            _set(ws.cell(row=row, column=ci, value=h_txt),
                 bg=hc, fnt=_font(bold=True, size=10, color='FFFFFF'),
                 aln=_aln('center'), brd=_brd())
        ws.row_dimensions[row].height = 20
        row += 1

        # Mémoriser la première ligne de créneaux pour ce jour
        _cren_start_row = row

        for ni,(cren_name,cs,ce) in enumerate(creneaux):
            slot = jour_plan.get(cren_name)
            if slot is None:
                # Événements pendant ce créneau fermé (réunions, formations hors ouverture)
                date_str_closed = date.strftime('%Y-%m-%d')
                ev_closed = [ev for ev in evenements.get(date_str_closed, [])
                             if ev['debut'] < ce and ev['fin'] > cs]
                # Colonnes fermées A-E (Créneau + 4 sections)
                for ci, val in enumerate([cren_name,'—','—','—','—'], 1):
                    _set(ws.cell(row=row, column=ci, value=val or None),
                         bg=C['closed'],
                         fnt=_font(size=10 if ci==1 else 9, color=C['gray'], italic=ci>1),
                         aln=_aln(), brd=_brd())
                # 4 colonnes événement F-I
                ev_buckets_c = _dispatch_events(ev_closed, cs=cs, ce=ce)
                for ev_col_i, col_key in enumerate(['G','H','I','J']):
                    ev_val = ev_buckets_c[col_key]
                    _set(ws.cell(row=row, column=6+ev_col_i, value=ev_val),
                         bg=C['event_bg'] if ev_val else C['closed'],
                         fnt=_font(size=9, italic=True,
                                   color=_ev_txt_color() if ev_val else C['gray']),
                         aln=_aln('left'), brd=_brd())
                if ev_closed:
                    max_lines = max((len(v.split('\n')) for v in ev_buckets_c.values() if v), default=1)
                    ws.row_dimensions[row].height = max(20, 16*max_lines)
            else:
                assgn  = slot.get('assignment', {s:[] for s in SECTIONS})
                events = slot.get('events', [])
                alerts = slot.get('alerts', [])
                oos    = slot.get('out_of_section', {})

                # Col A = créneau
                _set(ws.cell(row=row, column=1, value=cren_name),
                     bg='FDFEFE', fnt=_font(bold=True, size=10), aln=_aln(), brd=_brd())

                for si, sect in enumerate(SECTIONS):
                    col = si + 2   # B=2 RDC, C=3 Adulte, D=4 MF, E=5 Jeunesse
                    agents  = assgn.get(sect, [])
                    has_al  = any(sect in al for al in alerts)
                    is_oos  = oos.get(sect, False)

                    if is_oos and agents:
                        # Plain text pour que SEARCH fonctionne
                        val = '  /  '.join(agents)
                        _set(ws.cell(row=row, column=col, value=val),
                             bg=C['oos_bg'],
                             fnt=_font(size=10, color=C['oos_txt'], bold=True),
                             aln=_aln('center'), brd=_brd())
                    elif agents:
                        bg   = SEC_BG[sect]
                        cell = ws.cell(row=row, column=col)
                        if len(agents) == 1:
                            # Un seul agent → plain text (SEARCH fonctionne)
                            _set(cell, value=agents[0],
                                 bg=bg,
                                 fnt=_font(size=10, color=agent_color(agents[0]), bold=True),
                                 aln=_aln('center'), brd=_brd())
                        else:
                            # Plusieurs agents → CellRichText pour couleurs distinctes
                            _set_rich(cell, _rich_agents(agents), bg, _brd())
                    elif has_al:
                        _set(ws.cell(row=row, column=col, value='ALERTE'),
                             bg=C['conge_bg'],
                             fnt=_font(size=10, color=C['sp_alert_t'], bold=True),
                             aln=_aln('center'), brd=_brd())
                    else:
                        _set(ws.cell(row=row, column=col, value='—'),
                             bg='FDF2F8',
                             fnt=_font(size=10, color='C0392B'),
                             aln=_aln('center'), brd=_brd())

                # 4 colonnes événement (F=Accueil, G=Animation, H=Réunion, I=Absence)
                ev_buckets = _dispatch_events(events, cs=cs, ce=ce)
                for ev_col_i, col_key in enumerate(['G','H','I','J']):
                    ev_val = ev_buckets[col_key]
                    n_lines = len(ev_val.split('\n')) if ev_val else 0
                    _set(ws.cell(row=row, column=6+ev_col_i, value=ev_val),
                         bg=C['event_bg'] if ev_val else 'FAFAFA',
                         fnt=_font(size=9, italic=True,
                                   color=_ev_txt_color() if ev_val else C['gray']),
                         aln=_aln('left'), brd=_brd())

                max_ev_lines = max((len(v.split('\n')) for v in ev_buckets.values() if v), default=0)
                h = max(20, 16*max(max_ev_lines, 1))
                ws.row_dimensions[row].height = h
            row += 1

        # Mémoriser la dernière ligne de créneaux pour ce jour
        jour_cren_rows[jour] = (_cren_start_row, row - 1)

        for ci in range(1, 10):
            ws.cell(row=row, column=ci).fill = _fill('F0F3F4')
        ws.row_dimensions[row].height = 4
        row += 1
    row += 1
    ws.merge_cells(f'A{row}:I{row}')
    _set(ws.cell(row=row, column=1,
                 value="RÉCAP — Heures de Service Public par agent (heures décimales)"),
         bg='34495E', fnt=_font(bold=True, size=11, color='FFFFFF'), aln=_aln('center'))
    ws.row_dimensions[row].height = 24
    row += 1

    hdrs2 = ['Agent','Section','Mardi','Mercredi','Jeudi','Vendredi','Samedi',
              '','','','','Total (h)']
    hcols = ['34495E','34495E','1A5276','1A5276','1A5276','1A5276','922B21',
             'AAAAAA','AAAAAA','AAAAAA','AAAAAA','145A32']
    for ci,(h_txt,hc) in enumerate(zip(hdrs2,hcols)):
        _set(ws.cell(row=row, column=ci+1, value=h_txt),
             bg=hc, fnt=_font(bold=True, size=10, color='FFFFFF'),
             aln=_aln(), brd=_brd())
    ws.row_dimensions[row].height = 20
    row += 1

    def primary_section(agent):
        sects = affectations.get(agent, [])
        return sects[0] if sects else '?'

    agents_shown = sorted(
        [a for a in sp_minmax_week.keys() if not is_vacataire(a)],
        key=lambda x: (primary_section(x), x)
    ) + [a for a in affectations.keys() if is_vacataire(a)]

    pa_sheet = f"Planning_Agent_Semaine_{week_num}"

    for i, agent in enumerate(agents_shown):
        bg   = C['alt1'] if i % 2 == 0 else C['alt2']
        sect = primary_section(agent)

        _set(ws.cell(row=row, column=1, value=agent),
             bg=bg, fnt=_font(bold=True, size=10), aln=_aln('left'), brd=_brd())
        _set(ws.cell(row=row, column=2,
                     value='VAC' if is_vacataire(agent) else sect),
             bg=C['vac_bg'] if is_vacataire(agent) else SEC_BG.get(sect, bg),
             fnt=_font(size=10, color=C['vac_txt'] if is_vacataire(agent) else '000000'),
             aln=_aln(), brd=_brd())

        total_h_static = 0.0
        for ji, jour in enumerate(JOURS_SP):
            col  = ji + 3
            cell = ws.cell(row=row, column=col)
            sp_min = sp_jour_cum.get(jour, {}).get(agent, 0)
            sp_h   = min_to_dec(sp_min)
            total_h_static += sp_h

            if week_dates.get(jour) is not None:
                # Formule dynamique si on a les refs Planning_Agent
                if agent_sp_cells and agent in agent_sp_cells:
                    cell_ref = agent_sp_cells[agent].get(jour)
                    if cell_ref:
                        cell.value = f"='{pa_sheet}'!{cell_ref}"
                        cell.number_format = '0.00'
                    else:
                        cell.value = sp_h if sp_h > 0 else 0
                        cell.number_format = '0.00'
                else:
                    cell.value = sp_h if sp_h > 0 else 0
                    cell.number_format = '0.00'

                bg_c = ('FFF9C4' if jour == 'Samedi' and sp_h > 0
                         else C['alt2'] if sp_h > 0 else bg)
                cell.fill      = _fill(bg_c)
                cell.font      = _font(size=10, bold=sp_h > 0)
                cell.alignment = _aln()
                cell.border    = _brd()
            else:
                cell.value     = '—'
                cell.fill      = _fill(bg)
                cell.font      = _font(size=9, color=C['gray'])
                cell.alignment = _aln()
                cell.border    = _brd()

        # Total — formule SUM des colonnes jour (col L=12, masquable)
        tc = ws.cell(row=row, column=12)
        # Colonnes C..G (indices 3..7) = 5 jours max → Total en col L (12)
        sum_cols = [get_column_letter(3 + ji)
                    for ji, jour in enumerate(JOURS_SP)
                    if week_dates.get(jour) is not None]
        if sum_cols:
            tc.value = f"=SUM({','.join(c+str(row) for c in sum_cols)})"
        else:
            tc.value = 0
        tc.number_format = '0.00'

        # Coloration total (basée sur valeur moteur pour le style)
        total_h = total_h_static
        if not is_vacataire(agent):
            mm     = sp_minmax_week.get(agent, {})
            min_sp = (mm.get('Min_MarSam',0) if is_sam_week else mm.get('Min_MarVen',0))
            max_sp = (mm.get('Max_MarSam',99) if is_sam_week else mm.get('Max_MarVen',99))
            if agent in sp_alerts:
                t_bg, t_fc = C['sp_alert'], C['sp_alert_t']
            elif total_h < min_sp:
                t_bg, t_fc = C['sp_alert'], C['sp_alert_t']
            elif total_h > max_sp:
                t_bg, t_fc = C['sp_warn'], C['sp_warn_t']
            else:
                t_bg, t_fc = C['sp_ok'], C['sp_ok_t']
        else:
            t_bg = C['vac_bg'] if total_h > 0 else bg
            t_fc = C['vac_txt'] if total_h > 0 else C['gray']

        tc.fill = _fill(t_bg)
        tc.font = _font(bold=True, size=10, color=t_fc)
        tc.alignment = _aln()
        tc.border    = _brd()
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1
    ws.merge_cells(f'A{row}:I{row}')
    _set(ws.cell(row=row, column=1,
                 value="Vert=OK  |  Rouge=sous min  |  Orange=sur max  |  "
                       "Violet=vacataire  |  Rose = agent hors section habituelle"),
         bg='F4F6F7', fnt=_font(size=9, italic=True, color='5D6D7E'), aln=_aln('left'))
    ws.row_dimensions[row].height = 16
    return ws, jour_cren_rows


# ══════════════════════════════════════════════════════════════
#  ONGLET PLANNING PAR AGENT — 1 PAR SEMAINE
# ══════════════════════════════════════════════════════════════

def write_planning_agent_week_sheet(wb, week_data, metadata, jour_cren_rows=None):
    """
    Retourne agent_sp_cells = {agent: {jour: 'C42'}}
    pour utilisation dans les formules cross-sheet du récap Semaine_N.
    """
    week_num     = week_data['week_num']
    week_dates   = week_data['week_dates']
    plan         = week_data['plan']
    sp_jour_cum  = week_data.get('sp_jour_cum', {})
    creneaux     = metadata['creneaux']
    affectations = metadata['affectations']
    horaires_ag  = metadata['horaires_agents']
    evenements   = metadata['evenements']
    sem_type     = week_data.get('semaine_type', '')

    mois_num = {'janvier':1,'février':2,'mars':3,'avril':4,'mai':5,'juin':6,
                'juillet':7,'août':8,'septembre':9,'octobre':10,
                'novembre':11,'décembre':12
                }.get(metadata['mois'].lower(), 1)
    mois_cap = MOIS_FR.get(mois_num, '')
    annee    = metadata['annee']

    sname = f"Planning_Agent_Semaine_{week_num}"
    if sname in wb.sheetnames:
        wb.remove(wb[sname])
    ws = wb.create_sheet(sname)
    ws.sheet_view.showGridLines = False

    jours_presents = [j for j in JOURS_SP if week_dates.get(j) is not None]
    nb_jours       = len(jours_presents)
    nb_cols        = 2 + nb_jours
    end_col        = get_column_letter(nb_cols)
    col_h_total    = nb_cols + 1   # colonne H (ou suivante) = total heures semaine

    agents_sorted = sorted(
        [a for a in affectations.keys() if not is_vacataire(a)],
        key=lambda x: (affectations.get(x, [''])[0], x)
    ) + [a for a in affectations.keys() if is_vacataire(a)]

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 15
    for ci in range(3, 3 + nb_jours):
        ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.column_dimensions[get_column_letter(col_h_total)].width = 12

    # Mapping jour → colonne lettre
    col_map = {}
    for ji, jour in enumerate(jours_presents):
        col_map[jour] = 3 + ji

    row = 1

    # ── Titre global ────────────────────────────────────────────
    ws.merge_cells(f'A{row}:{end_col}{row}')
    dates_list = sorted([d for d in week_dates.values() if d])
    d0, d1 = dates_list[0], dates_list[-1]
    titre = (f"PLANNING PAR AGENT — Semaine {week_num}  |  "
             f"{d0.day} au {d1.day} {mois_cap} {annee}  |  {sem_type}")
    _set(ws.cell(row=row, column=1, value=titre),
         bg=C['hdr_dark'], fnt=_font(bold=True, size=13, color='FFFFFF'),
         aln=_aln('center'))
    ws.row_dimensions[row].height = 30
    row += 1

    # ── Légende ─────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:{end_col}{row}')
    _set(ws.cell(row=row, column=1,
                 value="  SP = section (couleur)     Arr./Fin XXhYY = horaire décalé (gris)     "
                       "Gris uni = au bureau hors SP     Hachures = hors horaires     "
                       "Jaune = événement / congé     Rose = hors section habituelle"),
         bg='34495E', fnt=_font(size=9, italic=True, color='FFFFFF'), aln=_aln('left'))
    ws.row_dimensions[row].height = 15
    row += 2

    # Créneaux additionnels avant 10h
    EARLY_SLOTS = [
        ('8:30-9:00',  510, 540),
        ('9:00-10:00', 540, 600),
    ]

    # Créneau additionnel après 19h (affiché si agent finit après 19h00)
    LATE_SLOT = ('19:00-19:30', 1140, 1170)

    # Helper : total heures réellement travaillées sur la semaine
    # (congés exclus, samedis non roulés exclus, jours sans horaires exclus)
    def total_heures_semaine(agent):
        roulement_sem = metadata.get('roulement_all', {}).get(week_num, {})
        samedi_type   = week_data.get('samedi_type')

        total_min = 0
        for jour in jours_presents:
            date = week_dates.get(jour)
            if not date:
                continue
            date_str = date.strftime('%Y-%m-%d')

            # Samedi : vérifier le roulement
            if jour == 'Samedi':
                ag_roulement = roulement_sem.get(agent)
                if ag_roulement and samedi_type:
                    if ag_roulement not in (samedi_type, 'BOTH'):
                        continue  # samedi non travaillé ce roulement

            # Congé : si événement couvre ce jour pour cet agent
            ev_jour = evenements.get(date_str, [])
            is_conge = any(
                agent in ev['agents']
                and ev['nom'].lower() in ('congé', 'conge', 'rtt', 'vacation', 'formation')
                for ev in ev_jour
            )
            if is_conge:
                continue  # journée non travaillée

            h = horaires_ag.get(agent, {}).get(jour)
            if not h:
                continue
            sm, em, sa, ea = h
            if sm is not None and em is not None:
                total_min += em - sm
            if sa is not None and ea is not None:
                total_min += ea - sa

        h_int = total_min // 60
        m_int = total_min % 60
        return f"{h_int}h{m_int:02d}" if m_int else f"{h_int}h"

    # Durées en MINUTES entières pour les créneaux SP (évite pb décimal Excel FR)
    cren_durations_min = [(ce - cs) for (_, cs, ce) in creneaux]

    # Résultat : {agent: {jour: 'C42'}} pour cross-sheet refs
    agent_sp_cells = {}

    for agent in agents_sorted:
        sects      = affectations.get(agent, [])
        sect_prim  = sects[0] if sects else '?'
        hdr_bg     = C['vac_bg'] if is_vacataire(agent) else SEC_HDR.get(sect_prim, C['hdr_dark'])
        hdr_fc     = C['vac_txt'] if is_vacataire(agent) else 'FFFFFF'

        # ── En-tête agent ───────────────────────────────────────
        HDR_GREEN = '1A6B3A'
        ws.merge_cells(f'A{row}:{end_col}{row}')
        sects_label = sects[0] if sects else '?'
        _set(ws.cell(row=row, column=1,
                     value=f"  {agent.upper()}   —   {sects_label}"),
             bg=HDR_GREEN, fnt=_font(bold=True, size=12, color='FFFFFF'),
             aln=_aln('left'))
        # Colonne H : total heures travaillées sur la semaine
        total_h = total_heures_semaine(agent)
        h_cell = ws.cell(row=row, column=col_h_total)
        h_cell.value = f"⏱ {total_h}"
        h_cell.fill = _fill(HDR_GREEN)
        h_cell.font = _font(bold=True, size=10, color='A8E6BE')
        h_cell.alignment = _aln('center')
        ws.row_dimensions[row].height = 26
        row += 1

        # ── En-tête colonnes ────────────────────────────────────
        ws.cell(row=row, column=1).fill = _fill(C['hdr_dark2'])
        ws.cell(row=row, column=1).value = ''
        _set(ws.cell(row=row, column=2, value='Créneau'),
             bg=C['hdr_dark2'], fnt=_font(bold=True, size=10, color='FFFFFF'),
             aln=_aln(), brd=_brd())
        for jour in jours_presents:
            c    = col_map[jour]
            date = week_dates.get(jour)
            sam_t = plan.get(jour, {}).get('_samedi_type')
            lbl  = f"{jour[:3]}\n{date.day}/{date.month}"
            if sam_t:
                lbl += f"\n{sam_t}"
            bg_h = (C['hdr_rouge'] if sam_t == 'ROUGE'
                    else C['hdr_bleu'] if sam_t == 'BLEU'
                    else C['hdr_dark2'])
            _set(ws.cell(row=row, column=c, value=lbl),
                 bg=bg_h, fnt=_font(bold=True, size=9, color='FFFFFF'),
                 aln=_aln('center'), brd=_brd())
        ws.row_dimensions[row].height = 36
        row += 1

        # ── Lignes d'arrivée (8h30-9h et 9h-10h) ───────────────
        # Même formatage colonne B que les créneaux SP
        early_start_row = row  # pour la fusion sur la plage complète
        for slot_name, es, ee in EARLY_SLOTS:
            _set(ws.cell(row=row, column=1), bg='F8F9FA')
            ws.cell(row=row, column=1).value = ''
            # Colonne B : même style que les créneaux SP normaux
            _set(ws.cell(row=row, column=2, value=slot_name),
                 bg='FDFEFE', fnt=_font(bold=True, size=9),
                 aln=_aln(), brd=_brd())

            for jour in jours_presents:
                c        = col_map[jour]
                date     = week_dates.get(jour)
                date_str = date.strftime('%Y-%m-%d')
                cell     = ws.cell(row=row, column=c)

                # Jour férié → hachures
                if plan.get(jour, {}).get('_ferie'):
                    cell.fill = _hatch()
                    cell.value = None
                    cell.alignment = _aln()
                    cell.border = _brd()
                    continue

                # Vérifier congé sur la journée entière
                # (congé = événement couvrant le début de journée ou le créneau)
                ev_all = evenements.get(date_str, [])
                ev_agent = [ev for ev in ev_all if agent in ev['agents']]
                is_conge = any(
                    n.lower() in ('congé','conge','rtt','vacation')
                    for ev in ev_agent
                    for n in [ev['nom']]
                )

                # Horaires habituels de l'agent ce jour
                h  = horaires_ag.get(agent, {}).get(jour)
                sm = h[0] if h else None   # début matin (en minutes)

                # Événements non-congé pendant ce créneau d'arrivée
                ev_early = [ev for ev in ev_agent
                            if ev['debut'] < ee and ev['fin'] > es
                            and ev['nom'].lower() not in ('congé','conge','rtt','vacation')]

                if is_conge:
                    _set(cell, value='Congé',
                         bg=C['event_bg'],
                         fnt=_font(size=9, color=_ev_txt_color(), italic=True),
                         aln=_aln(), brd=_brd())

                elif ev_early:
                    ev_lbl = ev_early[0]['nom'][:20]
                    _set(cell, value=ev_lbl,
                         bg=C['event_bg'],
                         fnt=_font(size=9, italic=True, color=_ev_txt_color()),
                         aln=_aln(), brd=_brd())

                elif sm is not None and es <= sm < ee:
                    if sm == es:
                        _set(cell, value='',
                             bg='F0F0F0',
                             fnt=_font(size=8, color=C['gray']),
                             aln=_aln(), brd=_brd())
                    else:
                        _set(cell, value=f"Arr. {min_to_hhmm(sm)}",
                             bg='F0F0F0',
                             fnt=_font(size=9, italic=True, color='555555'),
                             aln=_aln(), brd=_brd())

                elif sm is not None and sm < es:
                    _set(cell, value='',
                         bg='F0F0F0',
                         fnt=_font(size=8, color=C['gray']),
                         aln=_aln(), brd=_brd())
                    cell.border = Border(
                        left=Side(style='medium', color='AAAAAA'),
                        right=Side(style='thin', color='CCCCCC'),
                        top=Side(style='thin', color='CCCCCC'),
                        bottom=Side(style='thin', color='CCCCCC'))

                else:
                    cell.fill = _hatch()
                    cell.value = None
                    cell.font = _font(size=10, color=C['off_txt'])
                    cell.alignment = _aln()
                    cell.border = _brd()

            ws.row_dimensions[row].height = 26
            row += 1

        # ── Lignes créneaux SP (10h → 19h) ─────────────────────
        cren_start_row = row  # première ligne SP (pour la formule TOTAL)

        for cren_idx, (cren_name, cs, ce) in enumerate(creneaux):

            dur_h = round((ce - cs) / 60, 4)
            dur_cell = ws.cell(row=row, column=1)
            dur_cell.value = dur_h
            dur_cell.number_format = '0.00'
            dur_cell.fill = _fill('F8F9FA')
            dur_cell.font = _font(size=8, color='F8F9FA')
            dur_cell.alignment = _aln()
            _set(ws.cell(row=row, column=2, value=cren_name),
                 bg='FDFEFE', fnt=_font(bold=True, size=9), aln=_aln(), brd=_brd())

            for jour in jours_presents:
                c        = col_map[jour]
                date     = week_dates.get(jour)
                date_str = date.strftime('%Y-%m-%d')
                cell     = ws.cell(row=row, column=c)
                slot     = plan.get(jour, {}).get(cren_name)

                # Jour férié → hachures
                if plan.get(jour, {}).get('_ferie'):
                    cell.fill = _hatch()
                    cell.value = None
                    cell.alignment = _aln()
                    cell.border = _brd()
                    continue

                ev_names = [ev['nom'] for ev in evenements.get(date_str, [])
                            if agent in ev['agents']
                            and ev['debut'] <= cs < ev['fin']]
                is_conge = any(n.lower() in ('congé','conge','rtt','vacation')
                               for n in ev_names)
                event    = [n for n in ev_names
                            if n.lower() not in ('congé','conge','rtt','vacation')]

                h = horaires_ag.get(agent, {}).get(jour)
                in_horaires = False
                if h:
                    sm,em,sa,ea = h
                    in_m = sm is not None and em is not None and cs >= sm and ce <= em
                    in_a = sa is not None and ea is not None and cs >= sa and ce <= ea
                    in_horaires = in_m or in_a

                sp_section  = None
                is_oos_cell = False
                if slot and isinstance(slot, dict) and slot.get('open'):
                    for sect in SECTIONS:
                        if agent in slot.get('assignment', {}).get(sect, []):
                            sp_section = sect
                            break
                    if sp_section:
                        is_oos_cell = slot.get('out_of_section', {}).get(sp_section, False)

                if is_conge:
                    _set(cell, value='Congé',
                         bg=C['event_bg'],
                         fnt=_font(size=9, color=_ev_txt_color(), italic=True),
                         aln=_aln(), brd=_brd())
                elif event:
                    ev_label = event[0][:20]
                    _set(cell, value=ev_label,
                         bg=C['event_bg'],
                         fnt=_font(size=9, italic=True, color=_ev_txt_color()),
                         aln=_aln(), brd=_brd())
                elif sp_section and is_oos_cell:
                    sec_label = 'M & F' if sp_section == 'MF' else sp_section
                    _set(cell, value=f"SP\n{sec_label} ⚠",
                         bg=C['oos_bg'],
                         fnt=_font(size=9, bold=True, color=C['oos_txt']),
                         aln=_aln(), brd=_brd())
                elif not in_horaires:
                    # Hors horaires → hachures + sentinelle '·'
                    cell.fill  = _hatch()
                    cell.value = '·'
                    cell.font  = _font(size=6, color='F5F5F5')
                    cell.alignment = _aln()
                    cell.border    = _brd()
                else:
                    # SP ou bureau → formule IF dynamique cross-sheet
                    # 1 cellule = 1 créneau, pas de fusion (le créneau est dans col B)
                    if jour_cren_rows and jour in jour_cren_rows:
                        sem_sht = f'Semaine_{week_num}'
                        ag_esc  = agent.replace('"', '""')
                        r_sem   = jour_cren_rows[jour][0] + cren_idx
                        # IF imbriqués : B=RDC, C=Adulte, D=MF, E=Jeunesse
                        formula = (
                            '=IF(ISNUMBER(SEARCH("' + ag_esc + '",\'' + sem_sht + '\'!B' + str(r_sem) + ')),"SP"&CHAR(10)&"RDC",'
                            'IF(ISNUMBER(SEARCH("' + ag_esc + '",\'' + sem_sht + '\'!C' + str(r_sem) + ')),"SP"&CHAR(10)&"Adulte",'
                            'IF(ISNUMBER(SEARCH("' + ag_esc + '",\'' + sem_sht + '\'!D' + str(r_sem) + ')),"SP"&CHAR(10)&"M & F",'
                            'IF(ISNUMBER(SEARCH("' + ag_esc + '",\'' + sem_sht + '\'!E' + str(r_sem) + ')),"SP"&CHAR(10)&"Jeunesse",""))))'
                        )
                        cell.value     = formula
                        cell.alignment = _aln()
                        cell.border    = _brd()
                        # Fond initial selon section connue du moteur
                        # CF la mettra à jour si l'agent change de section dans Semaine_N
                        if sp_section:
                            cell.fill = _fill(SEC_BG.get(sp_section, 'FFFFFF'))
                            cell.font = _font(size=9, bold=True, color='1A1A1A')
                        else:
                            # Au bureau : fond gris statique (pas de CF bureau)
                            cell.fill = _fill('F0F0F0')
                            cell.font = _font(size=8, color='888888')
                    else:
                        # Fallback statique
                        if sp_section:
                            sec_label = 'M & F' if sp_section == 'MF' else sp_section
                            _set(cell, value=f"SP\n{sec_label}",
                                 bg=SEC_BG.get(sp_section, 'FFFFFF'),
                                 fnt=_font(size=9, bold=True, color='1A1A1A'),
                                 aln=_aln(), brd=_brd())
                        else:
                            cell.value = ''
                            cell.fill  = _fill('F0F0F0')
                            cell.font  = _font(size=8, color=C['gray'])
                            cell.alignment = _aln()
                            cell.border    = _brd()

            ws.row_dimensions[row].height = 28
            row += 1

        cren_end_row = row - 1  # dernière ligne SP (inclusive)

        # ── Mise en forme conditionnelle sur la plage créneaux ──────
        if jours_presents and jour_cren_rows:
            col_start = col_map[jours_presents[0]]
            col_end   = col_map[jours_presents[-1]]
            cf_range  = (f"{get_column_letter(col_start)}{cren_start_row}:"
                         f"{get_column_letter(col_end)}{cren_end_row}")
            anc = f"{get_column_letter(col_start)}{cren_start_row}"

            # CF sections : colore quand la formule IF retourne "SP\nXxx"
            # Fond initial = couleur section (si SP connu du moteur), sinon gris
            # → pas de règle CF pour "" (bureau) car elle écraserait les fonds section
            sect_rules = [
                (SEC_BG['RDC'],      f'ISNUMBER(SEARCH("RDC",{anc}))'),
                (SEC_BG['Adulte'],   f'ISNUMBER(SEARCH("Adulte",{anc}))'),
                (SEC_BG['MF'],       f'ISNUMBER(SEARCH("M & F",{anc}))'),
                (SEC_BG['Jeunesse'], f'ISNUMBER(SEARCH("Jeunesse",{anc}))'),
            ]
            for color, formula in sect_rules:
                ws.conditional_formatting.add(cf_range, FormulaRule(
                    formula=[formula],
                    fill=PatternFill(fill_type='solid', fgColor=color),
                    font=Font(size=9, bold=True, color='1A1A1A'),
                    stopIfTrue=True
                ))

        # ── Créneau tardif (après 19h) ──────────────────────────
        late_name, ls, le = LATE_SLOT
        has_late = False
        for jour in jours_presents:
            h = horaires_ag.get(agent, {}).get(jour)
            if h:
                ea = h[3]  # fin après-midi
                if ea is not None and ea > 1140:
                    has_late = True
                    break

        if has_late:
            _set(ws.cell(row=row, column=1), bg='F8F9FA')
            ws.cell(row=row, column=1).value = ''
            _set(ws.cell(row=row, column=2, value=late_name),
                 bg='FDFEFE', fnt=_font(bold=True, size=9),
                 aln=_aln(), brd=_brd())
            for jour in jours_presents:
                c    = col_map[jour]
                cell = ws.cell(row=row, column=c)
                h    = horaires_ag.get(agent, {}).get(jour)
                ea   = h[3] if h else None
                date_str_late = date.strftime('%Y-%m-%d')
                ev_late_agent = [ev for ev in evenements.get(date_str_late, [])
                                 if agent in ev['agents']
                                 and ev['debut'] < le and ev['fin'] > ls
                                 and ev['nom'].lower() not in ('congé','conge','rtt','vacation')]

                if ev_late_agent:
                    ev_lbl = ev_late_agent[0]['nom'][:20]
                    _set(cell, value=ev_lbl,
                         bg=C['event_bg'],
                         fnt=_font(size=9, italic=True, color=_ev_txt_color()),
                         aln=_aln(), brd=_brd())
                elif ea is not None and ea > 1140:
                    _set(cell, value=f"Fin {min_to_hhmm(ea)}",
                         bg='F0F0F0',
                         fnt=_font(size=9, italic=True, color='555555'),
                         aln=_aln(), brd=_brd())
                elif ea is not None and ea == 1140:
                    _set(cell, value='',
                         bg='F0F0F0',
                         fnt=_font(size=8, color=C['gray']),
                         aln=_aln(), brd=_brd())
                else:
                    cell.fill = _hatch()
                    cell.value = None
                    cell.font = _font(size=10, color=C['off_txt'])
                    cell.alignment = _aln()
                    cell.border = _brd()
            ws.row_dimensions[row].height = 26
            row += 1

        # ── Ligne TOTAL SP ──────────────────────────────────────
        _set(ws.cell(row=row, column=1), bg='1A6B3A')
        ws.cell(row=row, column=1).value = ''
        _set(ws.cell(row=row, column=2, value='TOTAL SP'),
             bg='1A6B3A', fnt=_font(bold=True, size=10, color='FFFFFF'),
             aln=_aln(), brd=_brd())

        agent_sp_cells[agent] = {}

        # ── TOTAL SP : formule dynamique cross-sheet ──────────────
        # Cherche le nom de l'agent dans Semaine_N colonnes C(RDC),D(Adulte),E(MF),F(Jeunesse)
        # Multiplie par les durées en col A (heures décimales, police blanche)
        # Entièrement recalculé par Excel si Semaine_N est modifié

        sem_sheet = f"Semaine_{week_num}"   # nom de l'onglet source
        # Sections = colonnes B,C,D,E dans Semaine_N (après suppression col A N°)
        SEC_COLS_SEM = ['B', 'C', 'D', 'E']

        for jour in jours_presents:
            c          = col_map[jour]
            col_ltr    = get_column_letter(c)
            total_cell = ws.cell(row=row, column=c)
            sp_j       = sp_jour_cum.get(jour, {}).get(agent, 0)

            if week_dates.get(jour) is not None:
                # Lignes créneaux dans Semaine_N pour ce jour
                if jour_cren_rows and jour in jour_cren_rows:
                    r1, r2 = jour_cren_rows[jour]
                else:
                    # Fallback : valeur statique moteur
                    sp_h = round(sp_j / 60, 2)
                    total_cell.value = sp_h
                    total_cell.number_format = '0.00'
                    agent_sp_cells[agent][jour] = f"{col_ltr}{row}"
                    bg_t = 'D5F5E3' if sp_j > 0 else 'F0F3F4'
                    fc_t = '1E8449' if sp_j > 0 else C['gray']
                    total_cell.fill = _fill(bg_t)
                    total_cell.font = _font(bold=sp_j>0, size=9, color=fc_t)
                    total_cell.alignment = _aln()
                    total_cell.border = _brd()
                    continue

                # =SUMPRODUCT(
                #   ((ISNUMBER(SEARCH("agent",Sem!C6:C16))
                #    +ISNUMBER(SEARCH("agent",Sem!D6:D16))
                #    +ISNUMBER(SEARCH("agent",Sem!E6:E16))
                #    +ISNUMBER(SEARCH("agent",Sem!F6:F16)))>0)
                #   * A{cren_start}:A{cren_end})
                search_parts = '+'.join(
                    'ISNUMBER(SEARCH("' + agent + '",' + "'" + sem_sheet + "'!" + sc + str(r1) + ':' + sc + str(r2) + '))'
                    for sc in SEC_COLS_SEM
                )
                formula = (
                    '=IFERROR(SUMPRODUCT((' + search_parts + '>0)'
                    '*A' + str(cren_start_row) + ':A' + str(cren_end_row) + '),0)'
                )
                total_cell.value         = formula
                total_cell.number_format = '0.00'

                bg_t = '155C30' if sp_j > 0 else '1A6B3A'
                fc_t = 'A8E6BE' if sp_j > 0 else '7FC89A'
                total_cell.fill      = _fill(bg_t)
                total_cell.font      = _font(bold=sp_j > 0, size=10, color=fc_t)
                total_cell.alignment = _aln()
                total_cell.border    = _brd()

                # Stocker la référence pour le récap cross-sheet Semaine_N
                agent_sp_cells[agent][jour] = f"{col_ltr}{row}"
            else:
                _set(total_cell, value='—',
                     bg='1A6B3A',
                     fnt=_font(size=9, color='7FC89A'),
                     aln=_aln(), brd=_brd())

        ws.row_dimensions[row].height = 24
        row += 1

        # ── Fusion cellules consécutives par colonne jour ──────
        # Couvre EARLY slots + SP créneaux + LATE slot
        # Pas de fusion — 1 ligne = 1 créneau pour que les formules IF restent dynamiques

        # ── Séparateur ──────────────────────────────────────────
        for ci in range(1, nb_cols + 1):
            ws.cell(row=row, column=ci).fill = _fill('C0C0C0')
        ws.row_dimensions[row].height = 3
        row += 2

    return ws, agent_sp_cells




# ══════════════════════════════════════════════════════════════
#  POST-PROCESSING : inlineStr → sharedStrings
# ══════════════════════════════════════════════════════════════

def _xml_decode(s):
    """Decode XML entities (named + numeric) via html.unescape + newline."""
    import html
    # html.unescape gère &amp; &lt; &gt; &quot; + toutes les entites numeriques &#NNN; &#xHHH;
    # On remplace d'abord &#10; par un placeholder pour le newline
    return html.unescape(s)


def _xml_encode(s):
    """Encode string for XML."""
    return (s.replace('&', '&amp;')
             .replace('<', '&lt;')
             .replace('>', '&gt;')
             .replace('"', '&quot;')
             .replace('\n', '&#10;'))


def _convert_inlinestr_to_shared_strings(buf, pa_sheet_names, all_sp_cells=None, weeks_data=None):
    """
    Convertit les cellules inlineStr des onglets Planning_Agent en sharedStrings.
    openpyxl ecrit tout en inlineStr ; SEARCH() ne fonctionne pas sur ce format.
    Apres conversion, les formules SUMPRODUCT+SEARCH deviennent dynamiques.

    Transformation :
      <c r="C8" s="42" t="inlineStr"><is><t>SP&#10;Jeunesse</t></is></c>
      => <c r="C8" s="42" t="s"><v>17</v></c>
    """
    buf.seek(0)
    in_bytes = buf.read()

    # Lire les manifestes
    with zipfile.ZipFile(io.BytesIO(in_bytes), 'r') as zin:
        wb_xml   = zin.read('xl/workbook.xml').decode()
        rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode()
        ct_xml   = zin.read('[Content_Types].xml').decode()
        all_files = zin.namelist()

    # Mapping sheet name -> chemin ZIP
    sheet_names = re.findall(r'<sheet [^>]*name="([^"]+)"', wb_xml)
    sheet_rids  = re.findall(r'<sheet [^>]*r:id="([^"]+)"', wb_xml)
    rel_ids     = re.findall(r'<Relationship[^>]*\sId="([^"]+)"', rels_xml)
    rel_targets = re.findall(r'<Relationship[^>]*\sTarget="([^"]+)"', rels_xml)
    rid_to_target = dict(zip(rel_ids, rel_targets))
    name_to_path = {}
    for sname, rid in zip(sheet_names, sheet_rids):
        target = rid_to_target.get(rid, '')
        if target:
            name_to_path[sname] = target.lstrip('/')

    pa_paths = {sn: name_to_path[sn] for sn in pa_sheet_names if sn in name_to_path}

    # Construire table shared strings (partir de l'existant si present)
    str_table = []
    str_index = {}

    if 'xl/sharedStrings.xml' in all_files:
        with zipfile.ZipFile(io.BytesIO(in_bytes), 'r') as zin:
            ss_xml = zin.read('xl/sharedStrings.xml').decode()
        for m in re.finditer(r'<si><t[^>]*>(.*?)</t></si>', ss_xml, re.DOTALL):
            txt = _xml_decode(m.group(1))
            if txt not in str_index:
                str_index[txt] = len(str_table)
                str_table.append(txt)

    # Collecter tous les inlineStr des feuilles PA
    with zipfile.ZipFile(io.BytesIO(in_bytes), 'r') as zin:
        for sname, path in pa_paths.items():
            if path not in all_files:
                continue
            xml = zin.read(path).decode()
            for m in re.finditer(
                    r'<c [^>]*t="inlineStr"[^>]*><is><t[^>]*>(.*?)</t></is></c>',
                    xml, re.DOTALL):
                txt = _xml_decode(m.group(1))
                if txt not in str_index:
                    str_index[txt] = len(str_table)
                    str_table.append(txt)

    # Construire sharedStrings.xml
    NS = 'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
    si_parts = []
    for s in str_table:
        s_enc = _xml_encode(s)
        space = ' xml:space="preserve"' if (' ' in s or '\n' in s) else ''
        si_parts.append('<si><t' + space + '>' + s_enc + '</t></si>')
    n = len(str_table)
    new_ss_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                  '<sst ' + NS + ' count="' + str(n) + '" uniqueCount="' + str(n) + '">'
                  + ''.join(si_parts) + '</sst>')

    # Convertir les cellules inlineStr -> t="s" dans chaque feuille PA
    def convert_sheet_xml(xml):
        def repl(m):
            c_open = m.group(1)   # ex: <c r="C8" s="42" t="inlineStr">
            txt    = _xml_decode(m.group(2))
            idx    = str_index.get(txt, 0)
            # Retirer t="inlineStr" ET le > final pour reconstruire proprement
            c_open = re.sub(r'\s*t="inlineStr"', '', c_open)
            c_open = c_open.rstrip('>').rstrip()
            return c_open + ' t="s"><v>' + str(idx) + '</v></c>'
        return re.sub(
            r'(<c [^>]*t="inlineStr"[^>]*>)<is><t[^>]*>(.*?)</t></is></c>',
            repl, xml, flags=re.DOTALL)

    # Mettre a jour Content_Types et rels si necessaire
    SS_CT  = ('application/vnd.openxmlformats-officedocument'
              '.spreadsheetml.sharedStrings+xml')
    SS_REL = ('http://schemas.openxmlformats.org/officeDocument'
              '/2006/relationships/sharedStrings')

    if 'sharedStrings' not in ct_xml:
        ct_xml = ct_xml.replace(
            '</Types>',
            '<Override PartName="/xl/sharedStrings.xml"'
            ' ContentType="' + SS_CT + '"/></Types>')

    if 'sharedStrings' not in rels_xml:
        max_rid = max((int(r) for r in re.findall(r'Id="rId(\d+)"', rels_xml)),
                      default=0)
        new_rid = 'rId' + str(max_rid + 1)
        rels_xml = rels_xml.replace(
            '</Relationships>',
            '<Relationship Id="' + new_rid + '" Type="' + SS_REL + '"'
            ' Target="sharedStrings.xml"/></Relationships>')

    # Construire mapping cellule -> valeur calculee pour les formules TOTAL SP
    # Cela permet a Excel d'afficher les valeurs immediatement a l'ouverture
    # (sans attendre un recalcul manuel), tout en gardant la formule dynamique
    cached_vals = {}  # {pa_sheet_name: {cell_ref: float}}
    if all_sp_cells and weeks_data:
        for wk in weeks_data:
            wn  = wk['week_num']
            sn  = f"Planning_Agent_Semaine_{wn}"
            sp_jour_cum = wk.get('sp_jour_cum', {})
            asc = all_sp_cells.get(wn, {})
            cached_vals[sn] = {}
            for agent, jour_map in asc.items():
                for jour, cell_ref in jour_map.items():
                    val = round(sp_jour_cum.get(jour, {}).get(agent, 0) / 60, 4)
                    cached_vals[sn][cell_ref] = val

    def inject_cached(xml, sheet_name):
        """Injecte les valeurs calculees dans <v></v> des cellules formule."""
        vals = cached_vals.get(sheet_name, {})
        for cell_ref, val in vals.items():
            vs  = f"{val:.4f}".rstrip('0').rstrip('.') or '0'
            esc = re.escape(cell_ref)
            xml = re.sub(
                r'(<c r="' + esc + r'"[^>]*><f>[^<]*</f>)<v>[^<]*</v>(</c>)',
                r'\g<1><v>' + vs + r'</v>\g<2>',
                xml
            )
        return xml

    # Reecrire le ZIP
    out_buf = io.BytesIO()
    ss_written = False
    with zipfile.ZipFile(io.BytesIO(in_bytes), 'r') as zin, \
         zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as zout:

        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == '[Content_Types].xml':
                data = ct_xml.encode('utf-8')
            elif item.filename == 'xl/_rels/workbook.xml.rels':
                data = rels_xml.encode('utf-8')
            elif item.filename == 'xl/sharedStrings.xml':
                data = new_ss_xml.encode('utf-8')
                ss_written = True
            else:
                for sname, path in pa_paths.items():
                    if item.filename == path:
                        xml_str = convert_sheet_xml(data.decode('utf-8'))
                        xml_str = inject_cached(xml_str, sname)
                        data = xml_str.encode('utf-8')
                        break
            zout.writestr(item, data)

        # Ajouter sharedStrings.xml s'il n'existait pas
        if not ss_written:
            zout.writestr('xl/sharedStrings.xml', new_ss_xml.encode('utf-8'))

    out_buf.seek(0)
    return out_buf


def _patch_recap_formulas(wb, week_data, metadata, agent_sp_cells):
    """
    Passe 3 : réécrit les cellules du récap SP dans Semaine_N
    avec des formules cross-sheet pointant vers Planning_Agent_Semaine_N.
    Ex : C77 = ='Planning_Agent_Semaine_1'!C19
         H77 = =SUM(C77:G77)
    """
    week_num     = week_data['week_num']
    week_dates   = week_data['week_dates']
    sp_jour_cum  = week_data.get('sp_jour_cum', {})
    sp_alerts    = week_data.get('sp_alerts', {})
    affectations = metadata.get('affectations', {})
    sp_minmax_all  = metadata.get('sp_minmax_all', {})
    sp_minmax_week = sp_minmax_all.get(week_num, sp_minmax_all.get(2, {}))
    is_sam_week  = 'Samedi' in week_dates

    sname = f"Semaine_{week_num}"
    pa_sheet = f"Planning_Agent_Semaine_{week_num}"
    if sname not in wb.sheetnames:
        return
    ws = wb[sname]

    # Trouver les lignes du récap (chercher l'en-tête "Agent")
    recap_hdr_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == 'Agent':
            recap_hdr_row = r
            break
    if recap_hdr_row is None:
        return

    # Agents dans le même ordre que write_week_sheet
    def primary_section(agent):
        sects = affectations.get(agent, [])
        return sects[0] if sects else '?'

    agents_shown = sorted(
        [a for a in sp_minmax_week.keys() if not is_vacataire(a)],
        key=lambda x: (primary_section(x), x)
    ) + [a for a in affectations.keys() if is_vacataire(a)]

    data_start_row = recap_hdr_row + 1

    for i, agent in enumerate(agents_shown):
        r   = data_start_row + i
        bg  = C['alt1'] if i % 2 == 0 else C['alt2']
        sp_total_static = 0.0

        for ji, jour in enumerate(JOURS_SP):
            col  = ji + 3   # C=3, D=4, E=5, F=6, G=7
            cell = ws.cell(row=r, column=col)
            sp_min = sp_jour_cum.get(jour, {}).get(agent, 0)
            sp_h   = min_to_dec(sp_min)
            sp_total_static += sp_h

            if week_dates.get(jour) is not None:
                cell_ref = (agent_sp_cells.get(agent, {}).get(jour)
                            if agent_sp_cells else None)
                if cell_ref:
                    cell.value = f"='{pa_sheet}'!{cell_ref}"
                else:
                    cell.value = sp_h if sp_h > 0 else 0
                cell.number_format = '0.00'

                bg_c = ('FFF9C4' if jour == 'Samedi' and sp_h > 0
                         else C['alt2'] if sp_h > 0 else bg)
                cell.fill      = _fill(bg_c)
                cell.font      = _font(size=10, bold=sp_h > 0)
                cell.alignment = _aln()
                cell.border    = _brd()
            # Jours absents : laisser tel quel (déjà '—')

        # Colonne L : SUM dynamique (col 12, masquable entre jours et total)
        tc = ws.cell(row=r, column=12)
        sum_cols = [get_column_letter(3 + ji)
                    for ji, jour in enumerate(JOURS_SP)
                    if week_dates.get(jour) is not None]
        if sum_cols:
            tc.value = f"=SUM({','.join(c + str(r) for c in sum_cols)})"
        tc.number_format = '0.00'

        # Couleur colonne H (basée sur valeur statique moteur)
        total_h = sp_total_static
        if not is_vacataire(agent):
            mm     = sp_minmax_week.get(agent, {})
            min_sp = mm.get('Min_MarSam', 0) if is_sam_week else mm.get('Min_MarVen', 0)
            max_sp = mm.get('Max_MarSam', 99) if is_sam_week else mm.get('Max_MarVen', 99)
            if agent in sp_alerts:
                t_bg, t_fc = C['sp_alert'], C['sp_alert_t']
            elif total_h < min_sp:
                t_bg, t_fc = C['sp_alert'], C['sp_alert_t']
            elif total_h > max_sp:
                t_bg, t_fc = C['sp_warn'], C['sp_warn_t']
            else:
                t_bg, t_fc = C['sp_ok'], C['sp_ok_t']
        else:
            t_bg = C['vac_bg'] if total_h > 0 else bg
            t_fc = C['vac_txt'] if total_h > 0 else C['gray']

        tc.fill      = _fill(t_bg)
        tc.font      = _font(bold=True, size=10, color=t_fc)
        tc.alignment = _aln()
        tc.border    = _brd()


# ══════════════════════════════════════════════════════════════
#  POINT D'ENTRÉE
# ══════════════════════════════════════════════════════════════

def generate_excel(source_filepath, weeks_data, metadata):
    wb = load_workbook(source_filepath)

    # Supprimer anciens onglets générés
    to_remove = [s for s in wb.sheetnames
                 if s.startswith('Semaine_') or s.startswith('Planning_Agent')]
    for name in to_remove:
        wb.remove(wb[name])

    # Passe 1 : onglets Semaine → récupère les numéros de lignes créneaux
    all_jour_cren_rows = {}
    for week_data in weeks_data:
        _ws, jcr = write_week_sheet(wb, week_data, metadata)
        all_jour_cren_rows[week_data['week_num']] = jcr

    # Passe 2 : onglets Planning_Agent → formules cross-sheet vers Semaine
    #           récupère agent_sp_cells = {agent: {jour: 'C19'}}
    all_sp_cells = {}
    for week_data in weeks_data:
        jcr = all_jour_cren_rows.get(week_data['week_num'], {})
        _, asc = write_planning_agent_week_sheet(wb, week_data, metadata, jour_cren_rows=jcr)
        all_sp_cells[week_data['week_num']] = asc

    # Passe 3 : injecter les formules cross-sheet dans le récap de chaque Semaine
    #           maintenant qu'on connaît les refs Planning_Agent (C19, C37, ...)
    for week_data in weeks_data:
        asc = all_sp_cells.get(week_data['week_num'], {})
        _patch_recap_formulas(wb, week_data, metadata, asc)

    # Réordonner : Semaine_1..5 → Planning_Agent_1..5 → autres
    week_sheets  = [f"Semaine_{i}" for i in range(1, len(weeks_data)+1)
                    if f"Semaine_{i}" in wb.sheetnames]
    agent_sheets = [f"Planning_Agent_Semaine_{i}" for i in range(1, len(weeks_data)+1)
                    if f"Planning_Agent_Semaine_{i}" in wb.sheetnames]
    other_sheets = [s for s in wb.sheetnames
                    if s not in week_sheets and s not in agent_sheets]
    wb._sheets   = ([wb[s] for s in week_sheets]
                    + [wb[s] for s in agent_sheets]
                    + [wb[s] for s in other_sheets])


    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    buf = io.BytesIO()
    wb.save(buf)

    # Note : _convert_inlinestr_to_shared_strings n'est plus appelée.
    # Les formules cross-sheet (IF/SEARCH vers Semaine_N) fonctionnent
    # nativement sans conversion inlineStr — les cellules Semaine_N
    # sont des sharedStrings dès que Excel recalcule le fichier.

    buf.seek(0)
    return buf

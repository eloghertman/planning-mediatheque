"""
planning_engine_cpsat.py
Moteur de calcul planning médiathèque — CP-SAT (Google OR-Tools)
Remplace l'ancien moteur glouton.
"""

import datetime
import re
import unicodedata
from collections import defaultdict

import openpyxl
from ortools.sat.python import cp_model

# ══════════════════════════════════════════════════════════════
#  CONSTANTES
# ══════════════════════════════════════════════════════════════

SECTIONS     = ['RDC', 'Adulte', 'MF', 'Jeunesse']
JOURS_SEMAINE = ['Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
JOURS_VAC    = {'Mercredi', 'Samedi'}   # jours où les vacataires peuvent travailler

# Poids des contraintes molles (plus = plus prioritaire)
# Nettoyé 08/2026 : les entrées H1/H2/H3/K1/K2/I2 ont été supprimées car mortes
# (remplacées par le bonus vacataire V1/V2 et l'équité en 2 passes, tous deux
# codés directement dans solve_day plutôt que via ce dictionnaire partagé).
POIDS = {
    'G1_planning_type':        100,   # respecter le planning type
    'G2_meme_section_repl':     50,   # remplaçant même section principale
    'J1_section_principale':    30,   # agent dans sa section principale
    'J3_responsable':           25,   # responsables déprioritisés
    'I1_non_fragmentation':     20,   # blocs continus préférés
}


# ══════════════════════════════════════════════════════════════
#  UTILITAIRES TEMPS
# ══════════════════════════════════════════════════════════════

def hhmm_to_min(t):
    """Convertit un objet time, une string 'HH:MM' ou 'HH:MM:SS' en minutes depuis minuit."""
    if t is None or (isinstance(t, str) and t.strip() in ('', '\xa0')):
        return None
    if isinstance(t, datetime.time):
        return t.hour * 60 + t.minute
    if isinstance(t, str):
        t = t.strip().replace('h', ':').replace('H', ':')
        parts = t.split(':')
        return int(parts[0]) * 60 + int(parts[1]) if len(parts) >= 2 else None
    return None


def parse_creneau(s):
    """Convertit '10:00-12:30' ou '10H-12H30' en (cs, ce) en minutes. None si invalide."""
    if not s:
        return None
    s = str(s).strip()
    if '-' not in s:
        return None

    def to_min(t):
        t = t.strip().upper()
        m = re.match(r'(\d{1,2})H(\d{0,2})', t)
        if m:
            return int(m.group(1)) * 60 + (int(m.group(2)) if m.group(2) else 0)
        if ':' in t:
            p = t.split(':')
            try:
                return int(p[0]) * 60 + int(p[1])
            except (ValueError, IndexError):
                return None
        return None

    parts = s.split('-', 1)
    cs = to_min(parts[0])
    ce = to_min(parts[1])
    if cs is None or ce is None or ce <= cs:
        return None
    return (cs, ce)


def load_excel_data(filepath):
    """Charge le fichier Excel et retourne un dict {nom_onglet: worksheet}."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    return {ws.title: ws for ws in wb.worksheets}


def parse_parametres(raw):
    """
    Retourne un dict avec les paramètres du mois.
    Gère deux listes de créneaux séparées (Mar/Jeu/Ven vs Mer/Sam)
    et le tableau "Présence Vacataire".
    """
    ws = raw['Paramètres']
    rows = list(ws.iter_rows(values_only=True))

    # ── Lecture des paramètres clé-valeur (col A → col B)
    params = {}
    presence_vac_start = None
    for i, row in enumerate(rows):
        if row[0] and row[1] is not None:
            params[str(row[0]).strip()] = row[1]
        # Détecter le début du tableau Présence Vacataire
        if row[0] and 'Présence' in str(row[0]) and 'Vacataire' in str(row[0]):
            presence_vac_start = i

    # ── Créneaux : supporte liste unique OU deux listes séparées
    def parse_liste_creneaux(key):
        val = params.get(key, '')
        result = []
        for c in str(val).split(';'):
            parsed = parse_creneau(c.strip())
            if parsed:
                result.append(parsed)
        return result

    creneaux_mjv = parse_liste_creneaux('Liste_des_créneaux_mardi_jeudi_vendredi') or                    parse_liste_creneaux('Liste_des_créneaux')
    creneaux_ms  = parse_liste_creneaux('Liste_des_créneaux_mercredi_samedi') or creneaux_mjv

    # Créneaux unifiés (union, triés) pour les jours sans liste spécifique
    creneaux_all = sorted(set(creneaux_mjv) | set(creneaux_ms))

    # ── Samedis et semaines
    samedis = {}
    for i in range(1, 6):
        v = params.get(f'Samedi_{i}')
        if v: samedis[i] = str(v).strip().upper()

    semaines = {}
    for i in range(1, 6):
        v = params.get(f'Semaine_{i}')
        if v: semaines[i] = str(v).strip()

    # ── Mode vacataires (fallback si pas de tableau Présence)
    mode_vac_str = str(params.get('Mode_vacataires', 'samedi')).lower()
    mode_vac = set()
    for j in JOURS_SEMAINE:
        if j.lower() in mode_vac_str:
            mode_vac.add(j)

    # ── Tableau Présence Vacataire
    # Format : Date (col B) | Vacataire (col C) | Heure début (col D) | Heure fin (col E)
    presences_vac = {}  # {date_str: {agent: (cs, ce)}}
    if presence_vac_start is not None:
        # Corrigé 09/2026 : la ligne "Présence Vacataire" contient À LA FOIS le
        # titre (col A) ET les en-têtes de colonnes (Date/Vacataire/Heure...)
        # sur la MÊME ligne — il n'y a donc qu'UNE seule ligne d'en-tête, pas
        # deux. L'ancien "+2" sautait par erreur la toute première ligne de
        # données du tableau à chaque fichier (ex: "6 mai, Vacataire 1" était
        # invisible pour le moteur). Corrigé en "+1".
        for row in rows[presence_vac_start + 1:]:
            if not any(c for c in row):
                continue
            # Chercher les colonnes avec date + vacataire
            date_val = row[1] if len(row) > 1 else None
            vac_val  = row[2] if len(row) > 2 else None
            hd_val   = row[3] if len(row) > 3 else None
            hf_val   = row[4] if len(row) > 4 else None

            if not date_val or not vac_val:
                continue

            # Parser la date
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.date().strftime('%Y-%m-%d')
            elif isinstance(date_val, datetime.date):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                d = _parse_fr_date(str(date_val))
                if d: date_str = d.strftime('%Y-%m-%d')
                else: continue

            # Normaliser le nom vacataire
            # Corrigé 09/2026 : l'ancienne version ('Vacataire 2' si '2' dans le nom,
            # sinon 'Vacataire 1' par défaut) écrasait silencieusement tout vacataire
            # au-delà de 2 (ex: 'Vacataire 3' devenait 'Vacataire 1', fusionnant à
            # tort ses données de présence avec celles du vrai Vacataire 1). On
            # extrait maintenant le numéro réel, quel qu'il soit.
            vac_name = str(vac_val).strip()
            vac_up = vac_name.upper()
            if 'VACATAIRE' in vac_up or 'VACATA' in vac_up:
                m = re.search(r'\d+', vac_name)
                vac_name = f'Vacataire {m.group()}' if m else 'Vacataire 1'

            # Heures (format "10h", "13h30", "10:00", ou datetime.time)
            cs = _parse_fr_time(hd_val) if hd_val else 600   # défaut 10h
            ce = _parse_fr_time(hf_val) if hf_val else 1140  # défaut 19h

            presences_vac.setdefault(date_str, {})[vac_name] = (cs, ce)

    return {
        'mois':          str(params.get('Mois', '')).strip(),
        'annee':         int(params.get('Année', 2026)),
        'creneaux':      creneaux_all,
        'creneaux_mjv':  creneaux_mjv,   # Mar/Jeu/Ven
        'creneaux_ms':   creneaux_ms,    # Mer/Sam
        'samedis':       samedis,
        'semaines':      semaines,
        'mode_vac':      mode_vac,       # fallback
        'presences_vac': presences_vac,  # {date_str: {agent: (cs, ce)}}
    }


def parse_affectations(raw):
    """
    Retourne :
    - affectations  : {agent: [section1, section2, ...]}  (ordre = priorité)
    - categories    : {agent: 'A' | None}  (A = sections 1 et 2 équivalentes)
    - responsables  : set d'agents responsables
    - pause_flexible: set d'agents avec pause flexible
    - priorite_rdc  : {agent: int} — départage RDC quand plusieurs agents sont à
      égalité de rang de section (colonne "Priorité_remplacement_RDC", 08/2026).
      Nombre plus petit = préféré. Lu par en-tête (pas par position de colonne)
      pour rester robuste si d'autres colonnes sont ajoutées/déplacées.
    """
    ws = raw['Affectations']
    affectations = {}
    categories   = {}
    responsables = set()
    pause_flex   = set()
    priorite_rdc = {}

    header_skipped = False
    col_priorite = None
    for row in ws.iter_rows(values_only=True):
        if not row[0]:
            continue
        agent = str(row[0]).strip()
        if agent == 'Agent':
            header_skipped = True
            for i, h in enumerate(row):
                if h and 'priorité_remplacement_rdc' in str(h).strip().lower().replace(' ', '_'):
                    col_priorite = i
            continue
        if not header_skipped:
            continue

        sections = [str(row[i]).strip() for i in range(2, 6) if row[i] and str(row[i]).strip()]
        affectations[agent]  = sections
        categories[agent]    = str(row[1]).strip() if row[1] else None
        if row[6] and str(row[6]).strip() == 'OUI':
            responsables.add(agent)
        if row[7] and str(row[7]).strip() == 'OUI':
            pause_flex.add(agent)
        if col_priorite is not None and col_priorite < len(row) and row[col_priorite] not in (None, ''):
            try:
                priorite_rdc[agent] = int(row[col_priorite])
            except (ValueError, TypeError):
                pass

    return affectations, categories, responsables, pause_flex, priorite_rdc


def parse_horaires_agents(raw):
    """
    Retourne {agent: {jour: (debut_matin, fin_matin, debut_apm, fin_apm)}}
    Toutes les valeurs en minutes depuis minuit (ou None si absent).
    """
    ws = raw['Horaires_Des_Agents']
    horaires = defaultdict(dict)
    header_skipped = False

    for row in ws.iter_rows(values_only=True):
        if not row[0]:
            continue
        if str(row[0]).strip() == 'Agent':
            header_skipped = True
            continue
        if not header_skipped:
            continue

        agent = str(row[0]).strip()
        jour  = str(row[1]).strip() if row[1] else None
        if not jour:
            continue

        dm = hhmm_to_min(row[2])
        fm = hhmm_to_min(row[3])
        da = hhmm_to_min(row[4])
        fa = hhmm_to_min(row[5])

        horaires[agent][jour] = (dm, fm, da, fa)

    return dict(horaires)


def _normalise_nom(s):
    """Normalise un nom pour comparaison (minuscules, sans accents ni espaces superflus)."""
    if not s:
        return ''
    s = str(s).strip()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    return s.lower()


# La directrice ne doit jamais apparaître comme agent planifiable, y compris
# quand on lit ses horaires dans la grille collaborative.
_AGENTS_EXCLUS_HORAIRES = {'eloise'}

# Nom de l'onglet "grille" (fiches par agent, 4 blocs de service côte à côte :
# ADULTES, JEUNESSE, MUSIQUE, DIRECTION/ADMINISTRATIF), tel que maintenu par
# l'équipe dans le document collaboratif "horaires d'équipes".
ONGLET_HORAIRES_GRILLE = "horaires d'équipes"

# Repérage des 4 blocs de service dans la grille : [colonne_jour, début_matin,
# fin_matin, début_après-midi, fin_après-midi]. Chaque bloc se répète tous les
# 8 lignes (5 lignes de jours + en-tête + 2 lignes d'espacement), pour 5
# emplacements d'agent possibles par bloc.
_HORAIRES_GRILLE_BLOCS = {
    'ADULTES': ['A', 'B', 'C', 'D', 'E'],
    'JEUNESSE': ['H', 'I', 'J', 'K', 'L'],
    'MUSIQUE': ['O', 'P', 'Q', 'R', 'S'],
    'DIRECTION/ADMIN': ['V', 'W', 'X', 'Y', 'Z'],
}
_HORAIRES_GRILLE_GROUPES = [(8, 14), (16, 22), (24, 30), (32, 38), (40, 46)]


def parse_horaires_agents_grille(raw):
    """
    Lit l'onglet "horaires d'équipes" (grille de fiches par agent) et retourne
    exactement le même format que parse_horaires_agents :
        {agent: {jour: (debut_matin, fin_matin, debut_apm, fin_apm)}}
    Toutes les valeurs en minutes depuis minuit (ou None si absent).

    Remplace l'ancienne liste à plat "Horaires_Des_Agents" : l'onglet visuel
    collaboratif est désormais lu directement, sans retranscription manuelle.
    """
    if ONGLET_HORAIRES_GRILLE not in raw:
        return {}
    ws = raw[ONGLET_HORAIRES_GRILLE]

    horaires = defaultdict(dict)

    for cols in _HORAIRES_GRILLE_BLOCS.values():
        day_col, m1, m2, a1, a2 = cols
        for (start, end) in _HORAIRES_GRILLE_GROUPES:
            name = ws[f'{m1}{start}'].value
            if name is None or isinstance(name, (int, float)):
                continue  # case vide, ou table "Fermeture 19h" (pas un agent)
            if isinstance(name, str) and name.strip() in ('', '\xa0'):
                continue
            agent = str(name).strip()
            if _normalise_nom(agent) in _AGENTS_EXCLUS_HORAIRES:
                continue  # la directrice n'est jamais un agent planifiable

            for r in range(start + 1, start + 6):  # Mardi -> Samedi
                jour = ws[f'{day_col}{r}'].value
                if not jour:
                    continue
                jour = str(jour).strip()

                dm = hhmm_to_min(ws[f'{m1}{r}'].value)
                fm = hhmm_to_min(ws[f'{m2}{r}'].value)
                da = hhmm_to_min(ws[f'{a1}{r}'].value)
                fa = hhmm_to_min(ws[f'{a2}{r}'].value)

                if dm is None and fm is None and da is None and fa is None:
                    continue  # agent absent ce jour-là (case entièrement vide)

                horaires[agent][jour] = (dm, fm, da, fa)

    return dict(horaires)


def parse_roulement_samedi(raw):
    """
    Retourne :
    - roulement_type : {agent: 'ROUGE' | 'BLEU'}
    - exceptions     : {semaine_num: {agent: 'ROUGE' | 'BLEU'}}
    """
    ws = raw['Roulement_Samedi']
    roulement_type = {}
    exceptions     = defaultdict(dict)

    mode = None   # 'type' ou 'exceptions'
    current_sem = None

    for row in ws.iter_rows(values_only=True):
        if not any(c for c in row):
            continue
        c0 = str(row[0] or '').strip()
        c1 = str(row[1] or '').strip()
        c2 = str(row[2] or '').strip()

        if c0 == 'Roulement type':
            mode = 'type'
            continue
        if c0 == 'Exceptions par semaine':
            mode = 'exceptions'
            current_sem = None
            continue

        if mode == 'type':
            if c1 and c1 != 'Agent' and c2 and c2 != 'Roulement':
                roulement_type[c1] = c2.upper()

        elif mode == 'exceptions':
            if c1 == 'Semaine':
                continue
            # Numéro de semaine — corrigé 09/2026 : la cellule peut contenir
            # "semaine_1", "Semaine 1", "S1"... et pas seulement un nombre pur
            # ("1"). L'ancien `.isdigit()` ratait tout format textuel et
            # laissait `current_sem` à None, ce qui faisait disparaître
            # silencieusement TOUTE exception de roulement (ex: Stéphane
            # Bleu en semaine 1 de septembre 2026, jamais appliqué).
            if row[1] is not None:
                m = re.search(r'\d+', str(row[1]))
                if m:
                    current_sem = int(m.group())
            # Agent + roulement
            agent = str(row[2] or '').strip().rstrip()
            roul  = str(row[3] or '').strip().upper() if len(row) > 3 else ''
            if agent and roul and current_sem is not None:
                exceptions[current_sem][agent] = roul

    return roulement_type, dict(exceptions)


def parse_besoins_jeunesse(raw):
    """
    Retourne {periode: {jour_cle: {creneau_str: nb_agents}}}
    periode = 'Hors Vacances scolaires' | 'Vacances Scolaires'
    jour_cle = 'Mardi' | 'Mercredi' | 'Jeudi' | 'Vendredi' | 'Samedi_rouge' | 'samedi bleu'
    """
    ws = raw['Besoins_Jeunesse']
    result = {}
    current_periode = None
    headers = []

    for row in ws.iter_rows(values_only=True):
        if not any(c for c in row):
            continue
        c0 = str(row[0] or '').strip()
        c1 = str(row[1] or '').strip()

        # Ligne de période
        if 'Vacances' in c0 or 'Hors' in c0:
            current_periode = c0.rstrip()
            result[current_periode] = {}
            headers = []
            continue

        # Ligne d'en-têtes
        if c1 == 'Créneau' or c1 == 'créneau':
            headers = [str(row[i] or '').strip() for i in range(2, 8)]
            for h in headers:
                if h:
                    result[current_periode][h] = {}
            continue

        # Ligne de données
        if current_periode and headers and c1 and ':' in c1:
            cren = c1.strip()
            for i, h in enumerate(headers):
                if h and row[i + 2] is not None:
                    try:
                        result[current_periode][h][cren] = int(row[i + 2])
                    except (ValueError, TypeError):
                        pass

    return result


def _parse_fr_date(s, annee_defaut=None):
    """Parse "mardi 5 mai 2026" → datetime.date ou None.
    Si le texte ne contient pas d'année (ex: "mardi 1 septembre"), utilise
    `annee_defaut` si fourni (corrigé 09/2026 : le fichier septembre n'écrit
    jamais l'année dans l'onglet Événements, contrairement à celui de mai)."""
    MOIS = {'janvier':1,'février':2,'mars':3,'avril':4,'mai':5,'juin':6,
             'juillet':7,'août':8,'septembre':9,'octobre':10,'novembre':11,'décembre':12}
    s = str(s).strip().lower()
    parts = s.split()
    for i, p in enumerate(parts):
        if p.isdigit():
            day = int(p)
            mois = MOIS.get(parts[i+1] if i+1<len(parts) else '')
            yr   = int(parts[i+2]) if i+2<len(parts) and parts[i+2].isdigit() else None
            if yr is None:
                yr = annee_defaut
            if mois and yr:
                try:
                    return datetime.date(yr, mois, day)
                except ValueError:
                    pass
    return None

def _parse_fr_time(s):
    """Parse "10h", "10h30", "13h30", "10:00" → minutes ou None."""
    if s is None: return None
    if isinstance(s, datetime.time):
        return s.hour*60 + s.minute
    s = str(s).strip().lower()
    if not s or s in ('non défini',''):
        return None
    # "13h30" ou "10h"
    m = re.match(r'(\d{1,2})h(\d{0,2})', s)
    if m:
        return int(m.group(1))*60 + (int(m.group(2)) if m.group(2) else 0)
    # "10:00"
    if ':' in s:
        p = s.split(':')
        try: return int(p[0])*60+int(p[1])
        except: pass
    return None

def parse_evenements(raw, annee_defaut=None):
    """
    Retourne liste de {date_str, cs, ce, nom, agents: []}
    Gère les dates en format texte français ("mardi 5 mai 2026" OU, sans
    année, "mardi 1 septembre" — corrigé 09/2026, cf. _parse_fr_date)
    et les heures "10h", "13h30".
    """
    ws = raw.get('Événements') or raw.get('Evenements')
    if ws is None:
        return []
    events = []
    header_skipped = False

    for row in ws.iter_rows(values_only=True):
        if not row[0]:
            continue
        c0 = str(row[0]).strip()
        if c0.lower() in ('date',):
            header_skipped = True
            continue
        if not header_skipped:
            continue

        # Date
        date_val = row[0]
        if isinstance(date_val, (datetime.datetime, datetime.date)):
            d = date_val.date() if isinstance(date_val, datetime.datetime) else date_val
        else:
            d = _parse_fr_date(c0, annee_defaut=annee_defaut)
        if d is None:
            continue
        date_str = d.strftime('%Y-%m-%d')

        cs  = _parse_fr_time(row[1])
        ce  = _parse_fr_time(row[2])
        nom = str(row[3] or '').strip()
        agents_str = str(row[4] or '').strip()
        agents = [a.strip() for a in re.split(r'[,;/]', agents_str)
                  if a.strip() and a.strip() not in ('', 'None')] if agents_str else []

        if cs is not None and ce is not None and nom:
            events.append({'date': date_str, 'cs': cs, 'ce': ce,
                           'nom': nom, 'agents': agents})

    return events


def parse_planning_type(raw):
    """
    Nouveau format (2026) :
    Col A = créneau ('10H-12H'), Col B = RDC, C = Adulte, D = M&F, E = Jeunesse
    Jours en col A avec espaces : '  MARDI', '  SAMEDI — SEMAINE ROUGE'
    Plusieurs agents séparés par ' / ' dans la même cellule.
    """
    ws = raw.get('Planning_type') or raw.get('planning_type')
    if ws is None:
        return {}

    result    = {}
    cur_jour  = None
    sam_count = 0

    JOUR_MAP = {
        'MARDI': 'Mardi', 'MERCREDI': 'Mercredi',
        'JEUDI': 'Jeudi', 'VENDREDI': 'Vendredi',
    }

    def get_agents(val):
        if not val:
            return []
        s = str(val).strip()
        # Séparer sur / et nettoyer
        agents = []
        for a in re.split(r'\s*/\s*', s):
            a = a.strip()
            if not a or re.match(r'^\d', a):
                continue
            # Normaliser vacataires (même correctif 09/2026 qu'en amont : extraire
            # le vrai numéro plutôt que d'écraser tout n°3+ en 'Vacataire 1')
            a_up = a.upper()
            if 'VACATAIRE' in a_up:
                m = re.search(r'\d+', a)
                a = f'Vacataire {m.group()}' if m else 'Vacataire 1'
            # Nettoyer mentions "à partir de..."
            a = re.split(r'\s+[àa]\s+', a)[0].strip()
            if a:
                agents.append(a)
        return agents

    for row in ws.iter_rows(values_only=True):
        if not any(c for c in row):
            continue
        c0 = str(row[0] or '').strip()

        # Détecter le jour
        c0_up = c0.upper()
        detected = False
        for key, val in JOUR_MAP.items():
            if key in c0_up:
                cur_jour = val
                result.setdefault(cur_jour, {})
                detected = True
                break

        if not detected:
            if 'SAMEDI' in c0_up:
                sam_count += 1
                cur_jour = 'Samedi_ROUGE' if sam_count == 1 else 'Samedi_BLEU'
                result.setdefault(cur_jour, {})
                continue

        if detected or cur_jour is None:
            continue

        # Ignorer lignes d'en-tête et lignes fermées
        if c0 in ('Créneau', 'créneau', '') or 'fermé' in c0.lower():
            continue
        if c0.startswith('ROULEMENT') or c0.startswith('HEURES') or c0 == 'Agent':
            cur_jour = None
            continue
        if 'Section' in c0:
            cur_jour = None
            continue

        # Ligne de données : col A = créneau horaire
        if not ('-' in c0 and ('H' in c0.upper() or ':' in c0)):
            continue

        cren_str = c0.strip()
        if cren_str not in result[cur_jour]:
            result[cur_jour][cren_str] = {s: [] for s in SECTIONS}

        result[cur_jour][cren_str]['RDC']      = get_agents(row[1] if len(row) > 1 else None)
        result[cur_jour][cren_str]['Adulte']   = get_agents(row[2] if len(row) > 2 else None)
        result[cur_jour][cren_str]['MF']       = get_agents(row[3] if len(row) > 3 else None)
        # Jeunesse : 3 colonnes séparées (E, F, G — une par agent), comme
        # partout ailleurs dans le projet — PAS une seule cellule avec des
        # noms séparés par '/'. Corrigé 08/2026 : l'ancien code ne lisait
        # que la colonne E et perdait silencieusement les agents des
        # colonnes F et G quand le PT en prévoyait 2 ou 3.
        jeunesse_agents = []
        for col_idx in (4, 5, 6):
            jeunesse_agents.extend(get_agents(row[col_idx] if len(row) > col_idx else None))
        result[cur_jour][cren_str]['Jeunesse'] = jeunesse_agents

    return result


# ══════════════════════════════════════════════════════════════
#  CONSTRUCTION DU CALENDRIER DU MOIS
# ══════════════════════════════════════════════════════════════

def parse_jours_speciaux(raw):
    """
    Lit l'onglet Jours_speciaux.
    Retourne {date_str: {'ferie': bool, 'vacances': bool}}
    Les dates sont en format texte français : "vendredi 8 mai 2026"
    """
    ws = raw.get('Jours_speciaux')
    if ws is None:
        return {}
    result = {}
    MOIS_FR = {
        'janvier':1,'février':2,'mars':3,'avril':4,'mai':5,'juin':6,
        'juillet':7,'août':8,'septembre':9,'octobre':10,'novembre':11,'décembre':12
    }
    for row in ws.iter_rows(values_only=True):
        if not row[0] or str(row[0]).strip() in ('Date','date'):
            continue
        date_str_raw = str(row[0]).strip().lower()
        # Parser "vendredi 8 mai 2026" → extraire jour, mois, année
        parts = date_str_raw.split()
        try:
            # Format: "jour_semaine jour_num mois_fr annee"
            # ou "jour_num mois_fr annee"
            if len(parts) >= 3:
                # Trouver le numéro du jour
                for i, p in enumerate(parts):
                    if p.isdigit():
                        day_num = int(p)
                        mois_str = parts[i+1] if i+1 < len(parts) else ''
                        year_str = parts[i+2] if i+2 < len(parts) else ''
                        mois_num = MOIS_FR.get(mois_str)
                        if mois_num and year_str.isdigit():
                            d = datetime.date(int(year_str), mois_num, day_num)
                            date_key = d.strftime('%Y-%m-%d')
                            result[date_key] = {
                                'ferie':    str(row[1] or '').strip().upper() == 'OUI',
                                'vacances': str(row[2] or '').strip().lower() in ('vacances','vac','vacation'),
                            }
                            break
        except (ValueError, IndexError):
            continue
    return result


def build_calendar(mois_str, annee, samedis_params):
    """
    Retourne les semaines SP du mois, chacune Mardi→Samedi.
    Commence au PREMIER MARDI du mois (les jours avant sont exclus).
    samedis_params[N] = type du N-ième Samedi des semaines complètes.
    """
    MOIS_FR_NUM = {
        'janvier':1,'fevrier':2,'février':2,'mars':3,'avril':4,'mai':5,'juin':6,
        'juillet':7,'aout':8,'août':8,'septembre':9,'octobre':10,'novembre':11,'decembre':12,'décembre':12
    }
    mois_clean = str(mois_str).strip().lower().split()[0]
    mois_num   = MOIS_FR_NUM.get(mois_clean, 5)
    premier    = datetime.date(annee, mois_num, 1)

    JOURS_FR = {0:'Lundi',1:'Mardi',2:'Mercredi',3:'Jeudi',
                4:'Vendredi',5:'Samedi',6:'Dimanche'}
    JOURS_SP = ['Mardi','Mercredi','Jeudi','Vendredi','Samedi']

    # Trouver le premier Mardi du mois
    d = premier
    while JOURS_FR[d.weekday()] != 'Mardi':
        d += datetime.timedelta(days=1)

    # Construire les semaines depuis le premier Mardi
    semaines  = []
    current   = []
    week_num  = 1
    sam_count = 0

    # Corrigé 09/2026 (règle précisée par l'utilisatrice) : une semaine
    # ENTAMÉE dans le mois (son Mardi tombe dans le mois demandé) doit être
    # terminée jusqu'au Samedi, même si Jeudi/Vendredi/Samedi débordent sur
    # le mois suivant. Le mois suivant démarrera alors à son propre premier
    # Mardi, après la fin de cette semaine à cheval.
    # Condition de boucle : on continue tant qu'on est dans le mois, OU
    # qu'une semaine est en cours de construction (current non vide) — dans
    # ce 2e cas on ne démarre jamais une NOUVELLE semaine hors du mois,
    # on finit juste celle déjà entamée (elle se termine forcément au
    # prochain Samedi, donc pas de risque de boucle infinie).
    while d.month == mois_num or current:
        jour_fr = JOURS_FR[d.weekday()]
        if jour_fr in JOURS_SP:
            sam_type = None
            if jour_fr == 'Samedi':
                sam_count += 1
                sam_type   = samedis_params.get(sam_count)
            current.append({'date': d.strftime('%Y-%m-%d'),
                             'jour': jour_fr, 'samedi_type': sam_type})
            if jour_fr == 'Samedi':
                semaines.append({'num': week_num, 'jours': current})
                current  = []
                week_num += 1
        d += datetime.timedelta(days=1)

    if current:  # dernière semaine sans samedi
        semaines.append({'num': week_num, 'jours': current})

    # Garder uniquement les semaines avec un samedi défini
    n_sam, result, sam_vus = len(samedis_params), [], 0
    for sem in semaines:
        has_sam = any(j['jour'] == 'Samedi' for j in sem['jours'])
        if has_sam:
            sam_vus += 1
            if sam_vus <= n_sam:
                result.append(sem)
        else:
            result.append(sem)

    for i, sem in enumerate(result, 1):
        sem['num'] = i
    return result


def is_vacataire(agent):
    return 'Vacataire' in agent or 'vacataire' in agent


# Mots-clés identifiant un événement d'ABSENCE (congé, RTT, formation...) dans
# l'onglet Événements — à ne jamais compter comme du "travail équivalent" pour
# l'équité (§ ev_minutes_agent dans solve_day). Mêmes mots-clés que la colonne
# J du générateur Excel (excel_writer.py / generate_planning_excel_septembre.py),
# dupliqués ici pour que le moteur reste autonome (pas de dépendance croisée).
_MOTS_CLES_ABSENCE = ('congé', 'conge', 'rtt', 'vacation', 'absence', 'formation')


def _est_evenement_absence(nom):
    """True si le nom de l'événement correspond à une absence (congé, RTT,
    formation...) et non à un vrai événement travaillé (accueil, animation,
    réunion)."""
    n = (nom or '').lower()
    return any(mc in n for mc in _MOTS_CLES_ABSENCE)


def _norm_jour_cle(s):
    """Normalise une clé de jour ('Samedi_rouge', 'samedi bleu', 'SAMEDI_BLEU'...)
    pour une comparaison insensible à la casse et aux séparateurs (espace/tiret bas)."""
    return str(s).strip().lower().replace(' ', '_').replace('-', '_')


def get_besoins_jour_normalise(besoins_periode, jour_key):
    """Cherche jour_key dans besoins_periode (dict {jour_cle: {...}}) en tolérant
    les variations de casse/séparateur entre les en-têtes de colonnes Excel."""
    cible = _norm_jour_cle(jour_key)
    for k, v in besoins_periode.items():
        if _norm_jour_cle(k) == cible:
            return v
    return {}


def agent_disponible(agent, jour, cs, ce, horaires_agents, evenements,
                     date_str, pause_flex, presences_vac=None):
    """
    Retourne True si l'agent peut être placé sur ce créneau (cs, ce) ce jour-là.
    Vérifie : horaires contractuels, pause contractuelle, événements bloquants.
    """
    if not is_vacataire(agent):
        h = horaires_agents.get(agent, {}).get(jour)
        if not h:
            return False  # pas de contrat ce jour

        dm, fm, da, fa = h

        # Vérifier que le créneau est dans les heures de travail
        dans_matin = (dm is not None and fm is not None and cs >= dm and ce <= fm)
        dans_apm   = (da is not None and fa is not None and cs >= da and ce <= fa)
        # Si pas de pause réelle (fm==da) ou pause flexible : autoriser tout créneau dans [dm, fa]
        dans_global = (dm is not None and fa is not None and cs >= dm and ce <= fa and
                       (fm == da or agent in pause_flex))

        if not (dans_matin or dans_apm or dans_global):
            return False

        # Pause contractuelle (sans pause flexible) : ne pas placer pendant la pause
        if agent not in pause_flex and fm is not None and da is not None and fm < da:
            en_pause = (cs >= fm and ce <= da)
            if en_pause:
                return False
    else:
        # Vacataires : disponibilité définie par le tableau Présence Vacataire
        pv = presences_vac or {}
        # Si présence explicite définie → utiliser ces horaires
        if date_str in pv and agent in pv[date_str]:
            vac_cs, vac_ce = pv[date_str][agent]
            if not (cs >= vac_cs and ce <= vac_ce):
                return False
        else:
            # Fallback : horaires contractuels
            h = horaires_agents.get(agent, {}).get(jour)
            if h:
                dm, fm, da, fa = h
                if dm is not None and fa is not None:
                    if not (cs >= dm and ce <= fa):
                        return False

    # Événements bloquants
    for ev in evenements:
        if ev['date'] != date_str:
            continue
        # Si la liste d'agents est vide → événement général sans impact sur les agents
        # Si la liste est non vide et l'agent n'y figure pas → pas concerné
        if not ev['agents'] or agent not in ev['agents']:
            continue
        # Chevauchement → agent bloqué
        if cs < ev['ce'] and ce > ev['cs']:
            return False

    return True


# ══════════════════════════════════════════════════════════════
#  MOTEUR CP-SAT — UNE JOURNÉE
# ══════════════════════════════════════════════════════════════

def solve_day(jour, date_str, creneaux_ouverts, agents_eligibles,
              affectations, categories, responsables, pause_flex, priorite_rdc,
              horaires_agents, evenements, besoins_jeunesse,
              planning_type_jour, roulement_agents,
              samedi_type=None, periode='Hors Vacances scolaires',
              mode_vac=None, swap_map=None, presences_vac=None,
              cumul_hebdo_avant=None):
    """
    swap_map : {agent_absent: agent_remplacant} pour ce jour
               ex: {'Guillaume': 'Robin'} si Guillaume est BLEU ce samedi ROUGE
               et Robin a pris sa place
    cumul_hebdo_avant : {agent: minutes de dépassement déjà cumulées cette
               semaine, avant ce jour} — fourni par compute_full_planning,
               remis à zéro à chaque nouvelle semaine. Sert à l'équité
               hebdomadaire (§7bis, 08/2026) : évite qu'un même agent soit
               systématiquement choisi comme remplaçant tous les jours de la
               semaine, en tenant compte de ce qu'il a déjà fait en plus les
               jours précédents.
    """
    if swap_map is None:
        swap_map = {}
    if cumul_hebdo_avant is None:
        cumul_hebdo_avant = {}
    """
    Résout le planning d'une journée avec CP-SAT.

    Paramètres :
    - creneaux_ouverts : liste de (cs, ce)
    - agents_eligibles : liste d'agents pouvant travailler ce jour
    - planning_type_jour : {creneau_str: {section: [agents]}}
    - roulement_agents : {agent: 'ROUGE'|'BLEU'} pour ce jour (samedi)

    Retourne : {creneau_idx: {section: [agents]}} ou None si infaisable
    """
    if mode_vac is None:
        mode_vac = JOURS_VAC

    model  = cp_model.CpModel()
    agents = list(agents_eligibles)
    n_cren = len(creneaux_ouverts)

    # ── Variables de décision ──────────────────────────────────
    # x[a][c][s] = 1 si agent a travaille au créneau c en section s
    x = {}
    for a in agents:
        for c in range(n_cren):
            for s in SECTIONS:
                x[a, c, s] = model.new_bool_var(f'x_{a}_{c}_{s}')

    # ══ CONTRAINTES DURES ════════════════════════════════════

    # A2 : vacataires jamais en RDC
    for a in agents:
        if is_vacataire(a):
            for c in range(n_cren):
                model.add(x[a, c, 'RDC'] == 0)

    # A3 : Stéphane uniquement MF
    if 'Stéphane' in agents:
        for c in range(n_cren):
            for s in SECTIONS:
                if s != 'MF':
                    model.add(x['Stéphane', c, s] == 0)

    # A1 : sections habilitées uniquement
    for a in agents:
        sects_ok = set(affectations.get(a, []))
        for c in range(n_cren):
            for s in SECTIONS:
                if s not in sects_ok:
                    model.add(x[a, c, s] == 0)

    # A4 : max 1 agent par section/créneau pour RDC, Adulte, MF
    for c in range(n_cren):
        for s in ['RDC', 'Adulte', 'MF']:
            model.add_at_most_one(x[a, c, s] for a in agents)

    # D13 : 1 agent = 1 section par créneau
    for a in agents:
        for c in range(n_cren):
            model.add_at_most_one(x[a, c, s] for s in SECTIONS)

    # B1/B2 : disponibilité contractuelle
    for a in agents:
        cs_ce_list = creneaux_ouverts
        for c, (cs, ce) in enumerate(cs_ce_list):
            if not agent_disponible(a, jour, cs, ce, horaires_agents,
                                    evenements, date_str, pause_flex,
                                    presences_vac=presences_vac):
                for s in SECTIONS:
                    model.add(x[a, c, s] == 0)

    # B3 : vacataires uniquement les jours autorisés
    # Si presences_vac est défini → un vacataire est autorisé ce jour s'il est listé
    # Si presences_vac vide → fallback sur mode_vac
    _pv = presences_vac or {}
    for a in agents:
        if not is_vacataire(a):
            continue
        # Autorisé si présence explicite définie pour cette date
        if date_str in _pv and a in _pv[date_str]:
            continue
        # Autorisé si mode_vac global inclut ce jour
        if jour in (mode_vac or set()):
            continue
        # Sinon → bloqué
        for c in range(n_cren):
            for s in SECTIONS:
                model.add(x[a, c, s] == 0)

    # D1 : roulement samedi ROUGE/BLEU
    if jour == 'Samedi' and samedi_type:
        for a in agents:
            if is_vacataire(a):
                continue
            roul_agent = roulement_agents.get(a)
            if roul_agent and roul_agent != samedi_type:
                for c in range(n_cren):
                    for s in SECTIONS:
                        model.add(x[a, c, s] == 0)


    # C3 : pause déjeuner ≥ 1h (12h-14h)
    # Réguliers sauf Delphine + vacataires le samedi et mercredi
    # ⚠️ Corrigé 08/2026 : l'ancienne version comparait une somme sur (créneau × section)
    # à son propre nombre total de variables — toujours vraie car un agent ne peut être
    # que dans 1 section à la fois (D13), donc la pause n'était jamais réellement imposée.
    # Nouvelle version : calcule la durée réellement travaillée (en minutes) dans la
    # fenêtre 12h-14h et exige qu'il reste au moins 60 minutes libres.
    # ⚠️ Corrigé une 2e fois 08/2026 (remarque utilisatrice sur Barbara, mercredi) :
    # un agent dont les horaires du jour (Horaires_Des_Agents) montrent une présence
    # RÉELLEMENT continue (pas de coupure entre le matin et l'après-midi, fm==da,
    # même logique que la pause flexible ci-dessus dans agent_disponible) n'a PAS à
    # se voir imposer une pause artificielle — il n'y en a pas dans son contrat.
    def a_pause_naturelle(agent):
        h = horaires_agents.get(agent, {}).get(jour)
        if not h:
            return False
        _, fm, da, _ = h
        return fm is not None and da is not None and fm != da
    agents_pause_oblig = [a for a in agents
                          if ((not is_vacataire(a) and a != 'Delphine'
                               and a not in pause_flex and a_pause_naturelle(a)) or
                             (is_vacataire(a) and jour in ('Samedi', 'Mercredi')))]
    pause_creneaux = [c for c, (cs, ce) in enumerate(creneaux_ouverts)
                      if cs >= 720 and ce <= 840]  # 12h-14h = 720-840 min
    for a in agents_pause_oblig:
        if pause_creneaux:
            pause_total_dur = sum(creneaux_ouverts[c][1] - creneaux_ouverts[c][0]
                                  for c in pause_creneaux)
            if pause_total_dur >= 60:
                duree_travaillee = sum(
                    (creneaux_ouverts[c][1] - creneaux_ouverts[c][0]) * x[a, c, s]
                    for c in pause_creneaux for s in SECTIONS
                )
                # Au moins 60 minutes NON travaillées dans la fenêtre 12h-14h
                model.add(duree_travaillee <= pause_total_dur - 60)

    # ══ Pénalités et alertes (souple) — déclarées ici pour être utilisées
    #    par F1/F2 (Jeunesse) et D_FILL (RDC/Adulte/MF) ci-dessous
    penalites = []
    penalites_stabilite = []  # objectif de 2e RANG (passe 2/3) — ne jamais déplacer un agent
                               # présent et normalement à sa place, sauf nécessité réelle.
                               # Voir §résolution en 3 passes plus bas.
    penalites_qualite = []  # objectif de 3e RANG (passe 3/4) — QUI remplace le mieux
                             # (G2/J1/J3/I1/préférence vacataires) — indépendant des heures
    penalites_equite = []  # objectif de 4e RANG (passe 4/4) — équité des heures SEULEMENT
                            # (dépassement ET manque par rapport au PT). Résolue en DERNIER
                            # pour ne jamais pouvoir dégrader le choix qualitatif du meilleur
                            # remplaçant (§13.5 — corrige la régression Stéphanie/Macha du
                            # 11/08 : avant, l'équité en minutes pouvait "acheter" une moins
                            # bonne section ou une responsable pour économiser quelques
                            # minutes de dépassement ailleurs)
    alertes = []  # [(cren_idx, section, message)] — créneaux non pourvus malgré le besoin
    ALERTE_POIDS = 5000  # pénalité très forte : remplir RDC/Adulte/MF (D_FILL) — priorité maximale
    JEUNESSE_POIDS = 200  # pénalité plus faible : atteindre le nombre visé en Jeunesse cède
                           # le pas si ça entre en conflit avec le remplissage RDC/Adulte/MF
    jeunesse_requis = {}   # {cren_idx: besoin} — mémorisé pour vérif post-résolution
    fill_requis = {}       # {(cren_idx, section): True} — sections PT à vérifier post-résolution

    # F1/F2 : besoins Jeunesse
    # Hors Vacances Scolaires → le PT définit le nombre EXACT d'agents en Jeunesse
    # Vacances Scolaires (y compris jour spécial ponctuel via Jours_speciaux) →
    #   Besoins_Jeunesse donne le nombre par créneau
    est_vacances = 'Hors' not in periode

    if est_vacances:
        # Vacances scolaires : utiliser Besoins_Jeunesse
        jour_key = jour
        if jour == 'Samedi' and samedi_type:
            # Recherche insensible à la casse/espaces/tirets : le fichier Excel peut
            # écrire "Samedi_rouge", "samedi bleu", "Samedi_Bleu", etc.
            def _norm(s):
                return s.lower().replace('_', ' ').replace('-', ' ').strip()
            cible = _norm(f'samedi {samedi_type}')
            jour_key = next((k for k in besoins_jeunesse.get(
                next((p for p in besoins_jeunesse if 'Hors' not in p), ''), {})
                if _norm(k) == cible), f'Samedi_{samedi_type.lower()}')
        periode_key = next((k for k in besoins_jeunesse if 'Hors' not in k), None)
        if periode_key:
            besoins_jour = besoins_jeunesse.get(periode_key, {}).get(jour_key, {})
            # Pré-parser les tranches du tableau Besoins_Jeunesse en (cs, ce, besoin)
            besoins_ranges = []
            for cren_str, besoin in besoins_jour.items():
                bparsed = parse_creneau(cren_str)
                if bparsed:
                    besoins_ranges.append((bparsed[0], bparsed[1], besoin))
            for c, (cs, ce) in enumerate(creneaux_ouverts):
                # Le créneau du planning peut être plus large que les tranches du
                # tableau Besoins_Jeunesse (ex: créneau 17:00-19:00 vs tranches
                # 17:00-18:00 et 18:00-19:00). On prend le MINIMUM des tranches
                # contenues dans le créneau (évite de sur-exiger sur tout le bloc
                # à cause d'un pic ponctuel sur une sous-tranche).
                sous_tranches = [b for (bcs, bce, b) in besoins_ranges
                                  if bcs >= cs and bce <= ce]
                if sous_tranches:
                    besoin = min(sous_tranches)
                else:
                    # Correspondance exacte en repli si aucune sous-tranche trouvée
                    cren_str = f'{cs//60:02d}:{cs%60:02d}-{ce//60:02d}:{ce%60:02d}'
                    besoin = besoins_jour.get(cren_str, 0)
                jeunesse_vars = [x[a, c, 'Jeunesse'] for a in agents]
                sum_j = sum(jeunesse_vars)
                jeunesse_requis[c] = besoin
                if besoin > 0:
                    shortfall = model.new_int_var(0, besoin, f'shortfall_jeu_{c}')
                    model.add(shortfall >= besoin - sum_j)
                    model.add(sum_j <= besoin)
                    penalites.append(JEUNESSE_POIDS * shortfall)
                else:
                    model.add(sum_j == 0)
    else:
        # Hors vacances : nombre exact d'agents Jeunesse = ce que dit le PT
        for c, (cs, ce) in enumerate(creneaux_ouverts):
            # Compter le nombre d'agents Jeunesse prévus dans le PT pour ce créneau
            nb_pt_jeunesse = 0
            for cren_str, sect_agents in planning_type_jour.items():
                pt_parsed = parse_creneau(cren_str)
                if not pt_parsed:
                    continue
                pt_cs, pt_ce = pt_parsed
                if cs >= pt_cs and ce <= pt_ce:
                    nb_pt_jeunesse = len([a for a in sect_agents.get('Jeunesse', [])
                                          if a and a.strip()])
            # Plafonner au nombre d'agents Jeunesse réellement disponibles à ce créneau
            # (évite l'infaisabilité quand les agents PT sont absents)
            jeunesse_dispo = [a for a in agents
                              if 'Jeunesse' in affectations.get(a, [])
                              and agent_disponible(a, jour, cs, ce, horaires_agents,
                                                   evenements, date_str, pause_flex,
                                                   presences_vac=presences_vac)]
            nb_possible = len(jeunesse_dispo)
            nb_requis = min(nb_pt_jeunesse, nb_possible)
            jeunesse_vars = [x[a, c, 'Jeunesse'] for a in agents]
            sum_j = sum(jeunesse_vars)
            jeunesse_requis[c] = nb_requis
            if nb_requis > 0:
                shortfall = model.new_int_var(0, nb_requis, f'shortfall_jeu_{c}')
                model.add(shortfall >= nb_requis - sum_j)
                model.add(sum_j <= nb_requis)
                penalites.append(JEUNESSE_POIDS * shortfall)
            else:
                model.add(sum_j == 0)

    # K3 (dure) : vacataire seul en Jeunesse uniquement 12h-14h
    for c, (cs, ce) in enumerate(creneaux_ouverts):
        is_in_12_14 = (cs >= 720 and ce <= 840)
        if not is_in_12_14:
            for a_vac in [a for a in agents if is_vacataire(a)]:
                # Si vacataire en Jeunesse → au moins 1 régulier aussi en Jeunesse
                reguliers_j = [x[a, c, 'Jeunesse'] for a in agents if not is_vacataire(a)]
                model.add(x[a_vac, c, 'Jeunesse'] <= sum(reguliers_j))

    # ══ PRÉ-CALCUL PT INDEXÉ (partagé dures + molles) ══════════
    # Convertir planning_type_jour en {cren_idx: {section: [agents]}}
    pt_indexed = {}
    for cren_str, sections_agents in planning_type_jour.items():
        parsed = parse_creneau(cren_str)
        if not parsed:
            continue
        for c, (cs, ce) in enumerate(creneaux_ouverts):
            if cs >= parsed[0] and ce <= parsed[1]:
                pt_indexed.setdefault(c, {s: [] for s in SECTIONS})
                for s in SECTIONS:
                    pt_indexed[c][s] = sections_agents.get(s, [])

    # ══ D_FILL (SOUPLE) : si le PT prévoit quelqu'un dans une section,
    #    le solveur DOIT très fortement y mettre exactement 1 agent (remplaçant si absent).
    #    Rendue souple (au lieu de dure) : si c'est structurellement impossible ce créneau-là
    #    (aucun agent dispo et habilité), le créneau reste vide avec une ALERTE plutôt que
    #    de rendre toute la journée infaisable.
    for c, sections_dict in pt_indexed.items():
        for s in ['RDC', 'Adulte', 'MF']:
            pt_agents_ici = [a for a in sections_dict.get(s, []) if a and a.strip()]
            if not pt_agents_ici:
                continue  # PT ne prévoit personne ici → pas de contrainte
            agents_possibles = [a for a in agents
                                 if s in affectations.get(a, [])
                                 and not (is_vacataire(a) and s == 'RDC')]
            if not agents_possibles:
                alertes.append((c, s, 'aucun agent habilité disponible'))
                continue  # Aucun agent possible → alerte, on laisse vide
            sum_x = sum(x[a, c, s] for a in agents)
            fill_requis[(c, s)] = True
            # Le "au plus 1" est déjà garanti par A4 (add_at_most_one) plus haut.
            # Ici on pousse fortement vers "exactement 1" sans jamais bloquer le solveur.
            penalites.append(ALERTE_POIDS * (1 - sum_x))

    # Verrou manquant corrigé (08/2026, détecté en validant l'équité hebdo §7bis) :
    # sur RDC/Adulte/MF, rien n'empêchait auparavant d'affecter un agent à un
    # créneau où le PT ne prévoit PERSONNE (contrairement à Jeunesse, qui avait
    # déjà `sum_j == 0` dans ce cas, cf. plus haut). Ce trou ne coûtait rien au
    # calcul : le solveur pouvait laisser le créneau vide OU le remplir sans
    # aucune différence de score, au bon vouloir de son ordre de recherche
    # interne — un simple changement du modèle (comme l'ajout des variables
    # d'équité hebdo) pouvait donc faire basculer ce choix arbitraire. Corrigé
    # en forçant 0 agent, symétriquement à ce qui existe déjà pour Jeunesse.
    for c in range(n_cren):
        for s in ['RDC', 'Adulte', 'MF']:
            if (c, s) not in fill_requis:
                model.add(sum(x[a, c, s] for a in agents) == 0)

    # C1/C2 : durées consécutives — DEUX seuils désormais distincts (corrigé 08/2026,
    # suite à la proposition de l'utilisatrice sur vendredi 4/09 : Marie-France
    # aurait dû rester au RDC tout l'après-midi plutôt que de fragmenter entre
    # Robin/Stéphanie).
    # - Seuil IDÉAL (2h30 en semaine, 4h mercredi/samedi) : dépassement TOLÉRÉ,
    #   pénalité déplacée en passe 3/4 (qualité) — comparé équitablement à G2/J1/J3
    #   au lieu d'écraser automatiquement toute autre considération comme avant.
    # - Seuil TOLÉRÉ (4h partout, 5h Barbara le samedi) : plafond DUR, jamais
    #   dépassé, quelle que soit la pression ailleurs dans l'optimisation.
    ideal_consec_defaut = 4 * 60 if jour in ('Mercredi', 'Samedi') else 2 * 60 + 30

    def ideal_consec_pour(agent):
        if agent == 'Barbara' and jour == 'Samedi':
            return 5 * 60  # exception validée — pas de préférence de raccourcissement ici
        return ideal_consec_defaut

    def tolere_consec_pour(agent):
        if agent == 'Barbara' and jour == 'Samedi':
            return 5 * 60  # exception validée
        return 4 * 60  # plafond dur commun à tous les jours désormais

    for a in agents:
        ideal_c  = ideal_consec_pour(a)
        tolere_c = tolere_consec_pour(a)
        for c_start in range(n_cren):
            total_dur = 0
            c_end = c_start
            while c_end < n_cren:
                cs_e, ce_e = creneaux_ouverts[c_end]
                if c_end > c_start:
                    cs_prev, ce_prev = creneaux_ouverts[c_end - 1]
                    if cs_e != ce_prev:
                        break
                total_dur += ce_e - cs_e
                consec_vars = [x[a, c, s] for c in range(c_start, c_end + 1)
                               for s in SECTIONS]
                limit = c_end - c_start
                if total_dur > tolere_c:
                    # Plafond quasi-dur : très fortement découragé (poids proche de
                    # D_FILL/Jeunesse, tier 1) mais pas absolument bloquant — sinon
                    # une vraie impossibilité de couverture (ex: samedi 5/09, Jeunesse)
                    # devient une alerte plutôt qu'un dépassement exceptionnel de 4h,
                    # ce qui va à l'encontre du principe "une alerte est pire qu'un
                    # léger débordement". Revient à l'esprit d'origine, mais isolé du
                    # seuil idéal (voir ci-dessous) pour ne plus l'écraser inutilement.
                    viol = model.new_bool_var(f'consec_tolere_viol_{a}_{c_start}_{c_end}')
                    model.add(sum(consec_vars) <= limit).only_enforce_if(viol.negated())
                    model.add(sum(consec_vars) <= limit + len(SECTIONS)).only_enforce_if(viol)
                    penalites.append(150 * viol)
                elif total_dur > ideal_c:
                    # Préférence SOUPLE : comparée en passe 3/4 (qualité), plus en
                    # passe 1 — ne domine plus automatiquement G2/J1/J3.
                    viol = model.new_bool_var(f'consec_viol_{a}_{c_start}_{c_end}')
                    model.add(sum(consec_vars) <= limit).only_enforce_if(viol.negated())
                    model.add(sum(consec_vars) <= limit + len(SECTIONS)).only_enforce_if(viol)
                    penalites_qualite.append(40 * viol)
                c_end += 1


    # PLAFOND QUOTIDIEN (nouveau, sécurité anti-surcharge) : un régulier ne peut
    # jamais dépasser un total d'heures données dans la même journée, même si le
    # rattrapage hebdomadaire (plus bas) pousse dans cette direction. Ne s'applique
    # PAS aux vacataires (leurs présences longues type 10h-19h sont normales et
    # définies par le tableau Présence Vacataire, pas par ce plafond).
    PLAFOND_JOUR_MINUTES = 420  # 7h — à ajuster si besoin
    for a in agents:
        if is_vacataire(a):
            continue
        total_jour = sum((creneaux_ouverts[c][1] - creneaux_ouverts[c][0]) * x[a, c, s]
                          for c in range(n_cren) for s in SECTIONS)
        model.add(total_jour <= PLAFOND_JOUR_MINUTES)

    # G1 : préférer l'agent du PT dans sa section
    # Sur les jours où un/des vacataire(s) sont présents (peu importe le jour de la semaine,
    # déterminé par le tableau Présence Vacataire — seule source de vérité) : G1 réduit sur
    # Adulte/MF pour laisser les vacataires remplacer les réguliers du PT
    # (règle "maximiser heures vacataires", cf. règle utilisatrice 08/2026).
    # G1 reste fort sur RDC et Jeunesse (vacataires jamais en RDC ; Jeunesse traité par bonus).
    vacataire_present = any(is_vacataire(a) for a in agents)
    def g1_poids(s):
        if vacataire_present and s in ('Adulte', 'MF'):
            return 30  # réduit → vacataires préférés sur sections secondaires
        return POIDS['G1_planning_type']  # 100 → suit le PT strictement

    # Pour chaque créneau PT, pénaliser si l'agent PT n'est pas à sa place
    # swap_map : si agent PT absent mais remplacé par swap → le remplaçant hérite de la préférence PT
    for c, sections_dict in pt_indexed.items():
        for s, pt_agents in sections_dict.items():
            for a_pt in pt_agents:
                # Résoudre le swap : si a_pt est absent mais swappé → utiliser le remplaçant
                a_effectif = swap_map.get(a_pt, a_pt)

                if a_effectif in agents:
                    # L'agent normalement prévu (ou son remplaçant swap) est PRÉSENT ce
                    # jour-là : le déplacer de sa place n'est jamais anodin → pénalité de
                    # STABILITÉ, résolue en passe 2 (avant G2/J1/J3/I1/équité), pour qu'elle
                    # ne puisse jamais être "battue" par une somme de petites préférences.
                    not_in_pt = model.new_bool_var(f'not_in_pt_{a_effectif}_{c}_{s}')
                    model.add(not_in_pt == 1 - x[a_effectif, c, s])
                    penalites_stabilite.append(g1_poids(s) * not_in_pt)
                elif a_pt in agents:
                    not_in_pt = model.new_bool_var(f'not_in_pt_{a_pt}_{c}_{s}')
                    model.add(not_in_pt == 1 - x[a_pt, c, s])
                    penalites_stabilite.append(g1_poids(s) * not_in_pt)
                else:
                    # Agent PT réellement absent, pas de swap → pénalité si remplacement
                    # par agent de section différente (G2). Ici il ne s'agit plus de
                    # stabilité (personne à sa place n'est déplacée) mais de la QUALITÉ
                    # du remplacement → reste avec G2/J1/J3/I1/équité (passe 3).
                    wrong_sect = []
                    for a in agents:
                        sect_prim = (affectations.get(a) or [''])[0]
                        if sect_prim != s:
                            wrong_sect.append(x[a, c, s])
                    if wrong_sect:
                        v = model.new_bool_var(f'wrong_sect_{c}_{s}')
                        model.add(sum(wrong_sect) >= 1).only_enforce_if(v)
                        model.add(sum(wrong_sect) == 0).only_enforce_if(v.negated())
                        penalites_qualite.append(POIDS['G2_meme_section_repl'] * v)

    # H2 : équité des heures travaillées par rapport au planning-type (règle
    # utilisatrice 08/2026). Quand des agents doivent dépasser leurs heures PT
    # pour remplacer des absents, ce dépassement doit être RÉPARTI entre les
    # agents interchangeables plutôt que concentré sur un seul (ex: si Léa fait
    # +2h de remplacement, essayer que Chloé en fasse aussi +2h plutôt que 0).
    # Les responsables de section sont exclus (traités à part, cf. J3).
    pt_minutes_agent = {}
    for c, sections_dict in pt_indexed.items():
        cs_c, ce_c = creneaux_ouverts[c]
        dur = ce_c - cs_c
        for s, pt_agents in sections_dict.items():
            for a_pt in pt_agents:
                pt_minutes_agent[a_pt] = pt_minutes_agent.get(a_pt, 0) + dur

    agents_equite = [a for a in agents if not is_vacataire(a) and a not in responsables]

    # Minutes d'événements du jour, par agent (règle utilisatrice 08/2026) :
    # un agent occupé par un accueil de classe, une animation ou une réunion
    # est tout autant "chargé" que s'il faisait du service public au comptoir
    # — sans ça, il aurait l'air "disponible" pour un remplacement alors qu'il
    # a déjà donné son heure de travail, juste ailleurs. Équivalence stricte
    # 1h événement = 1h service public, quel que soit le type d'événement
    # (Accueil/Animation/Réunion) — pas de pondération différenciée.
    # ⚠️ L'onglet Événements contient AUSSI les congés/absences/formations
    # (ex: "congé" sur 9h-19h) — ce ne sont PAS des heures travaillées, il ne
    # faut surtout pas les compter ici (sinon un agent absent toute la journée
    # se retrouverait crédité de 10h de "travail équivalent"). On exclut donc
    # les mêmes mots-clés d'absence que ceux utilisés côté génération Excel
    # (colonne J : congé/RTT/vacation/absence/formation).
    evenements_jour = [ev for ev in evenements if ev['date'] == date_str
                        and not _est_evenement_absence(ev['nom'])]
    ev_minutes_agent = {
        a: sum(ev['ce'] - ev['cs'] for ev in evenements_jour if a in ev['agents'])
        for a in agents_equite
    }

    depas_par_agent = {}  # {agent: variable CP-SAT du dépassement NET du jour}
                           # conservé pour (a) l'équité hebdo ci-dessous et
                           # (b) être retourné à compute_full_planning, qui
                           # cumule ces valeurs jour après jour dans la semaine
    if agents_equite:
        depassements_pos = []
        for a in agents_equite:
            travail = sum((creneaux_ouverts[c][1] - creneaux_ouverts[c][0]) * x[a, c, s]
                          for c in range(n_cren) for s in SECTIONS)
            pt_a = pt_minutes_agent.get(a, 0)
            ev_a = ev_minutes_agent.get(a, 0)
            depas = model.new_int_var(-2000, 2000, f'depas_{a}')
            # Dépassement NET = (service public réel + événements du jour) −
            # service public prévu au planning-type. Les événements comptent
            # donc comme du "déjà fait", exactement comme du service public
            # en plus, aussi bien pour la franchise du jour que pour le
            # cumul hebdomadaire (depas_par_agent est repris tel quel plus
            # bas pour les deux niveaux, et retourné à compute_full_planning).
            model.add(depas == travail - pt_a + ev_a)
            depas_par_agent[a] = depas
            depas_pos = model.new_int_var(0, 2000, f'depas_pos_{a}')
            model.add(depas_pos >= depas)
            model.add(depas_pos >= 0)
            depassements_pos.append(depas_pos)
        max_depas = model.new_int_var(0, 2000, 'max_depassement_jour')
        for dp in depassements_pos:
            model.add(max_depas >= dp)
        # Franchise + objectif SECONDAIRE (résolution en 2 passes, voir plus bas) :
        # l'équité ne doit JAMAIS dégrader une solution par ailleurs meilleure sur
        # les priorités structurelles (D_FILL, Jeunesse, G1...). Calculée à part
        # dans `penalites_equite`, optimisée seulement APRÈS avoir fixé la valeur
        # optimale des priorités principales.
        GRACE_EQUITE = 60  # minutes de tolérance avant de pénaliser le déséquilibre (PAR JOUR)
        depas_au_dela_franchise = model.new_int_var(0, 2000, 'depas_au_dela_franchise')
        model.add(depas_au_dela_franchise >= max_depas - GRACE_EQUITE)
        model.add(depas_au_dela_franchise >= 0)
        penalites_equite.append(depas_au_dela_franchise)

        # ── Équité HEBDOMADAIRE (nouveau, 08/2026) ──────────────────────────
        # Même principe que ci-dessus, mais en ajoutant à chaque agent le
        # dépassement DÉJÀ accumulé les jours précédents de la même semaine
        # (cumul_hebdo_avant, fourni par compute_full_planning, remis à zéro
        # au début de chaque semaine). Objectif : éviter qu'un même agent soit
        # choisi comme remplaçant tous les jours de la semaine — sans jamais
        # empêcher un remplacement réellement nécessaire (même mécanisme de
        # 2e passe que ci-dessus : ça ne fait que départager entre choix par
        # ailleurs équivalents pour les priorités structurelles).
        # Franchise plus généreuse (3h) car appréciée sur toute la semaine,
        # cumulée avec la franchise journalière (60 min) — pas à sa place.
        FRANCHISE_HEBDO = 180  # 3h de tolérance sur la semaine
        cumuls_pos_hebdo = []
        for a in agents_equite:
            avant = cumul_hebdo_avant.get(a, 0)
            cumul_total = model.new_int_var(-4000, 4000, f'cumul_hebdo_{a}')
            model.add(cumul_total == depas_par_agent[a] + avant)
            cumul_pos = model.new_int_var(0, 4000, f'cumul_hebdo_pos_{a}')
            model.add(cumul_pos >= cumul_total)
            model.add(cumul_pos >= 0)
            cumuls_pos_hebdo.append(cumul_pos)
        max_cumul_hebdo = model.new_int_var(0, 4000, 'max_cumul_hebdo')
        for cp in cumuls_pos_hebdo:
            model.add(max_cumul_hebdo >= cp)
        depas_hebdo_au_dela_franchise = model.new_int_var(0, 4000, 'depas_hebdo_au_dela_franchise')
        model.add(depas_hebdo_au_dela_franchise >= max_cumul_hebdo - FRANCHISE_HEBDO)
        model.add(depas_hebdo_au_dela_franchise >= 0)
        penalites_equite.append(depas_hebdo_au_dela_franchise)

        # ── MANQUE hebdomadaire (nouveau, règle utilisatrice) ───────────────
        # Symétrique du dépassement ci-dessus : un agent qui reste SOUS son
        # propre planning-type cumulé sur la semaine (cumul_total très négatif)
        # doit aussi être détecté et rattrapé en priorité — pas seulement ceux
        # qui dépassent. Important : `cumul_total` n'existe déjà que pour les
        # jours où l'agent est réellement éligible (agents_eligibles exclut ses
        # propres jours d'absence, cf. compute_full_planning) → un agent absent
        # 2 jours sur 5 n'est PAS pénalisé pour ces 2 jours, seulement comparé
        # à son propre planning-type des jours où il est effectivement présent.
        # Même franchise (3h) que le dépassement, par cohérence.
        # Effet de bord recherché : entre plusieurs agents de section équivalente
        # candidats à un même remplacement, le solveur préfère naturellement
        # celui dont le manque est le plus grand (réduit davantage l'objectif).
        manque_pos_hebdo = []
        for a in agents_equite:
            avant = cumul_hebdo_avant.get(a, 0)
            cumul_total_m = model.new_int_var(-4000, 4000, f'cumul_hebdo_m_{a}')
            model.add(cumul_total_m == depas_par_agent[a] + avant)
            manque = model.new_int_var(0, 4000, f'manque_hebdo_{a}')
            model.add(manque >= -cumul_total_m)
            model.add(manque >= 0)
            manque_pos_hebdo.append(manque)
        max_manque_hebdo = model.new_int_var(0, 4000, 'max_manque_hebdo')
        for mp in manque_pos_hebdo:
            model.add(max_manque_hebdo >= mp)
        # Aucune franchise côté manque (règle utilisatrice) : contrairement au
        # dépassement (toléré jusqu'à 3h avant d'être corrigé), le moindre écart
        # sous le planning-type doit être rattrapé dès qu'une marge de manœuvre
        # existe. Reste toujours un objectif de départage (passe 3) : n'entre en
        # jeu que si ça n'empire pas la couverture des besoins ni la stabilité.
        manque_hebdo_au_dela_franchise = model.new_int_var(0, 4000, 'manque_hebdo_au_dela_franchise')
        model.add(manque_hebdo_au_dela_franchise >= max_manque_hebdo)
        model.add(manque_hebdo_au_dela_franchise >= 0)
        penalites_equite.append(manque_hebdo_au_dela_franchise)

    # J1 : section principale prioritaire
    # Règle vacataires : quand un vacataire est présent (peu importe le jour de la semaine),
    # il n'a PAS de pénalité J1 → il remplit librement les sections laissées par les réguliers.
    # Les réguliers gardent leur préférence de section (J1 actif).
    # → Résultat naturel : réguliers dans leurs sections primaires, vacataires dans le reste.
    for a in agents:
        sects = affectations.get(a, [])
        if not sects:
            continue

        # Vacataires présents → aucune pénalité de section
        if is_vacataire(a) and vacataire_present:
            continue

        cat = categories.get(a)
        sects_equiv = set(sects[:2]) if cat == 'A' else {sects[0]}
        for c in range(n_cren):
            for s in SECTIONS:
                if s in sects_equiv:
                    continue  # section primaire/équivalente → pas de pénalité
                if s not in sects:
                    continue  # pas habilité → géré par A1
                rang = sects.index(s) + 1
                if rang == 2 and cat != 'A':
                    penalites_qualite.append(POIDS['J1_section_principale'] * x[a, c, s])
                else:
                    # Section 3 ou 4 → pénalité forte
                    penalites_qualite.append(70 * x[a, c, s])

                # Départage RDC (08/2026, colonne "Priorité_remplacement_RDC" de
                # l'onglet Affectations — remplace l'ancienne règle codée en dur
                # Adulte>Jeunesse). Nombre plus petit = préféré. Poids
                # volontairement petit (2/point) : ne doit jamais l'emporter sur
                # une vraie différence de rang J1 (30 vs 70), seulement départager
                # entre agents de même rang. Modifiable chaque mois directement
                # dans le fichier Excel, sans toucher au code.
                if s == 'RDC' and a in priorite_rdc:
                    penalites_qualite.append(2 * priorite_rdc[a] * x[a, c, s])

    # J3 : responsables déprioritisés
    for a in responsables:
        if a in agents:
            for c in range(n_cren):
                for s in SECTIONS:
                    penalites_qualite.append(POIDS['J3_responsable'] * x[a, c, s])

    # K1 (redéfinie 08/2026 v2, règle utilisatrice précisée) :
    #   - Vacataire 1 : maximisé (suit le PT / comble un maximum de créneaux),
    #     priorité de remplacement Jeunesse > M&F > Adulte
    #   - Vacataire 2 (et 3+ si présents un jour) : dernier recours UNIQUEMENT
    #     (comble les impossibilités de planning que même les réguliers ne
    #     peuvent pas couvrir). Le choix fin de qui il remplace ensuite se fait
    #     à la main par l'utilisatrice (créneaux les plus longs / journées les
    #     plus chargées à soulager en priorité — trop fin pour être automatisé
    #     pour le moment).
    VAC_BONUS = {'Jeunesse': 90, 'MF': 70, 'Adulte': 50}
    VAC2_DERNIER_RECOURS = 10  # petite pénalité : n'intervient que si nécessaire
    for a in agents:
        if not is_vacataire(a):
            continue
        if a == 'Vacataire 1':
            for c in range(n_cren):
                for s in ('Jeunesse', 'MF', 'Adulte'):
                    penalites_qualite.append(-VAC_BONUS[s] * x[a, c, s])
        else:
            # Vacataire 2, 3... : légère pénalité pour ne s'en servir qu'en
            # dernier recours (D_FILL=5000 et Jeunesse=200 restent prioritaires
            # et forceront quand même son usage si aucun régulier n'est possible)
            for c in range(n_cren):
                for s in SECTIONS:
                    penalites_qualite.append(VAC2_DERNIER_RECOURS * x[a, c, s])

    # I1 : non-fragmentation (pénalité si agent travaille des créneaux non consécutifs)
    for a in agents:
        for c in range(n_cren - 1):
            cs_c,  ce_c  = creneaux_ouverts[c]
            cs_n,  ce_n  = creneaux_ouverts[c + 1]
            if ce_c != cs_n:  # pas consécutifs
                # Pénalité si l'agent travaille c mais pas c+1, et travaille c+2 ou plus
                for c2 in range(c + 2, n_cren):
                    travaille_c  = sum(x[a, c, s]  for s in SECTIONS)
                    travaille_c2 = sum(x[a, c2, s] for s in SECTIONS)
                    gap = model.new_bool_var(f'gap_{a}_{c}_{c2}')
                    model.add(travaille_c  >= 1).only_enforce_if(gap)
                    model.add(travaille_c2 >= 1).only_enforce_if(gap)
                    penalites_qualite.append(POIDS['I1_non_fragmentation'] * gap)

    # Objectif PRINCIPAL : minimiser les pénalités structurelles pures
    # (couverture des besoins : D_FILL, Jeunesse, consécutif — rien d'autre)
    model.minimize(sum(penalites))

    # ══ RÉSOLUTION EN 4 PASSES (mis à jour 08/2026 — ex-3 passes) ══════════
    # Passe 1 : couverture des besoins seule (D_FILL, Jeunesse, consécutif).
    #           Rien ne doit jamais dégrader ça.
    # Passe 2 : à couverture FIXÉE, minimiser les déplacements inutiles
    #           d'agents déjà correctement placés (stabilité — G1 pour un
    #           agent présent).
    # Passe 3 : à couverture + stabilité FIXÉES, choisir le MEILLEUR
    #           remplaçant possible (G2/J1/J3/I1, préférence vacataires) —
    #           SANS regarder les heures de qui que ce soit.
    # Passe 4 (NOUVEAU — corrige régression du 11/08) : à qualité de
    #           remplacement FIXÉE, optimiser en dernier l'équité des heures
    #           (dépassement ET manque). Avant, l'équité était mélangée dans
    #           la même passe que G2/J1/J3 : comme elle se compte en MINUTES
    #           alors que G2/J1/J3 sont des poids fixes par créneau, quelques
    #           dizaines de minutes d'équité pouvaient "racheter" un moins bon
    #           choix de remplaçant — ex. préférer Stéphanie (responsable,
    #           section secondaire) à Macha (section primaire, disponible)
    #           simplement parce que Macha aurait dépassé son quota de
    #           quelques minutes. Désormais l'équité ne peut plus jamais
    #           changer QUI est choisi comme remplaçant — seulement départager
    #           entre choix par ailleurs strictement équivalents en qualité.
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30.0
    solver.parameters.num_search_workers  = 4
    # Graine fixe (08/2026) : sans ça, avec 4 chercheurs en parallèle, le
    # solveur peut trancher différemment entre deux solutions à égalité de
    # score d'un lancement à l'autre — même moteur, mêmes données, résultat
    # parfois différent. Fixer la graine rend le planning reproductible.
    solver.parameters.random_seed = 42
    status = solver.solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return None, [('*', '*', 'aucune solution trouvée (structurellement impossible)')], {}

    if penalites_stabilite:
        valeur_optimale = round(solver.objective_value)
        model.add(sum(penalites) <= valeur_optimale)
        model.minimize(sum(penalites_stabilite))
        solver_stab = cp_model.CpSolver()
        solver_stab.parameters.max_time_in_seconds = 30.0
        solver_stab.parameters.num_search_workers  = 4
        solver_stab.parameters.random_seed = 42
        status_stab = solver_stab.solve(model)
        if status_stab in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            solver = solver_stab  # utiliser la solution la plus stable (même couverture)

    if penalites_qualite:
        if penalites_stabilite:
            valeur_stabilite = sum(solver.value(v) for v in penalites_stabilite)
            model.add(sum(penalites_stabilite) <= valeur_stabilite)
        model.minimize(sum(penalites_qualite))
        solver_qual = cp_model.CpSolver()
        solver_qual.parameters.max_time_in_seconds = 30.0
        solver_qual.parameters.num_search_workers  = 4
        solver_qual.parameters.random_seed = 42
        status_qual = solver_qual.solve(model)
        if status_qual in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            solver = solver_qual  # utiliser la solution avec le meilleur remplaçant

    if penalites_equite:
        if penalites_qualite:
            valeur_qualite = sum(solver.value(v) for v in penalites_qualite)
            model.add(sum(penalites_qualite) <= valeur_qualite)
        model.minimize(sum(penalites_equite))
        solver2 = cp_model.CpSolver()
        solver2.parameters.max_time_in_seconds = 30.0
        solver2.parameters.num_search_workers  = 4
        solver2.parameters.random_seed = 42
        status2 = solver2.solve(model)
        if status2 in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            solver = solver2  # utiliser la solution équilibrée (même couverture + même stabilité + même qualité)

    # ══ EXTRACTION DE LA SOLUTION ═════════════════════════════
    result = {}
    for c in range(n_cren):
        result[c] = {s: [] for s in SECTIONS}
        for a in agents:
            for s in SECTIONS:
                if solver.value(x[a, c, s]) == 1:
                    result[c][s].append(a)

    # ══ VÉRIFICATION POST-RÉSOLUTION DES ALERTES ═══════════════
    # (créneaux où le besoin n'a pas pu être entièrement couvert malgré la pénalité forte)
    for (c, s) in fill_requis:
        if not result[c][s]:
            alertes.append((c, s, 'non pourvu malgré besoin du planning-type'))
    for c, besoin in jeunesse_requis.items():
        nb_reel = len(result[c]['Jeunesse'])
        if nb_reel < besoin:
            alertes.append((c, 'Jeunesse', f'{nb_reel}/{besoin} agent(s) seulement'))

    # Dépassement NET du jour par agent (peut être négatif si l'agent a
    # travaillé MOINS que son PT ce jour-là) — remonté à compute_full_planning
    # qui l'additionne au cumul de la semaine pour le jour suivant.
    depas_jour = {a: solver.value(depas_par_agent[a]) for a in depas_par_agent}

    return result, alertes, depas_jour


# ══════════════════════════════════════════════════════════════
#  POINT D'ENTRÉE PRINCIPAL
# ══════════════════════════════════════════════════════════════



def parse_horaires_ouverture(raw):
    """
    Retourne {jour: [(cs1,ce1), (cs2,ce2)]} — plages d'ouverture au public.
    """
    ws = raw.get('Horaire_ouverture_mediatheque')
    if ws is None:
        return {}
    result = {}
    for row in ws.iter_rows(values_only=True):
        if not row[1] or str(row[1]).strip() in ('Jour', 'jour', ''):
            continue
        jour = str(row[1]).strip().capitalize()
        if jour not in ['Mardi','Mercredi','Jeudi','Vendredi','Samedi']:
            continue
        plages = []
        # S1 : col 2-3, S2 : col 4-5
        for i in range(2, 6, 2):
            cs = hhmm_to_min(row[i]   if len(row) > i   else None)
            ce = hhmm_to_min(row[i+1] if len(row) > i+1 else None)
            if cs is not None and ce is not None and ce > cs:
                plages.append((cs, ce))
        if plages:
            result[jour] = plages
    return result

def compute_full_planning(filepath):
    """
    Calcule le planning complet du mois.
    Retourne (weeks_data, metadata) au même format que l'ancien moteur.
    """
    raw = load_excel_data(filepath)

    params         = parse_parametres(raw)
    affectations, categories, responsables, pause_flex, priorite_rdc = parse_affectations(raw)
    # Lecture directe de la grille collaborative "horaires d'équipes" ; repli sur
    # l'ancienne liste à plat "Horaires_Des_Agents" si le fichier de préparation
    # n'a pas encore été mis à jour avec le nouvel onglet.
    if ONGLET_HORAIRES_GRILLE in raw:
        horaires_agents = parse_horaires_agents_grille(raw)
    else:
        horaires_agents = parse_horaires_agents(raw)
    roulement_type, roulement_exceptions = parse_roulement_samedi(raw)
    besoins_jeunesse = parse_besoins_jeunesse(raw)
    evenements       = parse_evenements(raw, annee_defaut=params.get('annee'))
    planning_type    = parse_planning_type(raw)
    jours_speciaux   = parse_jours_speciaux(raw)

    calendrier = build_calendar(params['mois'], params['annee'], params['samedis'])

    agents_tous = list(affectations.keys())

    # Grille horaire spécifique aux jours "vacances" : construite à partir des
    # tranches du tableau Besoins_Jeunesse (plus fine que la liste standard),
    # car aucune liste de créneaux "vacances" n'existe dans l'onglet Paramètres.
    # ⚠️ Corrigé 08/2026 : construite désormais PAR JOUR (pas une grille unique
    # partagée), et les sous-tranches consécutives avec le MÊME besoin Jeunesse
    # sont fusionnées — évite de fragmenter un bloc homogène (ex: 15h30-17h
    # devenait à tort 15h30-16h + 16h-17h alors que le besoin ne change pas,
    # ce qui laissait le solveur choisir 2 agents différents sur un même bloc).
    def resoudre_besoins_jour(jour_x, samedi_type_x):
        periode_key = next((p for p in besoins_jeunesse if 'Hors' not in p), None)
        if not periode_key:
            return {}
        jours_dict = besoins_jeunesse.get(periode_key, {})
        if jour_x == 'Samedi' and samedi_type_x:
            def _norm(s):
                return s.lower().replace('_', ' ').replace('-', ' ').strip()
            cible = _norm(f'samedi {samedi_type_x}')
            jour_key = next((k for k in jours_dict if _norm(k) == cible),
                             f'Samedi_{samedi_type_x.lower()}')
        else:
            jour_key = jour_x
        return jours_dict.get(jour_key, {})

    def construire_grille_vacances_jour(jour_x, samedi_type_x):
        besoins_jour = resoudre_besoins_jour(jour_x, samedi_type_x)
        ranges = []
        for cren_str, besoin in besoins_jour.items():
            parsed = parse_creneau(cren_str)
            if parsed:
                ranges.append((parsed[0], parsed[1], besoin))
        ranges.sort()
        if not ranges:
            return []
        # Blocs "standards" (hors vacances) pour ce jour : la fusion des sous-tranches
        # ne doit JAMAIS dépasser ces limites, sinon ça fusionne aussi les blocs
        # RDC/Adulte/MF du planning-type qui, eux, gardent leurs propres frontières
        # (ex: fusionner au-delà casserait le remplissage de ces sections).
        blocs_standards = (params['creneaux_ms'] if jour_x in ('Mercredi', 'Samedi')
                            else params['creneaux_mjv'])

        merged = []
        for bs, be in blocs_standards:
            # Sous-tranches Besoins_Jeunesse contenues dans ce bloc standard
            sous = [(cs, ce, b) for (cs, ce, b) in ranges if cs >= bs and ce <= be]
            if not sous:
                merged.append((bs, be))
                continue
            cur = list(sous[0])
            for cs, ce, b in sous[1:]:
                if cs == cur[1] and b == cur[2]:
                    cur[1] = ce  # même besoin ET contigu → fusion (dans ce bloc uniquement)
                else:
                    merged.append((cur[0], cur[1]))
                    cur = [cs, ce, b]
            merged.append((cur[0], cur[1]))
        return merged

    weeks_data = []
    for semaine in calendrier:
        week_num  = semaine['num']
        periode   = params['semaines'].get(week_num, 'Hors Vacances scolaires')
        week_plan = {'week_num': week_num, 'jours': []}

        # Carnet de compte de l'équité HEBDOMADAIRE (08/2026) : {agent: minutes
        # de dépassement net cumulées depuis le début de CETTE semaine}. Remis
        # à zéro à chaque nouvelle semaine (nouvelle itération de cette boucle),
        # mis à jour jour après jour au fur et à mesure qu'on avance dans la
        # semaine — voir §7bis du contexte.
        cumul_hebdo = {}

        for jour_info in semaine['jours']:
            date_str   = jour_info['date']
            jour       = jour_info['jour']
            sam_type   = jour_info.get('samedi_type')

            # Roulement samedi (avec exceptions)
            roulement_agents = dict(roulement_type)
            for agent_exc, roul_exc in roulement_exceptions.get(week_num, {}).items():
                roulement_agents[agent_exc] = roul_exc.upper()

            # Période effective : le réglage par semaine (Semaine_N) sert de défaut,
            # mais un jour marqué "vacances" dans l'onglet Jours_speciaux prime dessus
            # (ex : un pont ponctuel en vacances au sein d'une semaine "Hors Vacances")
            periode_effective = periode
            js_info = jours_speciaux.get(date_str)
            if js_info and js_info.get('vacances'):
                periode_effective = 'Vacances Scolaires'

            # Agents éligibles ce jour
            agents_eligibles = []
            pv = params.get('presences_vac', {})
            use_presences = bool(pv)  # Si tableau défini → utiliser exclusivement
            for a in agents_tous:
                if is_vacataire(a):
                    if use_presences:
                        # Présence explicite uniquement
                        if date_str in pv and a in pv[date_str]:
                            agents_eligibles.append(a)
                    else:
                        # Fallback mode_vac global
                        if jour in params['mode_vac']:
                            agents_eligibles.append(a)
                else:
                    h = horaires_agents.get(a, {}).get(jour)
                    if h and any(v is not None for v in h):
                        agents_eligibles.append(a)

            # Planning type pour ce jour
            if jour == 'Samedi' and sam_type:
                pt_jour_key = f'Samedi_{sam_type}'
            else:
                pt_jour_key = jour
            pt_jour = planning_type.get(pt_jour_key, {})

            # Créneaux ouverts : grille "vacances" (fine, fusionnée, propre à ce
            # jour) si le jour est en mode vacances (via Jours_speciaux ou
            # Semaine_N), sinon liste standard selon le jour
            creneaux_vacances_jour = (construire_grille_vacances_jour(jour, sam_type)
                                       if 'Hors' not in periode_effective else [])
            if creneaux_vacances_jour:
                creneaux_ouverts = creneaux_vacances_jour
            elif jour in ('Mercredi', 'Samedi'):
                creneaux_ouverts = params['creneaux_ms']
            else:
                creneaux_ouverts = params['creneaux_mjv']

            # Résolution CP-SAT
            # Construire le swap_map pour ce samedi
            # Si agent A normalement ROUGE est passé BLEU (exception) et B est passé ROUGE
            # → B remplace A dans les slots PT de A
            swap_map = {}
            if jour == 'Samedi' and sam_type:
                exc = roulement_exceptions.get(week_num, {})
                # Agents qui ont changé de roulement ce samedi
                vers_autre = {a: r for a, r in exc.items() if r != sam_type}  # absents
                vers_ce_sam = {a: r for a, r in exc.items() if r == sam_type}  # présents par exception
                # Aussi les agents dont le type normal diffère du samedi actuel
                # absents_normal = ceux qui sont normalement de l'autre couleur MAIS
                # ont été swappés vers ce samedi
                normal_absents = [a for a, r in roulement_type.items()
                                  if r != sam_type and a not in exc]
                # Pour chaque absent normal qui a un swap entrant → construire le map
                # Heuristique : chercher dans vers_autre si l'agent PT ROUGE est absent
                for a_absent, r_absent in vers_autre.items():
                    # Trouver qui l'a remplacé (celui qui est passé vers ce samedi)
                    for a_repl, r_repl in vers_ce_sam.items():
                        if a_absent not in swap_map:
                            swap_map[a_absent] = a_repl

            solution, alertes, depas_jour = solve_day(
                jour=jour,
                date_str=date_str,
                creneaux_ouverts=creneaux_ouverts,
                agents_eligibles=agents_eligibles,
                affectations=affectations,
                categories=categories,
                responsables=responsables,
                pause_flex=pause_flex,
                priorite_rdc=priorite_rdc,
                horaires_agents=horaires_agents,
                evenements=evenements,
                besoins_jeunesse=besoins_jeunesse,
                planning_type_jour=pt_jour,
                roulement_agents=roulement_agents,
                samedi_type=sam_type,
                periode=periode_effective,
                mode_vac=params['mode_vac'],
                swap_map=swap_map,
                presences_vac=params.get('presences_vac', {}),
                cumul_hebdo_avant=cumul_hebdo,
            )

            # Mise à jour du carnet hebdo : on ajoute le dépassement NET de ce
            # jour à ce qui était déjà cumulé cette semaine, pour que le jour
            # suivant en tienne compte.
            for a, d in depas_jour.items():
                cumul_hebdo[a] = cumul_hebdo.get(a, 0) + d

            week_plan['jours'].append({
                'date':      date_str,
                'jour':      jour,
                'sam_type':  sam_type,
                'creneaux':  creneaux_ouverts,
                'solution':  solution,   # {cren_idx: {section: [agents]}}
                'infaisable': solution is None,
                'alertes':   alertes,    # [(cren_idx, section, message)]
                'cumul_hebdo_apres': dict(cumul_hebdo),  # utile pour debug/traçabilité
            })

        weeks_data.append(week_plan)

    metadata = {
        'mois':       params['mois'],
        'annee':      params['annee'],
        'evenements': evenements,
    }

    return weeks_data, metadata
